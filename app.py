from datetime import datetime, timedelta
import json
import logging
import uuid
import time
import threading
import unicodedata
from flask import Flask, flash, redirect, render_template, request, jsonify, send_file, session, url_for
from flask_bcrypt import Bcrypt
from flask_mail import Mail, Message
from flask_session import Session
from werkzeug.utils import secure_filename
import pandas as pd
import pyodbc, random
import os
from dotenv import load_dotenv
from connection import connect
from github_service import create_issue, close_issue, get_user_github_usernames, get_project_github_repo

# Carregar variáveis de ambiente do ficheiro .env (desenvolvimento)
load_dotenv()


def parse_responsible(resp_str):
    """Converte string de responsáveis (separada por ;) em lista.
    Handling de backward compatibility: se for string simples (sem ;), retorna [string]."""
    if not resp_str or resp_str == 'Unassigned':
        return []
    if isinstance(resp_str, list):
        return resp_str
    # Se já contém ponto e vírgula, dividir
    if ';' in resp_str:
        return [r.strip() for r in resp_str.split(';') if r.strip()]
    # Senão, é um único responsável (backward compatible)
    return [resp_str]


def format_responsible(resp_list):
    """Converte lista de responsáveis em string separada por ;."""
    if not resp_list:
        return 'Unassigned'
    if isinstance(resp_list, str):
        # Já é string, retornar conforme está
        return resp_list if resp_list else 'Unassigned'
    # Filtrar vazios e juntar com ;
    return ';'.join([r.strip() for r in resp_list if r and r.strip()])


try:
    print("Connection successful")
except Exception as e:
    print("Connection failed new APP")

app = Flask(__name__)

# Configure logging
logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Sessão persistente
app.config["SESSION_PERMANENT"] = True
app.permanent_session_lifetime = timedelta(days=30)

# Sessão no servidor (filesystem)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_COOKIE_NAME'] = 'session_rms'  # 🔹 nome exclusivo do cookie

Session(app)

app.config['MAIL_SERVER'] = 'viasmtp.borgwarner.net'
app.config['MAIL_PORT'] = 25
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_DEFAULT_SENDER'] = 'RMS@borgwarner.com'
app.config['UPLOAD_FOLDER'] = 'Uploads'  # Define the upload folder
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Limite de 16MB para uploads
# Remover restrições de extensão - permitir qualquer tipo de arquivo
ALLOWED_EXTENSIONS = set()  # Empty set means all extensions allowed

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

mail = Mail(app)

@app.before_request
def refresh_session():
    if 'username' in session:
        session.permanent = True
        session.modified = True

def send_ticket_status_email(
    to_email, to_name, ticket_code, ticket_title, new_status, extra_msg='',
    responsible=None, expected_date=None, notes=None, cc_emails=None
):
    subject = f"RMS - Status Update for Ticket {ticket_code}"

    # Cores institucionais
    colors = {
        'primary': '#051729',
        'primary_light': '#113561',
        'accent': '#2EFAD9',
        'success': '#28a745',
        'danger': '#dc3545',
        'warning': '#ffc107',
        'light_bg': '#f8f9fc'
    }

    status_msg = {
        'opened': f'has been <b style="color:{colors["success"]};">opened</b> and is now in our system.',
        'under_analysis': f'is now <b style="color:{colors["warning"]};">under analysis</b> by our team.',
        'approved': f'has been <b style="color:{colors["success"]};">approved</b> and is now in our processing queue.',
        'rejected': f'has been <b style="color:{colors["danger"]};">declined</b> at this time.',
        'in_progress': f'is now <b style="color:{colors["primary_light"]};">actively being worked on</b>.',
        'completed': f'has been <b style="color:{colors["success"]};">successfully completed</b>!',
        'waiting_dt': f'is now <b style="color:{colors["primary_light"]};">waiting for DT team input</b>.',
        'waiting_requester': f'is now <b style="color:{colors["warning"]};">waiting for your response</b>.',
        'waiting_line': f'is now <b style="color:{colors["primary_light"]};">Waiting for Line Availability</b>.',
        'waiting_maintenance': f'is now <b style="color:{colors["primary_light"]};">waiting for maintenance action</b>.'
    }

    action_items = {
        'opened': 'Your ticket has been received and is pending assignment. We will notify you when it is processed.',
        'under_analysis': 'Our team is currently evaluating your request. You will be notified once the analysis is complete and work is scheduled.',
        'approved': 'Our team will begin working on your request. You will receive updates as progress is made.',
        'rejected': 'Please review the feedback provided and feel free to submit a revised request if applicable.',
        'in_progress': 'Our team is actively working on your request. We will notify you once the work is completed.',
        'completed': 'Your request is now ready for use. Thank you for using our services.',
        'waiting_dt': 'The DT team is evaluating next steps. You will be notified of any updates.',
        'waiting_requester': 'Please provide the requested information or confirmation to proceed with your request.',
        'waiting_line': 'Your request is queued and awaiting availability. We appreciate your patience.',
        'waiting_maintenance': 'Your request is awaiting maintenance window scheduling. Thank you for your cooperation.'
    }

    header_colors = {
        'opened': '#e8f5e8',
        'under_analysis': '#fff8e6',
        'approved': '#e8f5e8',
        'rejected': '#f8e8e8',
        'in_progress': '#e8f0ff',
        'completed': '#e8f5e8',
        'waiting_dt': '#e8f0ff',
        'waiting_requester': '#fff8e6',
        'waiting_line': '#e8f0ff',
        'waiting_maintenance': '#e8f0ff'
    }
    header_color = header_colors.get(new_status, colors['light_bg'])

    # Bloco para responsável e data prevista
    responsible_block = ""
    if new_status in ['under_analysis', 'approved', 'in_progress', 'waiting_dt', 'waiting_requester', 'waiting_line', 'waiting_maintenance']:
        notes_line = f"<br><b style=\"color:{colors['primary_light']};\">Notes:</b> {notes}" if notes and notes.strip() else ""
        
        if new_status == 'under_analysis':
            responsible_info = f"""
                <div style="margin: 12px 0 0 0; font-size: 0.97rem;">
                    <b style="color:{colors['primary_light']};">Responsible:</b> {responsible or '<i>Not assigned</i>'}{notes_line}
                </div>
            """
        elif new_status in ['waiting_dt', 'waiting_requester', 'waiting_line', 'waiting_maintenance']:
            responsible_info = f"""
                <div style="margin: 12px 0 0 0; font-size: 0.97rem;">
                    <b style="color:{colors['primary_light']};">Responsible:</b> {responsible or '<i>Not assigned</i>'}{notes_line}
                </div>
            """
        else:
            responsible_info = f"""
                <div style="margin: 12px 0 0 0; font-size: 0.97rem;">
                    <b style="color:{colors['primary_light']};">Responsible:</b> {responsible or '<i>Not assigned</i>'}<br>
                    <b style="color:{colors['primary_light']};">Expected Date:</b> {expected_date or '<i>Not defined</i>'}{notes_line}
                </div>
            """
        responsible_block = responsible_info

    body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>RMS Status Update</title>
        <link href="https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600&display=swap" rel="stylesheet">
    </head>
    <body style="margin:0; padding:0; background:{colors['light_bg']}; font-family:'Montserrat', Arial, sans-serif;">
        <table width="100%" cellpadding="0" cellspacing="0" style="background:{colors['light_bg']}; padding:0;">
            <tr>
                <td align="center">
                    <table width="600" cellpadding="0" cellspacing="0" style="background:#fff; border-radius:8px; margin:32px 0; border:1px solid #e3e6f0;">
                        <tr>
                            <td style="padding:24px 32px 8px 32px;">
                                <h2 style="color:{colors['primary_light']}; font-weight:700; font-size:1.5rem; margin:0 0 8px 0; letter-spacing:0.5px;">
                                    RMS
                                </h2>
                                <p style="color:{colors['primary']}; font-size:1rem; margin:0 0 0 0;">
                                    Digital Transformation Team
                                </p>
                            </td>
                        </tr>
                        <tr>
                            <td style="background:{header_color}; padding:16px 32px;">
                                <h3 style="margin:0; color:{colors['primary_light']}; font-size:1.1rem; font-weight:600;">
                                    Ticket Status Update
                                </h3>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding:24px 32px;">
                                <p style="color:{colors['primary']}; font-size:1rem; margin:0 0 12px 0;">
                                    Hello <b>{to_name}</b>,
                                </p>
                                <p style="color:#333; font-size:0.98rem; margin:0 0 18px 0;">
                                    We wanted to update you on the status of your request:
                                </p>
                                <table width="100%" cellpadding="0" cellspacing="0" style="background:{colors['light_bg']}; border-left:4px solid {colors['accent']}; border-radius:4px; margin-bottom:18px;">
                                    <tr>
                                        <td style="padding:12px 16px;">
                                            <span style="font-weight:600; color:{colors['primary_light']};">{ticket_code}</span>
                                            <span style="color:#555;">- {ticket_title}</span>
                                            <div style="color:#444; margin-top:6px; font-size:0.97rem;">
                                                Your request {status_msg.get(new_status, f'status has been updated to {new_status.capitalize()}.')}
                                            </div>
                                            {responsible_block}
                                        </td>
                                    </tr>
                                </table>
                                <div style="background:#fff; border:1px solid {colors['accent']}; border-radius:4px; padding:14px 16px; margin-bottom:18px;">
                                    <b style="color:{colors['primary']}; font-size:0.97rem;">Next Steps</b>
                                    <div style="color:#555; font-size:0.96rem; margin-top:4px;">
                                        {action_items.get(new_status, 'We will keep you informed of any further updates.')}
                                    </div>
                                </div>
                                {f'''
                                <div style="background-color:{colors['warning']}22; border:1px solid {colors['warning']}66; border-radius:4px; padding:12px 16px; margin-bottom:18px;">
                                    <b style="color:{colors['warning']}; font-size:0.96rem;">Additional Information</b>
                                    <div style="color:{colors['primary_light']}; font-size:0.95rem; margin-top:4px;">
                                        {extra_msg}
                                    </div>
                                </div>
                                ''' if extra_msg else ''}
                                <div style="color:#888; font-size:0.93rem; margin-top:24px;">
                                    If you have any questions, please contact the Digital Transformation team.
                                </div>
                            </td>
                        </tr>
                        <tr>
                            <td style="background:{colors['light_bg']}; color:#6c757d; font-size:0.92rem; text-align:center; padding:16px 32px; border-top:1px solid #e3e6f0;">
                                This is an automated notification from RMS.<br>
                                Please do not reply to this email.<br>
                                <span style="color:#adb5bd; font-size:0.85rem;">© 2025 Digital Transformation Team. All rights reserved.</span>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

    try:
        # Preparar destinatários: apenas o requisitante no 'To'
        recipients = [to_email]
        cc_list = []
        if cc_emails:
            # cc_emails pode ser string ou lista
            if isinstance(cc_emails, str):
                cc_list = [email.strip() for email in cc_emails.split(',') if email.strip()]
            elif isinstance(cc_emails, list):
                cc_list = [email.strip() for email in cc_emails if email and email.strip()]

        # Criar mensagem com CC separado
        msg = Message(subject, recipients=recipients, cc=cc_list if cc_list else None, html=body)
        mail.send(msg)

        # Log detalhado
        print(f"[EMAIL SUCCESS] Status email sent to {to_email} for ticket {ticket_code}")
        if cc_list:
            print(f"[EMAIL CC] Also sent to CC: {', '.join(cc_list)}")
    except Exception as e:
        print(f"[EMAIL ERROR] Could not send email to {to_email}: {e}")
        print(f"[EMAIL ERROR] Ticket: {ticket_code}, Status: {new_status}, Recipient: {to_name}")

def allowed_file(filename):
    if not filename or '.' not in filename:
        return False

    # Lista de extensões perigosas que não são permitidas por segurança
    dangerous_extensions = {
        'exe', 'bat', 'cmd', 'com', 'pif', 'scr', 'vbs', 'js', 'jar',
        'msi', 'dmg', 'pkg', 'deb', 'rpm', 'sh', 'ps1', 'php'
    }

    extension = filename.rsplit('.', 1)[1].lower()
    return extension not in dangerous_extensions

def get_emails_from_cc_input(cc_input):
    """
    Converte input de CC (username ou email) em lista de emails.
    O input pode ser uma string com vírgulas ou uma lista.
    """
    if not cc_input:
        return []

    # Se for lista, usamos; se for string, split
    if isinstance(cc_input, str):
        items = [item.strip() for item in cc_input.split(',') if item.strip()]
    elif isinstance(cc_input, list):
        items = cc_input
    else:
        return []

    # Separar usernames de emails: se contém '@' assumimos que é email
    emails = []
    usernames_to_lookup = []

    for item in items:
        if '@' in item:
            emails.append(item)
        else:
            usernames_to_lookup.append(item)

    # Buscar emails dos usernames na base de dados
    if usernames_to_lookup:
        conn = connect()
        cursor = conn.cursor()
        placeholders = ','.join(['?'] * len(usernames_to_lookup))
        query = f"SELECT username, email FROM users WHERE username IN ({placeholders})"
        cursor.execute(query, usernames_to_lookup)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        username_to_email = {row[0]: row[1] for row in rows}

        for username in usernames_to_lookup:
            if username in username_to_email:
                emails.append(username_to_email[username])
            else:
                # Se não encontrar, usamos o próprio texto (permite inserção manual de email)
                emails.append(username)

    return emails

def row_to_dict(cursor, row):
    columns = [column[0] for column in cursor.description]
    return dict(zip(columns, row))

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/home')
def home():
    try:
        username = session.get('username')  # Check username instead of name
        if not username:
            flash('Please log in to view your tickets.', category='error')
            return redirect(url_for('index'))

        name = session.get('name')  # Still get name for display
        role = session.get('role')  # Get user role

        conn = connect()
        cursor = conn.cursor()

        # Fetch production lines
        cursor.execute('Exec GetAllProdLines')
        lines = cursor.fetchall()

        def row_to_dict(cursor, row):
            columns = [column[0] for column in cursor.description]
            return dict(zip(columns, row))

        # DTAI
        cursor.execute("""
            SELECT ai.id, ai.internal_code, ai.title, ai.requester, u1.name as requester_name,
                   ai.prod_line, ai.n_sap, ai.reason, ai.current_process, ai.action_to_improve,
                   ai.filename, ai.expected_date, ai.responsible, u2.name as responsible_name,
                   ai.status, ai.observations, ai.created_at, ai.updated_at, ai.notes, ai.requester_response
            FROM automation_improvement ai
            LEFT JOIN users u1 ON ai.requester = u1.name
            LEFT JOIN users u2 ON ai.responsible = u2.username
            WHERE ai.requester = ? AND ai.is_deleted = 0
        """, (name,))
        dtai_tickets = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        # DTAS
        cursor.execute("""
            SELECT asu.id, asu.internal_code, asu.title, asu.requester, u1.name as requester_name,
                   asu.prod_line, asu.n_sap, asu.reason, asu.current_process, asu.action_to_improve,
                   asu.observations_requester, asu.observations_dt,
                   asu.filename, asu.expected_date, asu.responsible, u2.name as responsible_name,
                   asu.status, asu.created_at, asu.updated_at, asu.notes, asu.requester_response
            FROM automation_support asu
            LEFT JOIN users u1 ON asu.requester = u1.name
            LEFT JOIN users u2 ON asu.responsible = u2.username
            WHERE asu.requester = ? AND asu.is_deleted = 0
        """, (name,))
        dtas_tickets = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        # DTNA
        cursor.execute("""
            SELECT na.id, na.internal_code, na.title, na.requester, u1.name as requester_name,
                   na.department, na.reason, na.current_process, na.objective,
                   na.filename, na.expected_date, na.responsible, u2.name as responsible_name,
                   na.status, na.observations, na.created_at, na.updated_at, na.notes, na.requester_response
            FROM new_application na
            LEFT JOIN users u1 ON na.requester = u1.name
            LEFT JOIN users u2 ON na.responsible = u2.username
            WHERE na.requester = ? AND na.is_deleted = 0
        """, (name,))
        dtna_tickets = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        # DTSI
        cursor.execute("""
            SELECT si.id, si.internal_code, si.title, si.requester, u1.name as requester_name,
                   si.app_name, si.reason, si.issue,
                   si.filename, si.expected_date, si.responsible, u2.name as responsible_name,
                   si.status, si.observations, si.created_at, si.updated_at, si.notes, si.requester_response
            FROM software_issue si
            LEFT JOIN users u1 ON si.requester = u1.name
            LEFT JOIN users u2 ON si.responsible = u2.username
            WHERE si.requester = ? AND si.is_deleted = 0
        """, (name,))
        dtsi_tickets = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        # DTIR (Software Internal Reports) - Only visible to Software and Admin category users
        dtir_tickets = []
        user_category = (session.get('category') or '').lower()
        
        # Security: Only Software and Admin categories can access DTIR
        allowed_dtir = user_category in ['software', 'admin']
        
        if allowed_dtir:
            try:
                cursor.execute("""
                    SELECT sir.id, sir.internal_code, sir.title, sir.requester,
                           sir.reporter, sir.reporter as reporter_name,
                           sir.description, sir.filename, sir.status, sir.created_at, sir.updated_at,
                           sir.requester_response, sir.notes, sir.expected_date, sir.responsible
                    FROM software_internal_reports sir
                    WHERE sir.requester = ? AND sir.is_deleted = 0
                """, (name,))
                dtir_tickets = [row_to_dict(cursor, row) for row in cursor.fetchall()]
            except Exception as e:
                logging.error(f"Error fetching DTIR tickets for {username}: {str(e)}")
                dtir_tickets = []
        else:
            if username and user_category not in ['software', 'admin']:
                logging.warning(f"DTIR access denied for {username} with category {user_category}")
            dtir_tickets = []

        cursor.close()
        conn.close()

        return render_template('homerequests.html',
                             lines=lines,
                             dtai_tickets=dtai_tickets,
                             dtas_tickets=dtas_tickets,
                             dtna_tickets=dtna_tickets,
                             dtsi_tickets=dtsi_tickets,
                             dtir_tickets=dtir_tickets,
                             user_role=role,
                             user_category=user_category)
    except Exception as e:
        flash(f'Error fetching tickets: {str(e)}', category='error')
        return redirect(url_for('index'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    conn = connect()
    cursor = conn.cursor()
    if request.method == 'POST' and 'username' in request.form and 'password' in request.form:
        username = request.form['username']
        password = request.form['password']
        cursor.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password))
        account = cursor.fetchone()
        if account:
            session.permanent = True
            session['id'] = account[0]
            session['username'] = account[1]
            session['email'] = account[3]
            session['role'] = account[4]
            session['auth_method'] = 'manual'
            session['name'] = account[6] if account[6] else account[1]  # Use username if name is None
            session['category'] = account[5] if len(account) > 5 and account[5] else ''  # Use empty string if category is NULL
            if session['role'] == 0:
                return redirect(url_for('home'))
            else:
                return redirect(url_for('homeadmin'))
        else:
            flash('Invalid credentials. Please check your username and password.', category='error')
    return redirect(url_for('index'))

@app.route('/windows_login')
def windows_login():
    user = request.environ.get('REMOTE_USER')
    if not user:
        flash("Windows authentication failed or is not active.", category="error")
        return redirect(url_for('index'))
    
    username = user.split("\\")[-1] if "\\" in user else user
    
    try:
        conn = connect()
        if not conn:
            flash('Erro ao conectar à base de dados', category='error')
            return redirect(url_for('index'))
        
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
        account = cursor.fetchone()
        conn.close()
        
        if not account:
            flash('User not found in the database.', category='error')
            return redirect(url_for('index'))
        
        session.permanent = True
        session['id'] = account[0]
        session['username'] = account[1]
        session['email'] = account[3]
        session['role'] = account[4]
        session['auth_method'] = 'windows'
        session['name'] = account[6] if account[6] else account[1]  # Use username if name is None
        # Category é a coluna 6 (index 5) na tabela users
        session['category'] = account[5] if len(account) > 5 and account[5] else ''  # Use empty string if category is NULL
        
        
        flash(f'Windows login successful! Welcome, {username}', category='success')
        
        if session['role'] == 0:
            return redirect(url_for('home'))
        else:
            return redirect(url_for('homeadmin'))
    except Exception as e:
        flash(f'Error in Windows authentication: {str(e)}', category='error')
        return redirect(url_for('index'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    conn = connect()
    cursor = conn.cursor()
    if request.method == 'POST':
        username = request.form['username']
        name = request.form['name'] 
        email = username + '@borgwarner.com'
        password1 = request.form['password1']
        password2 = request.form['password2']
        if password1 != password2:
            flash('Passwords do not match', category='warning')
        else:
            cursor.execute("SELECT email FROM users WHERE email = ?", (email,))
            account_validation = cursor.fetchall()
            if account_validation:
                flash('This email is already registered, contact the system administrator', category='error')
                return redirect(url_for('index'))
            else:
                cursor.execute(
                    "INSERT INTO users (username, password, email, name) VALUES (?, ?, ?, ?)",
                    (username, password1, email, name)
                )
                conn.commit()
                cursor.close()
                conn.close()
                flash('Account created successfully!', category='success')
                return redirect(url_for('index'))
    return redirect(url_for('index'))

@app.route('/logout')
def logout():
    username = session.get('username', 'unknown')
    auth_method = session.get('auth_method', 'manual')
    session.clear()
    return redirect(url_for('index'))

@app.route('/homeadmin')
def homeadmin():
    return redirect(url_for('new_tickets'))

@app.route('/addAutomationImprovement', methods=['GET', 'POST'])
def addAutomationImprovement():
    try:
        if request.method == 'POST':
            requester = session['name']
            title = request.form['title']
            prod_line = request.form['prod_line']
            equipament_SAP = request.form['equipamentSAP']
            reason = request.form['reason']
            subcategory = request.form.get('subcategory', '')
            current_process = request.form['current_process']
            improvement = request.form['improvement']

            # New conditional fields
            eight_d_number = request.form.get('8D_Number', '')
            complaint_number = request.form.get('complaint_number', '')
            hard_savings = request.form.get('hard_savings', None)
            other_description = request.form.get('other_description', '')

            # Obter valores de CC (suporta tanto select múltiplo quanto string legada)
            cc_values = request.form.getlist('cc_emails')
            if not cc_values:
                cc_values = request.form.get('cc_emails', '')
            # Se for string, converter para lista
            if isinstance(cc_values, str):
                cc_list = [v.strip() for v in cc_values.split(',') if v.strip()]
            else:
                cc_list = cc_values
            # String para armazenar no banco (valores originais separados por vírgula)
            cc_emails = ', '.join(cc_list) if cc_list else ''
            # Lista de emails para envio (converte usernames)
            cc_emails_list = get_emails_from_cc_input(cc_list)

            # Convert hard_savings to decimal if provided
            if hard_savings and hard_savings.strip():
                try:
                    hard_savings = float(hard_savings)
                except ValueError:
                    hard_savings = None
            else:
                hard_savings = None

            current_year = datetime.now().year

            # Handle file upload
            filename = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename != '':
                    if allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        file.save(file_path)
                    else:
                        return jsonify({
                            'success': False, 
                            'message': f'Invalid file type: {file.filename}. Dangerous file types are not allowed for security reasons.',
                            'keep_modal_open': True
                        })

            conn = connect()
            cursor = conn.cursor()
            cursor.execute("SELECT TOP 1 internal_code FROM automation_Improvement ORDER BY id DESC")
            last_code = cursor.fetchone()

            if last_code:
                last_number = int(last_code[0].split('-')[-1])
                new_number = last_number + 1
            else:
                new_number = 1
            new_code = f"DTAI-{current_year}-{new_number:03d}"

            cursor.execute('''
                INSERT INTO automation_improvement
                (internal_code, title, requester, prod_line, n_sap, reason, subcategory, current_process, action_to_improve,
                [8D_Number], complaint_number, hard_savings, other_description, cc_emails, filename, created_at, updated_at)
                OUTPUT INSERTED.ID
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE())
            ''', (new_code, title, requester, prod_line, equipament_SAP, reason, subcategory, current_process, improvement,
                  eight_d_number, complaint_number, hard_savings, other_description, cc_emails, filename))
            id = cursor.fetchone()[0]
            conn.commit()

            # Enviar email de abertura para o requester com cópia para os CCs
            try:
                cursor.execute("SELECT email, name FROM users WHERE name = ?", (requester,))
                user_row = cursor.fetchone()
                if user_row:
                    to_email, to_name = user_row
                    send_ticket_status_email(
                        to_email, to_name, new_code, title, 'opened',
                        cc_emails=cc_emails_list
                    )
            except Exception as e:
                print(f"[EMAIL ERROR] Failed to send creation email for ticket {new_code}: {str(e)}")

            cursor.close()
            conn.close()


            return jsonify({
                'success': True,
                'message': 'Ticket Created Successfully!',
                'ticket_code': new_code,
                'keep_modal_open': False
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f"ERROR: {str(e)}",
            'keep_modal_open': True
        })

@app.route('/addAutomationSupport', methods=['GET', 'POST'])
def addAutomationSupport():
    try:
        if request.method == 'POST':
            requester = session['name']
            title = request.form['title']
            prod_line = request.form['prod_line']
            equipament_SAP = request.form['equipamentSAP']
            reason = request.form['reason']
            subcategory = request.form.get('subcategory', '')
            current_process = request.form['current_process']
            corrective_process = request.form['corrective_process']
            machine_operational_state = request.form.get('machine_operational_state', '')
            observations = request.form['observations']

            # New conditional fields
            eight_d_number = request.form.get('8D_Number', '')
            complaint_number = request.form.get('complaint_number', '')
            hard_savings = request.form.get('hard_savings', None)
            other_description = request.form.get('other_description', '')

            # Obter valores de CC (suporta tanto select múltiplo quanto string legada)
            cc_values = request.form.getlist('cc_emails')
            if not cc_values:
                cc_values = request.form.get('cc_emails', '')
            # Se for string, converter para lista
            if isinstance(cc_values, str):
                cc_list = [v.strip() for v in cc_values.split(',') if v.strip()]
            else:
                cc_list = cc_values
            # String para armazenar no banco (valores originais separados por vírgula)
            cc_emails = ', '.join(cc_list) if cc_list else ''
            # Lista de emails para envio (converte usernames)
            cc_emails_list = get_emails_from_cc_input(cc_list)

            # Convert hard_savings to decimal if provided
            if hard_savings and hard_savings.strip():
                try:
                    hard_savings = float(hard_savings)
                except ValueError:
                    hard_savings = None
            else:
                hard_savings = None
                
            current_year = datetime.now().year

            # Handle file upload
            filename = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename != '':
                    if allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        file.save(file_path)
                    else:
                        return jsonify({
                            'success': False, 
                            'message': f'Invalid file type: {file.filename}. Dangerous file types are not allowed for security reasons.',
                            'keep_modal_open': True
                        })

            conn = connect()
            cursor = conn.cursor()
            cursor.execute("SELECT TOP 1 internal_code FROM automation_support ORDER BY id DESC")
            last_code = cursor.fetchone()

            if last_code:
                last_number = int(last_code[0].split('-')[-1])
                new_number = last_number + 1
            else:
                new_number = 1
            new_code = f"DTAS-{current_year}-{new_number:03d}"

            cursor.execute('''
                INSERT INTO automation_support
                (internal_code, title, requester, prod_line, n_sap, reason, subcategory, current_process, action_to_improve,
                observations_requester, [8D_Number], complaint_number, hard_savings, other_description, cc_emails, machine_operational_state, filename, created_at, updated_at)
                OUTPUT INSERTED.ID
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE())
            ''', (new_code, title, requester, prod_line, equipament_SAP, reason, subcategory, current_process, corrective_process,
                  observations, eight_d_number, complaint_number, hard_savings, other_description, cc_emails, machine_operational_state, filename))
            id = cursor.fetchone()[0]
            conn.commit()

            # Enviar email de abertura para o requester com cópia para os CCs
            try:
                cursor.execute("SELECT email, name FROM users WHERE name = ?", (requester,))
                user_row = cursor.fetchone()
                if user_row:
                    to_email, to_name = user_row
                    send_ticket_status_email(
                        to_email, to_name, new_code, title, 'opened',
                        cc_emails=cc_emails_list
                    )
            except Exception as e:
                print(f"[EMAIL ERROR] Failed to send creation email for ticket {new_code}: {str(e)}")

            cursor.close()
            conn.close()

            return jsonify({
                'success': True,
                'message': 'Ticket Created Successfully!',
                'ticket_code': new_code,
                'keep_modal_open': False
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f"ERROR: {str(e)}",
            'keep_modal_open': True
        })

@app.route('/addNewApplication', methods=['GET', 'POST'])
def addNewApplication():
    try:
        if request.method == 'POST':
            requester = session['name']
            title = request.form['title']
            department = request.form['department']
            reason = request.form['reason']
            current_process = request.form['current_process']
            objective = request.form['objective']  # Campo correto do formulário

            # Obter valores de CC (suporta tanto select múltiplo quanto string legada)
            cc_values = request.form.getlist('cc_emails')
            if not cc_values:
                cc_values = request.form.get('cc_emails', '')
            # Se for string, converter para lista
            if isinstance(cc_values, str):
                cc_list = [v.strip() for v in cc_values.split(',') if v.strip()]
            else:
                cc_list = cc_values
            # String para armazenar no banco (valores originais separados por vírgula)
            cc_emails = ', '.join(cc_list) if cc_list else ''
            # Lista de emails para envio (converte usernames)
            cc_emails_list = get_emails_from_cc_input(cc_list)

            current_year = datetime.now().year

            # Handle file upload
            filename = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename != '':
                    if allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        file.save(file_path)
                    else:
                        return jsonify({
                            'success': False, 
                            'message': f'Invalid file type: {file.filename}. Dangerous file types are not allowed for security reasons.',
                            'keep_modal_open': True
                        })

            conn = connect()
            cursor = conn.cursor()
            cursor.execute("SELECT TOP 1 internal_code FROM new_application ORDER BY id DESC")
            last_code = cursor.fetchone()

            if last_code:
                last_number = int(last_code[0].split('-')[-1])
                new_number = last_number + 1
            else:
                new_number = 1
            new_code = f"DTNA-{current_year}-{new_number:03d}"

            cursor.execute('''
                INSERT INTO new_application
                (internal_code, title, requester, department, reason, current_process, objective, cc_emails, filename, created_at, updated_at)
                OUTPUT INSERTED.ID
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE())
            ''', (new_code, title, requester, department, reason, current_process, objective, cc_emails, filename))
            id = cursor.fetchone()[0]
            conn.commit()

            # Enviar email de abertura para o requester com cópia para os CCs
            try:
                cursor.execute("SELECT email, name FROM users WHERE name = ?", (requester,))
                user_row = cursor.fetchone()
                if user_row:
                    to_email, to_name = user_row
                    send_ticket_status_email(
                        to_email, to_name, new_code, title, 'opened',
                        cc_emails=cc_emails_list
                    )
            except Exception as e:
                print(f"[EMAIL ERROR] Failed to send creation email for ticket {new_code}: {str(e)}")

            cursor.close()
            conn.close()

            return jsonify({
                'success': True,
                'message': 'Ticket Created Successfully!',
                'ticket_code': new_code,
                'keep_modal_open': False
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f"ERROR: {str(e)}",
            'keep_modal_open': True
        })

@app.route('/addSoftwareIssue', methods=['GET', 'POST'])
def addSoftwareIssue():
    try:
        if request.method == 'POST':
            requester = session['name']
            title = request.form['title']
            app_name = request.form['app']
            reason = request.form['reason']
            issue = request.form['issue']

            # Obter valores de CC (suporta tanto select múltiplo quanto string legada)
            cc_values = request.form.getlist('cc_emails')
            if not cc_values:
                cc_values = request.form.get('cc_emails', '')
            # Se for string, converter para lista
            if isinstance(cc_values, str):
                cc_list = [v.strip() for v in cc_values.split(',') if v.strip()]
            else:
                cc_list = cc_values
            # String para armazenar no banco (valores originais separados por vírgula)
            cc_emails = ', '.join(cc_list) if cc_list else ''
            # Lista de emails para envio (converte usernames)
            cc_emails_list = get_emails_from_cc_input(cc_list)

            current_year = datetime.now().year

            # Handle file upload
            filename = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename != '':
                    if allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        file.save(file_path)
                    else:
                        return jsonify({
                            'success': False, 
                            'message': f'Invalid file type: {file.filename}. Dangerous file types are not allowed for security reasons.',
                            'keep_modal_open': True
                        })

            conn = connect()
            cursor = conn.cursor()
            cursor.execute("SELECT TOP 1 internal_code FROM software_issue ORDER BY id DESC")
            last_code = cursor.fetchone()

            if last_code:
                last_number = int(last_code[0].split('-')[-1])
                new_number = last_number + 1
            else:
                new_number = 1
            new_code = f"DTSI-{current_year}-{new_number:03d}"

            cursor.execute('''
                INSERT INTO software_issue
                (internal_code, title, requester, app_name, reason, issue, cc_emails, filename, created_at, updated_at)
                OUTPUT INSERTED.ID
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE())
            ''', (new_code, title, requester, app_name, reason, issue, cc_emails, filename))
            id = cursor.fetchone()[0]
            conn.commit()

            # Enviar email de abertura para o requester com cópia para os CCs
            try:
                cursor.execute("SELECT email, name FROM users WHERE name = ?", (requester,))
                user_row = cursor.fetchone()
                if user_row:
                    to_email, to_name = user_row
                    send_ticket_status_email(
                        to_email, to_name, new_code, title, 'opened',
                        cc_emails=cc_emails_list
                    )
            except Exception as e:
                print(f"[EMAIL ERROR] Failed to send creation email for ticket {new_code}: {str(e)}")

            cursor.close()
            conn.close()

            return jsonify({
                'success': True,
                'message': 'Ticket Created Successfully!',
                'ticket_code': new_code,
                'keep_modal_open': False
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f"ERROR: {str(e)}",
            'keep_modal_open': True
        })

@app.route('/Uploads/<filename>')
def uploaded_file(filename):
    try:
        # Securely serve the file from the upload folder
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(filename))
        if not os.path.exists(file_path):
            flash('File not found.', category='error')
            return redirect(url_for('home'))
        return send_file(file_path, as_attachment=False)
    except Exception as e:
        flash(f'Error accessing file: {str(e)}', category='error')
        return redirect(url_for('home'))

@app.route('/addSoftwareInternalReport', methods=['GET', 'POST'])
def addSoftwareInternalReport():
    try:
        # Check if user has Software or Admin category
        user_category = (session.get('category') or '').lower()
        username = session.get('username')
        
        if not username:
            logging.error("addSoftwareInternalReport: User not authenticated")
            return jsonify({
                'success': False,
                'message': 'Not authenticated'
            }), 401
        
        # Only Software or Admin category can create DTIR
        if user_category not in ['software', 'admin']:
            logging.warning(f"addSoftwareInternalReport: User {username} with category '{user_category}' attempted to create DTIR")
            return jsonify({
                'success': False,
                'message': 'You do not have permission to create internal reports. Only Admin and Software users can access this feature.'
            }), 403
        
        if request.method == 'POST':
            title = request.form.get('title', '')
            description = request.form.get('description', '')
            reporter = request.form.get('reporter', '').strip()
            requester = session.get('name', '')

            # Obter valores de CC (suporta tanto select múltiplo quanto string legada)
            cc_values = request.form.getlist('cc_emails')
            if not cc_values:
                cc_values = request.form.get('cc_emails', '')
            # Se for string, converter para lista
            if isinstance(cc_values, str):
                cc_list = [v.strip() for v in cc_values.split(',') if v.strip()]
            else:
                cc_list = cc_values
            # String para armazenar no banco (valores originais separados por vírgula)
            cc_emails = ', '.join(cc_list) if cc_list else ''
            # Lista de emails para envio (converte usernames)
            cc_emails_list = get_emails_from_cc_input(cc_list)

            if not title or not description or not reporter:
                return jsonify({
                    'success': False,
                    'message': 'Title, reporter and description are required',
                    'keep_modal_open': True
                })
            
            current_year = datetime.now().year
            priority = request.form.get('priority', 'Medium')  # Default: Medium

            # Handle file upload
            filename = None
            if 'image' in request.files:
                file = request.files['image']
                if file and file.filename != '':
                    if allowed_file(file.filename):
                        filename = secure_filename(file.filename)
                        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        file.save(file_path)
                    else:
                        return jsonify({
                            'success': False,
                            'message': f'Invalid file type: {file.filename}. Dangerous file types are not allowed for security reasons.',
                            'keep_modal_open': True
                        })

            conn = connect()
            cursor = conn.cursor()

            # Ensure priority column exists in software_internal_reports table
            try:
                cursor.execute("""
                    IF NOT EXISTS (SELECT * FROM sys.columns WHERE Name = N'priority' AND Object_ID = Object_ID(N'software_internal_reports'))
                    BEGIN
                        ALTER TABLE software_internal_reports ADD priority NVARCHAR(20) DEFAULT 'Medium' NULL
                    END
                """)
                conn.commit()
            except Exception as e:
                logging.warning(f"Could not ensure priority column exists: {str(e)}")

            # Get the last internal code
            cursor.execute("SELECT TOP 1 internal_code FROM software_internal_reports ORDER BY id DESC")
            last_code = cursor.fetchone()

            if last_code:
                last_number = int(last_code[0].split('-')[-1])
                new_number = last_number + 1
            else:
                new_number = 1

            new_code = f"DTIR-{current_year}-{new_number:03d}"

            cursor.execute('''
                INSERT INTO software_internal_reports
                (internal_code, title, requester, reporter, description, cc_emails, filename, priority, status, approved, created_at, updated_at)
                OUTPUT INSERTED.ID
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, GETDATE(), GETDATE())
            ''', (new_code, title, requester, reporter, description, cc_emails, filename, priority))

            id = cursor.fetchone()[0]
            conn.commit()

            # Enviar email de abertura para o requester com cópia para os CCs
            try:
                cursor.execute("SELECT email, name FROM users WHERE name = ?", (requester,))
                user_row = cursor.fetchone()
                if user_row:
                    to_email, to_name = user_row
                    send_ticket_status_email(
                        to_email, to_name, new_code, title, 'opened',
                        cc_emails=cc_emails_list
                    )
            except Exception as e:
                print(f"[EMAIL ERROR] Failed to send creation email for ticket {new_code}: {str(e)}")

            cursor.close()
            conn.close()

            return jsonify({
                'success': True,
                'message': 'Internal Report Created Successfully!',
                'ticket_code': new_code,
                'keep_modal_open': False
            })
    except Exception as e:
        logging.error(f"Error in addSoftwareInternalReport: {str(e)}")
        return jsonify({
            'success': False,
            'message': f"ERROR: {str(e)}",
            'keep_modal_open': True
        })
    
@app.route('/delete_ticket/<internal_code>', methods=['DELETE'])
def delete_ticket(internal_code):
    try:
        username = session.get('username')
        if not username:
            return jsonify(success=False, message="Not authenticated"), 401

        conn = connect()
        cursor = conn.cursor()
        # Determine table by code prefix
        if internal_code.startswith('DTAI'):
            table = 'automation_improvement'
        elif internal_code.startswith('DTAS'):
            table = 'automation_support'
        elif internal_code.startswith('DTNA'):
            table = 'new_application'
        elif internal_code.startswith('DTSI'):
            table = 'software_issue'
        elif internal_code.startswith('DTIR'):
            table = 'software_internal_reports'
        else:
            return jsonify(success=False, message="Invalid ticket code"), 400

        # First, check if the ticket exists and get its status
        if internal_code.startswith('DTIR'):
            cursor.execute(f"SELECT status, reporter FROM {table} WHERE internal_code = ? AND is_deleted = 0", (internal_code,))
        else:
            cursor.execute(f"SELECT status, requester FROM {table} WHERE internal_code = ? AND is_deleted = 0", (internal_code,))
        ticket = cursor.fetchone()
        if not ticket:
            return jsonify(success=False, message="Ticket not found"), 404
        
        status, db_requester = ticket
        
        if status in (2, -1):
            return jsonify(success=False, message="Tickets cannot be deleted once they are completed"), 400
        
        # Now soft delete the ticket
        cursor.execute(f"UPDATE {table} SET is_deleted = 1 WHERE internal_code = ?", (internal_code,))
        conn.commit()
        updated = cursor.rowcount
        cursor.close()
        conn.close()
        if updated:
            return jsonify(success=True)
        else:
            return jsonify(success=False, message="Failed to delete ticket"), 500
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500
    
    
@app.route('/get_software_applications', methods=['GET'])
def get_software_applications():
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT application_name
            FROM [DT_request].[dbo].[planning]
            WHERE category = 'software'
            ORDER BY application_name ASC
        """)
        apps = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'applications': apps})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/completed_tickets', methods=['GET', 'POST'])
def completed_tickets():
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute('Exec GetAllCompletedRequests')
        columns = [column[0] for column in cursor.description]
        completed_requests = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Fetch completed DTIR tickets separately (SP may not include them or may return NULL for time)
        cursor.execute("""
            SELECT internal_code, title, requester, reporter, start_date, end_date,
                   time, responsible, observations, status
            FROM software_internal_reports
            WHERE status = 2 AND is_deleted = 0
        """)
        dtir_columns = [col[0] for col in cursor.description]
        dtir_rows = [dict(zip(dtir_columns, row)) for row in cursor.fetchall()]

        # Remove any DTIR rows from SP result (may have wrong/null time) and replace with direct query
        completed_requests = [r for r in completed_requests if not r.get('internal_code', '').startswith('DTIR')]
        completed_requests.extend(dtir_rows)

        # Extract unique responsibles for dropdown filter
        responsibles = sorted(set(
            req['responsible'].strip() 
            for req in completed_requests 
            if req.get('responsible') is not None and req['responsible'].strip()
        ))
        
        cursor.close()
        conn.close()
        return render_template('completed_tickets.html', completed_requests=completed_requests, responsibles=responsibles)
    except Exception as e:
        flash(f"Error: {str(e)}", category='error')
        return redirect(url_for('index'))

@app.route('/pending_tickets', methods=['GET', 'POST'])
def pending_tickets():
    try:
        name = session['name']
        conn = connect()
        cursor = conn.cursor()
        
        user_identifier = session.get('name') or session.get('username')
        cursor.execute("EXEC GetPendingRequests ?", (user_identifier,))
        
        columns = [column[0] for column in cursor.description]
        pending_requests = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        cursor.close()
        conn.close()
                
        return render_template('pending_tickets.html', pending_requests=pending_requests)
    except Exception as e:
        flash(f"Error: {str(e)}", category='error')
        return redirect(url_for('index'))

@app.route('/edit_task', methods=['POST'])
def edit_task():
    if 'name' not in session:
        return redirect(url_for('login'))
        
    try:
        request_code = request.form.get('request_code')
        status = request.form.get('status')
        approved = request.form.get('approved')  # Adicionar suporte para approved
        expected_date = request.form.get('expected_date')
        responsible = request.form.get('responsible')
        notes = request.form.get('notes', '')
        
        print(f"Received form data - request_code: {request_code}, status: {status}, approved: {approved}, expected_date: {expected_date}, responsible: {responsible}, notes: {notes}")
        
        if not request_code or not status:
            missing_fields = []
            if not request_code:
                missing_fields.append('request_code')
            if not status:
                missing_fields.append('status')
            flash(f'Missing required fields: {", ".join(missing_fields)}', 'error')
            return redirect(url_for('pending_tickets'))
            
        # Validate status value - should be 0, 1, 3, 4, 5, 6, or 7 for pending tasks
        if status not in ['0', '1', '3', '4', '5', '6', '7']:
            flash('Invalid status value. Status must be 0 (Waiting), 1 (In Progress), 3 (Under Analysis), 4 (Waiting from DT), 5 (Waiting from Requester), 6 (Waiting for Line Availability), or 7 (Waiting Maintenance Action)', 'error')
            return redirect(url_for('pending_tickets'))
        
        # Convert status to integer
        status_int = int(status)
        # Check if approved was explicitly passed
        has_approved_field = approved is not None and approved != ''
        
        # Buscar o status atual antes da atualização
        conn = connect()
        cursor = conn.cursor()
        
        # Determinar tabela baseada no código do ticket
        if request_code.startswith('DTAI'):
            table = 'automation_improvement'
            query_select = "SELECT status, notes FROM automation_improvement WHERE internal_code = ? AND is_deleted = 0"
            if has_approved_field:
                query_update = """UPDATE automation_improvement 
                            SET status = ?, approved = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                            WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
            else:
                query_update = """UPDATE automation_improvement 
                            SET status = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                            WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
        elif request_code.startswith('DTAS'):
            table = 'automation_support'
            query_select = "SELECT status, notes FROM automation_support WHERE internal_code = ? AND is_deleted = 0"
            if has_approved_field:
                query_update = """UPDATE automation_support 
                            SET status = ?, approved = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                            WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
            else:
                query_update = """UPDATE automation_support 
                            SET status = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                            WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
        elif request_code.startswith('DTNA'):
            table = 'new_application'
            query_select = "SELECT status, notes FROM new_application WHERE internal_code = ? AND is_deleted = 0"
            if has_approved_field:
                query_update = """UPDATE new_application 
                            SET status = ?, approved = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                            WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
            else:
                query_update = """UPDATE new_application 
                            SET status = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                            WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
        elif request_code.startswith('DTSI'):
            table = 'software_issue'
            query_select = "SELECT status, notes FROM software_issue WHERE internal_code = ? AND is_deleted = 0"
            if has_approved_field:
                query_update = """UPDATE software_issue 
                            SET status = ?, approved = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                            WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
            else:
                query_update = """UPDATE software_issue 
                            SET status = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                            WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
        elif request_code.startswith('DTIR'):
            table = 'software_internal_reports'
            query_select = "SELECT status, notes FROM software_internal_reports WHERE internal_code = ? AND is_deleted = 0"
            query_update = """UPDATE software_internal_reports 
                        SET status = ?, expected_date = ?, responsible = ?, notes = ?, updated_at = GETDATE()
                        WHERE internal_code = ? AND status IN (0, 1, 3, 4, 5, 6, 7) AND is_deleted = 0"""
        else:
            flash('Invalid request code format', 'error')
            return redirect(url_for('pending_tickets'))
        
        # Buscar status e notes atuais
        cursor.execute(query_select, (request_code,))
        current_data = cursor.fetchone()
        current_status = current_data[0] if current_data else None
        current_notes = current_data[1] if current_data else None
        
        # Validação: Bloquear mudança PARA "Under Analysis" (3) se o status atual NÃO for "Under Analysis"
        if status_int == 3 and current_status != 3:
            flash('Cannot change status back to Under Analysis. Once moved from Under Analysis, you cannot return to it.', 'error')
            return redirect(url_for('pending_tickets'))
        
        # Auto-set approved = 1 when moving OUT of Under Analysis (status 3)
        if current_status == 3 and status_int != 3:
            has_approved_field = True
            approved = '1'
        
        # Execute the update query com ou sem o parâmetro approved
        if table == 'software_internal_reports':
            # DTIR: no approved field
            cursor.execute(query_update, (status_int, expected_date or None, responsible or None, notes, request_code))
        elif has_approved_field:
            approved_int = 1 if approved == '1' else 0
            cursor.execute(query_update, (status_int, approved_int, expected_date or None, responsible or None, notes, request_code))
        else:
            cursor.execute(query_update, (status_int, expected_date or None, responsible or None, notes, request_code))
        
        if cursor.rowcount == 0:
            flash('Task not found or not editable (task may be completed or rejected)', 'error')
        else:
            conn.commit()
            
            # Buscar nome do responsável
            responsible_name = None
            if responsible:
                cursor.execute("SELECT name FROM users WHERE username = ?", (responsible,))
                row_resp = cursor.fetchone()
                if row_resp:
                    responsible_name = row_resp[0]

            # Formatar data para DD/MM/YYYY
            expected_date_fmt = ''
            if expected_date:
                try:
                    expected_date_fmt = datetime.strptime(expected_date, "%Y-%m-%d").strftime("%d/%m/%Y")
                except Exception:
                    expected_date_fmt = expected_date

            # Enviar email se o status mudou OU se as notas mudaram
            status_changed = current_status != status_int
            notes_changed = (current_notes or '') != (notes or '')
            
            if status_changed or notes_changed:
                cursor.execute(f"""
                    SELECT u.email, u.name, t.title
                    FROM users u
                    JOIN {table} t ON (u.name = t.requester OR u.username = t.requester)
                    WHERE t.internal_code = ?
                """, (request_code,))
                row = cursor.fetchone()
                if row:
                    to_email, to_name, ticket_title = row
                    
                    # Mapear status para email type
                    status_map = {
                        0: 'in_progress',
                        1: 'in_progress',
                        3: 'under_analysis',
                        4: 'waiting_dt',
                        5: 'waiting_requester',
                        6: 'waiting_line',
                        7: 'waiting_maintenance'
                    }
                    email_status = status_map.get(status_int, 'in_progress')
                    
                    send_ticket_status_email(
                        to_email, to_name, request_code, ticket_title, email_status,
                        responsible=responsible_name, expected_date=expected_date_fmt, notes=notes
                    )
            
            status_map_text = {
                0: 'Waiting',
                1: 'In Progress',
                3: 'Under Analysis',
                4: 'Waiting from DT',
                5: 'Waiting from Requester',
                6: 'Waiting for Line Availability',
                7: 'Waiting Maintenance Action'
            }
            status_text = status_map_text.get(status_int, 'Unknown')
            flash(f'Task {request_code} updated successfully to {status_text}!')

            # Add to week plan if checkbox was checked
            add_to_week_plan = request.form.get('add_to_week_plan', '')
            if add_to_week_plan == '1':
                if table in ('automation_improvement', 'automation_support'):
                    task_category = 'Automation'
                    desc_col = 'action_to_improve'
                else:
                    task_category = 'Software'
                    if table == 'software_issue':
                        desc_col = 'issue'
                    elif table == 'new_application':
                        desc_col = 'objective'
                    else:
                        desc_col = 'description'

                # Fetch ticket details including priority if available, and start_date for planned_start_date
                if table == 'software_internal_reports':
                    # Ensure priority column exists
                    try:
                        cursor.execute("""
                            IF NOT EXISTS (SELECT * FROM sys.columns WHERE Name = N'priority' AND Object_ID = Object_ID(N'software_internal_reports'))
                            BEGIN
                                ALTER TABLE software_internal_reports ADD priority NVARCHAR(20) DEFAULT 'Medium' NULL
                            END
                        """)
                        conn.commit()
                    except Exception as e:
                        logging.warning(f"Could not ensure priority column exists: {str(e)}")
                    cursor.execute(f"SELECT title, {desc_col}, priority, start_date FROM {table} WHERE internal_code = ?", (request_code,))
                else:
                    cursor.execute(f"SELECT title, {desc_col} FROM {table} WHERE internal_code = ?", (request_code,))

                ticket_row = cursor.fetchone()
                if ticket_row:
                    ticket_title_wp = ticket_row[0] or request_code
                    ticket_desc_wp = ticket_row[1] or ''

                    # Get priority (for DTIR it comes from ticket, for others default to Medium)
                    ticket_priority = 'Medium'
                    if table == 'software_internal_reports' and len(ticket_row) >= 3:
                        ticket_priority = ticket_row[2] or 'Medium'

                    # Get planned_start_date (if available, otherwise None)
                    planned_start_date_wp = None
                    if table == 'software_internal_reports' and len(ticket_row) >= 4 and ticket_row[3]:
                        planned_start_date_wp = ticket_row[3]

                    # Calculate week number and planned_end_date from expected_date
                    week_num_wp = None
                    planned_end_date_wp = None
                    if expected_date:
                        try:
                            dt_wp = datetime.strptime(expected_date, "%Y-%m-%d").date()
                            week_num_wp = dt_wp.isocalendar()[1]
                            planned_end_date_wp = dt_wp
                        except Exception:
                            pass

                    # Determine if this should be principal or child task based on existing tasks for this ticket
                    is_principal_wp = 1
                    principal_task_id_wp = None

                    cursor.execute("""
                        SELECT id, is_principal_task FROM tasks
                        WHERE ticket_internal_code = ?
                        ORDER BY created_at ASC
                    """, (request_code,))
                    existing_tasks_wp = cursor.fetchall()

                    if existing_tasks_wp:
                        # Find existing principal task
                        principal_task_wp = None
                        for existing_task in existing_tasks_wp:
                            if existing_task[1] == 1:  # is_principal_task = 1
                                principal_task_wp = existing_task
                                break

                        if principal_task_wp:
                            # Already has a principal, this is a subtask
                            principal_task_id_wp = principal_task_wp[0]
                            is_principal_wp = 0
                        else:
                            # No principal exists - promote the first existing task to principal
                            first_task_id = existing_tasks_wp[0][0]
                            cursor.execute("""
                                UPDATE tasks SET is_principal_task = 1
                                WHERE id = ?
                            """, (first_task_id,))
                            conn.commit()
                            # This new task will be a subtask pointing to the promoted task
                            principal_task_id_wp = first_task_id
                            is_principal_wp = 0

                    wp_username = session.get('username', 'unknown')
                    cursor.execute("""
                        INSERT INTO [DT_request].[dbo].[tasks] (
                            category, week_number, title, description, responsible,
                            priority, status, planned_end_date, planned_start_date,
                            ticket_internal_code, ticket_table, task_type, created_by, created_at, updated_at,
                            is_principal_task, principal_task_id
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE(), ?, ?)
                    """, (
                        task_category, week_num_wp, ticket_title_wp, ticket_desc_wp,
                        responsible or None, ticket_priority, 'To Do',
                        planned_end_date_wp, planned_start_date_wp,
                        request_code, table, wp_username, wp_username,
                        is_principal_wp, principal_task_id_wp
                    ))
                    conn.commit()

        cursor.close()
        conn.close()

    except ValueError as e:
        flash(f'Invalid status value: {str(e)}', 'error')
    except Exception as e:
        flash(f'Error updating task: {str(e)}', 'error')

    return redirect(url_for('pending_tickets'))

@app.route('/all_pending_tickets', methods=['GET', 'POST'])
def all_pending_tickets():
    try:
        conn = connect()
        cursor = conn.cursor()
        
        # Use the updated stored procedure to get all pending requests (including Under Analysis)
        cursor.execute('EXEC [GetAllPendingRequests]')
        
        columns = [column[0] for column in cursor.description]
        pending_requests = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Extract unique responsibles for dropdown filter
        responsibles = sorted(set(
            (req.get('responsible') or '').strip() 
            for req in pending_requests 
            if (req.get('responsible') or '').strip()
        ))
        
        cursor.close()
        conn.close()

        return render_template('all_pending_tickets.html', 
                             pending_requests=pending_requests,
                             responsibles=responsibles)
    except Exception as e:
        flash(f"Error fetching all pending tasks: {str(e)}", category='error')
        return redirect(url_for('index'))

@app.route('/get_task_details/<task_code>', methods=['GET'])
def get_task_details(task_code):
    if 'name' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
        
    try:
        conn = connect()
        cursor = conn.cursor()
        
        # Determine which table to query based on task code
        if task_code.startswith('DTAI'):
            query = """SELECT internal_code, title, requester, prod_line, n_sap, reason,
                            current_process, action_to_improve, expected_date,
                            responsible, status, filename, notes, created_at, updated_at,
                            requester_response,
                            NULL as department, NULL as objective, NULL as app_name, NULL as issue
                    FROM automation_improvement
                    WHERE internal_code = ? AND (approved = 1 OR approved IS NULL) AND is_deleted = 0"""
        elif task_code.startswith('DTAS'):
            query = """SELECT internal_code, title, requester, prod_line, n_sap, reason,
                              current_process, action_to_improve, ISNULL(observations_requester, '') as observations_requester,
                              ISNULL(observations_dt, '') as observations_dt, expected_date, responsible, status, filename, notes, created_at, updated_at,
                              requester_response,
                              NULL as department, NULL as objective, NULL as app_name, NULL as issue
                      FROM automation_support
                      WHERE internal_code = ? AND (approved = 1 OR approved IS NULL) AND is_deleted = 0"""
        elif task_code.startswith('DTNA'):
            query = """SELECT internal_code, title, requester, NULL as prod_line, NULL as n_sap,
                              reason, current_process, NULL as action_to_improve, observations,
                              expected_date, responsible, status, filename, notes, department, objective, created_at, updated_at,
                              requester_response,
                              NULL as app_name, NULL as issue
                      FROM new_application
                      WHERE internal_code = ? AND (approved = 1 OR approved IS NULL) AND is_deleted = 0"""
        elif task_code.startswith('DTSI'):
            query = """SELECT internal_code, title, requester, prod_line, NULL as n_sap, reason,
                              NULL as current_process, NULL as action_to_improve, observations,
                              expected_date, responsible, status, filename, notes, NULL as department,
                              NULL as objective, app_name, issue, created_at, updated_at, requester_response
                      FROM software_issue
                      WHERE internal_code = ? AND (approved = 1 OR approved IS NULL) AND is_deleted = 0"""
        elif task_code.startswith('DTIR'):
            query = """SELECT internal_code, title, requester, reporter, description,
                              description as action_to_improve, description as issue,
                              status, filename, created_at, updated_at,
                              notes, expected_date, responsible,
                              ISNULL(start_date, NULL) as start_date, ISNULL(end_date, NULL) as end_date,
                              ISNULL(time, NULL) as time, ISNULL(observations, '') as observations,
                              requester_response,
                              NULL as prod_line, NULL as n_sap, NULL as reason,
                              NULL as current_process,
                              NULL as department, NULL as objective, NULL as app_name
                      FROM software_internal_reports
                      WHERE internal_code = ? AND is_deleted = 0"""
        else:
            return jsonify({'error': 'Invalid task code'}), 400
        
        cursor.execute(query, (task_code,))
        columns = [column[0] for column in cursor.description]
        result = cursor.fetchone()
        
        if result:
            task_data = dict(zip(columns, result))
            
            # Ensure status is properly converted to integer
            if task_data['status'] is not None:
                task_data['status'] = int(task_data['status'])
            else:
                task_data['status'] = 0
            
            # Format expected_date properly
            if task_data.get('expected_date'):
                expected_date = task_data['expected_date']
                if hasattr(expected_date, 'strftime'):
                    task_data['expected_date'] = expected_date.strftime('%Y-%m-%d')
                elif isinstance(expected_date, str):
                    if expected_date.strip() == '' or expected_date.strip().lower() == 'being analyzed':
                        task_data['expected_date'] = ''
                    else:
                        try:
                            if ' ' in expected_date:
                                task_data['expected_date'] = expected_date.split(' ')[0]
                            elif 'T' in expected_date:
                                task_data['expected_date'] = expected_date.split('T')[0]
                            else:
                                task_data['expected_date'] = expected_date
                        except:
                            task_data['expected_date'] = expected_date
                else:
                    task_data['expected_date'] = str(expected_date) if expected_date else ''
            else:
                task_data['expected_date'] = ''
            
            # Add observations_requester field mapping for compatibility
            if task_code.startswith('DTAS') and 'observations_requester' not in task_data:
                task_data['observations_requester'] = task_data.get('observations_requester', '')
            elif not task_code.startswith('DTAS'):
                task_data['observations_requester'] = task_data.get('observations', '')

            # Parse responsible to array for multi-select support
            task_data['responsible'] = parse_responsible(task_data.get('responsible'))

            return jsonify(task_data)
        else:
            return jsonify({'error': 'Task not found'}), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/get_ticket_tasks/<ticket_code>', methods=['GET'])
def get_ticket_tasks(ticket_code):
    try:
        conn = connect()
        cursor = conn.cursor()
        
        # Get all tasks associated with this ticket
        cursor.execute("""
            SELECT id, title, description, status, CAST(time_spent AS FLOAT) as time_spent
            FROM tasks
            WHERE ticket_internal_code = ?
            ORDER BY updated_at DESC
        """, (ticket_code,))
        
        columns = [column[0] for column in cursor.description]
        tasks = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'tasks': tasks
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/check_duplicate', methods=['GET'])
def check_duplicate():
    username = session.get('username')
    if not username:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401

    title = request.args.get('title', '').strip()
    reason = request.args.get('reason', '').strip()
    ticket_type = request.args.get('type', '').upper()

    if not title or not ticket_type:
        return jsonify({'success': False, 'message': 'Missing required parameters'}), 400

    # Map ticket type to table and reason column
    type_map = {
        'DTAI': {'table': 'automation_improvement', 'reason_col': 'reason'},
        'DTAS': {'table': 'automation_support', 'reason_col': 'reason'},
        'DTNA': {'table': 'new_application', 'reason_col': 'reason'},
        'DTSI': {'table': 'software_issue', 'reason_col': 'reason'},
        'DTIR': {'table': 'software_internal_reports', 'reason_col': None}
    }

    if ticket_type not in type_map:
        return jsonify({'success': False, 'message': 'Invalid ticket type'}), 400

    table_info = type_map[ticket_type]
    table = table_info['table']
    reason_col = table_info['reason_col']

    conn = connect()
    cursor = conn.cursor()

    # Build query: search by title (and reason if provided)
    if reason_col:
        sql = f"SELECT internal_code, title, status FROM {table} WHERE title LIKE ? AND reason LIKE ? AND is_deleted = 0"
        params = [f'%{title}%', f'%{reason}%']
    else:
        sql = f"SELECT internal_code, title, status FROM {table} WHERE title LIKE ? AND is_deleted = 0"
        params = [f'%{title}%']

    sql += " ORDER BY created_at DESC"

    try:
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        duplicates = [{'internal_code': row[0], 'title': row[1], 'status': row[2]} for row in rows]
        return jsonify({'success': True, 'duplicates': duplicates})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/new_tickets', methods=['GET', 'POST'])
def new_tickets():
    try:
        username = session.get('username')
        user_category = (session.get('category') or '').lower()
        
        conn = connect()  
        cursor = conn.cursor()
        cursor.execute("EXEC GetNewRequests")
        columns = [column[0] for column in cursor.description]
        new_tickets = [dict(zip(columns, row)) for row in cursor.fetchall()]
        dtai = [r for r in new_tickets if r['category'] == 'DTAI']
        dtas = [r for r in new_tickets if r['category'] == 'DTAS']
        dtna = [r for r in new_tickets if r['category'] == 'DTNA']
        dtsi = [r for r in new_tickets if r['category'] == 'DTSI']

        # DTIR (Software Internal Reports) - Only visible to Software and Admin category
        dtir = []
        if user_category in ['software', 'admin']:
            cursor.execute("""
                SELECT id, internal_code, title, requester, reporter, description,
                       filename, status, approved, created_at, updated_at,
                       'DTIR' as category
                FROM software_internal_reports
                WHERE approved IS NULL AND is_deleted = 0
                AND (status IS NULL OR status = 3)
                ORDER BY created_at DESC
            """)
            dtir_columns = [column[0] for column in cursor.description]
            dtir = [dict(zip(dtir_columns, row)) for row in cursor.fetchall()]
        else:
            if username:
                logging.warning(f"DTIR access denied for {username} with category {user_category}")

        cursor.close()
        conn.close()
        return render_template('new_tickets.html',
                             dtai=dtai, dtas=dtas, dtna=dtna, dtsi=dtsi, dtir=dtir, user_category=user_category)
    except Exception as e:
        flash(f"Error fetching tickets: {str(e)}", category='error')
        return redirect(url_for('home'))
    
@app.route('/approve_request', methods=['POST'])
def approve_request():
    try:
        request_code = request.form['request_code']
        expected_date = request.form['expected_date']
        responsible = request.form['responsible']
        notes = request.form.get('notes', None)  # Campo opcional

        conn = connect()
        cursor = conn.cursor()
        
        # Determinar tabela e query baseada no código
        if 'DTAI' in request_code:
            table = 'automation_improvement'
            query = """UPDATE automation_improvement 
                       SET expected_date = ?, responsible = ?, approved = 1, status = 0, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ? AND is_deleted = 0"""
        elif 'DTAS' in request_code:
            table = 'automation_support'
            query = """UPDATE automation_support 
                       SET expected_date = ?, responsible = ?, approved = 1, status = 0, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ? AND is_deleted = 0"""
        elif 'DTNA' in request_code:
            table = 'new_application'
            query = """UPDATE new_application 
                       SET expected_date = ?, responsible = ?, approved = 1, status = 0, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ? AND is_deleted = 0"""
        elif 'DTSI' in request_code:
            table = 'software_issue'
            query = """UPDATE software_issue 
                       SET expected_date = ?, responsible = ?, approved = 1, status = 0, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ? AND is_deleted = 0"""
        elif 'DTIR' in request_code:
            table = 'software_internal_reports'
            query = """UPDATE software_internal_reports 
                       SET expected_date = ?, responsible = ?, approved = 1, status = 0, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ? AND is_deleted = 0"""
        else:
            flash('Invalid request code.', 'danger')
            return redirect(url_for('new_tickets'))
            
        cursor.execute(query, (expected_date, responsible, notes, request_code))
        conn.commit()
        
        username = session.get('username', 'unknown')
        
        # Buscar nome do responsável
        responsible_name = None
        if responsible:
            cursor.execute("SELECT name FROM users WHERE username = ?", (responsible,))
            row_resp = cursor.fetchone()
            if row_resp:
                responsible_name = row_resp[0]

        # Formatar data para DD/MM/YYYY
        expected_date_fmt = ''
        if expected_date:
            try:
                expected_date_fmt = datetime.strptime(expected_date, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                expected_date_fmt = expected_date

        # Enviar email de aprovação (skip for DTIR - internal report)
        if 'DTIR' not in request_code:
            cursor.execute(f"""
                SELECT u.email, u.name, t.title
                FROM users u
                JOIN {table} t ON (u.name = t.requester OR u.username = t.requester)
                WHERE t.internal_code = ?
            """, (request_code,))
            row = cursor.fetchone()
            if row:
                to_email, to_name, ticket_title = row
                send_ticket_status_email(
                    to_email, to_name, request_code, ticket_title, 'approved',
                    responsible=responsible_name, expected_date=expected_date_fmt, notes=notes
                )

        # Add to week plan if checkbox was checked
        add_to_week_plan = request.form.get('add_to_week_plan', '')
        if add_to_week_plan == '1':
            # Determine category and fetch ticket details
            if table in ('automation_improvement', 'automation_support'):
                task_category = 'Automation'
                desc_col = 'action_to_improve'
            else:
                task_category = 'Software'
                if table == 'software_issue':
                    desc_col = 'issue'
                elif table == 'new_application':
                    desc_col = 'objective'
                else:
                    desc_col = 'description'

            # Fetch ticket details including priority if available, and start_date for planned_start_date
            if table == 'software_internal_reports':
                # Ensure priority column exists
                try:
                    cursor.execute("""
                        IF NOT EXISTS (SELECT * FROM sys.columns WHERE Name = N'priority' AND Object_ID = Object_ID(N'software_internal_reports'))
                        BEGIN
                            ALTER TABLE software_internal_reports ADD priority NVARCHAR(20) DEFAULT 'Medium' NULL
                        END
                    """)
                    conn.commit()
                except Exception as e:
                    logging.warning(f"Could not ensure priority column exists: {str(e)}")
                cursor.execute(f"SELECT title, {desc_col}, priority, start_date FROM {table} WHERE internal_code = ?", (request_code,))
            else:
                cursor.execute(f"SELECT title, {desc_col} FROM {table} WHERE internal_code = ?", (request_code,))

            ticket_row = cursor.fetchone()
            if ticket_row:
                ticket_title_wp = ticket_row[0] or request_code
                ticket_desc_wp = ticket_row[1] or ''

                # Get priority (for DTIR it comes from ticket, for others default to Medium)
                ticket_priority = 'Medium'
                if table == 'software_internal_reports' and len(ticket_row) >= 3:
                    ticket_priority = ticket_row[2] or 'Medium'

                # Calculate week number and planned dates from expected_date
                # Both planned_start_date and planned_end_date use expected_date
                week_num_wp = None
                planned_end_date_wp = None
                planned_start_date_wp = None
                if expected_date:
                    try:
                        dt_wp = datetime.strptime(expected_date, "%Y-%m-%d").date()
                        week_num_wp = dt_wp.isocalendar()[1]
                        planned_end_date_wp = dt_wp
                        planned_start_date_wp = dt_wp  # Same as expected_date
                    except Exception:
                        pass

                # Determine if this should be principal or child task based on existing tasks for this ticket
                is_principal_wp = 1
                principal_task_id_wp = None

                cursor.execute("""
                    SELECT id, is_principal_task FROM tasks
                    WHERE ticket_internal_code = ?
                    ORDER BY created_at ASC
                """, (request_code,))
                existing_tasks_wp = cursor.fetchall()

                if existing_tasks_wp:
                    # Find existing principal task
                    principal_task_wp = None
                    for existing_task in existing_tasks_wp:
                        if existing_task[1] == 1:  # is_principal_task = 1
                            principal_task_wp = existing_task
                            break

                    if principal_task_wp:
                        # Already has a principal, this is a subtask
                        principal_task_id_wp = principal_task_wp[0]
                        is_principal_wp = 0
                    else:
                        # No principal exists - promote the first existing task to principal
                        first_task_id = existing_tasks_wp[0][0]
                        cursor.execute("""
                            UPDATE tasks SET is_principal_task = 1
                            WHERE id = ?
                        """, (first_task_id,))
                        conn.commit()
                        # This new task will be a subtask pointing to the promoted task
                        principal_task_id_wp = first_task_id
                        is_principal_wp = 0

                cursor.execute("""
                    INSERT INTO [DT_request].[dbo].[tasks] (
                        category, week_number, title, description, responsible,
                        priority, status, planned_end_date, planned_start_date,
                        ticket_internal_code, ticket_table, task_type, created_by, created_at, updated_at,
                        is_principal_task, principal_task_id
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE(), ?, ?)
                """, (
                    task_category, week_num_wp, ticket_title_wp, ticket_desc_wp,
                    responsible or None, ticket_priority, 'To Do',
                    planned_end_date_wp, planned_start_date_wp,
                    request_code, table, username, username,
                    is_principal_wp, principal_task_id_wp
                ))
                conn.commit()

        cursor.close()
        conn.close()
        flash('Request approved successfully!', 'success')
        return redirect(url_for('new_tickets'))
    except Exception as e:
        flash(f'Error while approving the request: {str(e)}', 'danger')
        return redirect(url_for('new_tickets'))

@app.route('/set_under_analysis', methods=['POST'])
def set_under_analysis():
    try:
        request_code = request.form['request_code']
        responsible = request.form['responsible']
        notes = request.form.get('notes', None)  # Campo opcional

        conn = connect()
        cursor = conn.cursor()
        
        # Determinar tabela e query baseada no código
        if 'DTAI' in request_code:
            table = 'automation_improvement'
            query = """UPDATE automation_improvement 
                       SET responsible = ?, status = 3, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ?"""
        elif 'DTAS' in request_code:
            table = 'automation_support'
            query = """UPDATE automation_support 
                       SET responsible = ?, status = 3, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ?"""
        elif 'DTNA' in request_code:
            table = 'new_application'
            query = """UPDATE new_application 
                       SET responsible = ?, status = 3, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ?"""
        elif 'DTSI' in request_code:
            table = 'software_issue'
            query = """UPDATE software_issue 
                       SET responsible = ?, status = 3, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ?"""
        elif 'DTIR' in request_code:
            table = 'software_internal_reports'
            query = """UPDATE software_internal_reports 
                       SET responsible = ?, status = 3, updated_at = GETDATE(), notes = ?
                       WHERE internal_code = ?"""
        else:
            flash('Invalid request code.', 'danger')
            return redirect(url_for('new_tickets'))
            
        cursor.execute(query, (responsible, notes, request_code))
        conn.commit()
        
        # Buscar nome do responsável
        responsible_name = None
        if responsible:
            cursor.execute("SELECT name FROM users WHERE username = ?", (responsible,))
            row_resp = cursor.fetchone()
            if row_resp:
                responsible_name = row_resp[0]

        # Enviar email de "under analysis" (skip for DTIR - internal report)
        if 'DTIR' not in request_code:
            cursor.execute(f"""
                SELECT u.email, u.name, t.title
                FROM users u
                JOIN {table} t ON (u.name = t.requester OR u.username = t.requester)
                WHERE t.internal_code = ?
            """, (request_code,))
            row = cursor.fetchone()
            if row:
                to_email, to_name, ticket_title = row
                send_ticket_status_email(
                    to_email, to_name, request_code, ticket_title, 'under_analysis',
                    responsible=responsible_name, notes=notes
                )        
        cursor.close()
        conn.close()
        flash('Request set to Under Analysis successfully!', 'success')
        return redirect(url_for('new_tickets'))
    except Exception as e:
        flash(f'Error while setting request to Under Analysis: {str(e)}', 'danger')
        return redirect(url_for('new_tickets'))

@app.route('/reject_request', methods=['GET', 'POST'])
def reject_request():
    try:
        request_code = request.form['request_code']
        observations = request.form.get('observations', '')
        conn = connect()
        cursor = conn.cursor()
        
        # Determinar tabela e query baseada no código
        if 'DTAI' in request_code:
            table = 'automation_improvement'
            query = """UPDATE automation_improvement 
                      SET status = -1, observations = ?
                      WHERE internal_code = ?"""
        elif 'DTAS' in request_code:
            table = 'automation_support'
            query = """UPDATE automation_support 
                      SET status = -1, observations_dt = ?
                      WHERE internal_code = ?"""
        elif 'DTNA' in request_code:
            table = 'new_application'
            query = """UPDATE new_application 
                      SET status = -1, observations = ?
                      WHERE internal_code = ?"""
        elif 'DTSI' in request_code:
            table = 'software_issue'
            query = """UPDATE software_issue 
                      SET status = -1, observations = ?
                      WHERE internal_code = ?"""
        elif 'DTIR' in request_code:
            table = 'software_internal_reports'
            query = """UPDATE software_internal_reports 
                      SET status = -1, observations = ?, updated_at = GETDATE()
                      WHERE internal_code = ?"""
        else:
            flash('Invalid request code.', 'danger')
            return redirect(url_for('new_tickets'))
            
        cursor.execute(query, (observations, request_code))
        conn.commit()
        
        username = session.get('username', 'unknown')
        
        # Enviar email de rejeição (skip for DTIR - internal report)
        if 'DTIR' not in request_code:
            cursor.execute(f"""
                SELECT u.email, u.name, t.title
                FROM users u
                JOIN {table} t ON (u.name = t.requester OR u.username = t.requester)
                WHERE t.internal_code = ?
            """, (request_code,))
            row = cursor.fetchone()
            if row:
                to_email, to_name, ticket_title = row
                send_ticket_status_email(to_email, to_name, request_code, ticket_title, 'rejected', observations)
        
        cursor.close()
        conn.close()
        flash('Request rejected successfully!', 'success')
        return redirect(url_for('new_tickets'))
    except Exception as e:
        flash(f'Error while rejecting the request: {str(e)}', 'danger')
        return redirect(url_for('new_tickets'))

@app.route('/conclude_request', methods=['GET', 'POST'])
def conclude_request():
    try:
        # Suportar tanto JSON quanto form-data
        if request.is_json:
            data = request.get_json()
            request_code = data.get('request_code')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            time = int(float(data.get('time', 0)))
            observations = data.get('observations', '')
        else:
            request_code = request.form['request_code']
            start_date = request.form['start_date']
            end_date = request.form['end_date']
            time = int(float(request.form['time']))
            observations = request.form.get('observations', '')

        if not request_code or not start_date or not end_date:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400

        conn = connect()
        cursor = conn.cursor()

        # Determinar tabela e query baseada no código
        if 'DTAI' in request_code:
            table = 'automation_improvement'
            query = """UPDATE automation_improvement
                       SET start_date = ?, end_date = ?, time = ?, observations = ?, status = 2
                       WHERE internal_code = ?"""
        elif 'DTAS' in request_code:
            table = 'automation_support'
            query = """UPDATE automation_support
                       SET start_date = ?, end_date = ?, time = ?, observations_dt = ?, status = 2
                       WHERE internal_code = ?"""
        elif 'DTNA' in request_code:
            table = 'new_application'
            query = """UPDATE new_application
                       SET start_date = ?, end_date = ?, time = ?, observations = ?, status = 2
                       WHERE internal_code = ?"""
        elif 'DTSI' in request_code:
            table = 'software_issue'
            query = """UPDATE software_issue
                       SET start_date = ?, end_date = ?, time = ?, observations = ?, status = 2
                       WHERE internal_code = ?"""
        elif 'DTIR' in request_code:
            table = 'software_internal_reports'
            query = """UPDATE software_internal_reports
                       SET start_date = ?, end_date = ?, time = ?, observations = ?, status = 2, updated_at = GETDATE()
                       WHERE internal_code = ?"""
        else:
            return jsonify({'success': False, 'message': 'Invalid request code'}), 400

        cursor.execute(query, (start_date, end_date, time, observations, request_code))
        conn.commit()

        # Sum the estimated hours from all tasks created from this ticket
        try:
            cursor.execute("""
                SELECT COALESCE(SUM(CAST(estimated_hours AS FLOAT)), 0) as total_hours
                FROM tasks
                WHERE ticket_internal_code = ? AND estimated_hours IS NOT NULL
            """, (request_code,))
            result = cursor.fetchone()
            total_task_hours = result[0] if result and result[0] is not None else 0

            # Update the ticket with total hours from tasks
            if total_task_hours > 0:
                total_time = total_task_hours + time

                if 'DTAI' in request_code:
                    cursor.execute("""UPDATE automation_improvement
                                     SET time = ? WHERE internal_code = ?""", (int(total_time), request_code))
                elif 'DTAS' in request_code:
                    cursor.execute("""UPDATE automation_support
                                     SET time = ? WHERE internal_code = ?""", (int(total_time), request_code))
                elif 'DTNA' in request_code:
                    cursor.execute("""UPDATE new_application
                                     SET time = ? WHERE internal_code = ?""", (int(total_time), request_code))
                elif 'DTSI' in request_code:
                    cursor.execute("""UPDATE software_issue
                                     SET time = ? WHERE internal_code = ?""", (int(total_time), request_code))

                conn.commit()
        except Exception as hours_error:
            # Log do erro mas não falha a operação principal
            print(f"Warning: Could not sum task hours for {request_code}: {str(hours_error)}")

        # Atualizar a task relacionada como "Done"
        try:
            cursor.execute("""
                UPDATE tasks
                SET status = 'Done', start_date = ?, end_date = ?, time_spent = ?
                WHERE ticket_internal_code = ? AND status != 'Done'
            """, (start_date, end_date, time, request_code))
            conn.commit()
        except Exception as task_error:
            # Log do erro mas não falha a operação principal
            print(f"Warning: Could not update related task for {request_code}: {str(task_error)}")

        # Enviar email de conclusão
        cursor.execute(f"""
            SELECT u.email, u.name, t.title
            FROM users u
            JOIN {table} t ON (u.name = t.requester OR u.username = t.requester)
            WHERE t.internal_code = ?
        """, (request_code,))
        row = cursor.fetchone()
        if row:
            to_email, to_name, ticket_title = row
            send_ticket_status_email(to_email, to_name, request_code, ticket_title, 'completed', observations)

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'message': 'Ticket concluded successfully!'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error while concluding the request: {str(e)}'}), 500


@app.route('/waiting_from_requester', methods=['POST'])
def waiting_from_requester():
    try:
        request_code = request.form['request_code']
        observations = request.form.get('observations', '')
        conn = connect()
        cursor = conn.cursor()
        
        # Determinar tabela e query baseada no código
        if 'DTAI' in request_code:
            table = 'automation_improvement'
            query = """UPDATE automation_improvement 
                      SET status = 5, observations = ?
                      WHERE internal_code = ?"""
        elif 'DTAS' in request_code:
            table = 'automation_support'
            query = """UPDATE automation_support 
                      SET status = 5, observations_dt = ?
                      WHERE internal_code = ?"""
        elif 'DTNA' in request_code:
            table = 'new_application'
            query = """UPDATE new_application 
                      SET status = 5, observations = ?
                      WHERE internal_code = ?"""
        elif 'DTSI' in request_code:
            table = 'software_issue'
            query = """UPDATE software_issue 
                      SET status = 5, observations = ?
                      WHERE internal_code = ?"""
        else:
            flash('Invalid request code.', 'danger')
            return redirect(url_for('new_tickets'))
            
        cursor.execute(query, (observations, request_code))
        conn.commit()
        
        username = session.get('username', 'unknown')
        
        # Enviar email de notificação
        cursor.execute(f"""
            SELECT u.email, u.name, t.title
            FROM users u
            JOIN {table} t ON (u.name = t.requester OR u.username = t.requester)
            WHERE t.internal_code = ?
        """, (request_code,))
        row = cursor.fetchone()
        if row:
            to_email, to_name, ticket_title = row
            send_ticket_status_email(to_email, to_name, request_code, ticket_title, 'waiting_requester', observations)
        
        cursor.close()
        conn.close()
        flash('Request set to waiting from requester successfully!', 'success')
        return redirect(url_for('all_pending_tickets'))
    except Exception as e:
        flash(f'Error while setting request to waiting: {str(e)}', 'danger')
        return redirect(url_for('all_pending_tickets'))


@app.route('/reply_waiting', methods=['POST'])
def reply_waiting():
    try:
        request_code = request.form.get('request_code')
        response_text = request.form.get('response_text', '').strip()
        if not request_code or not response_text:
            return jsonify({'success': False, 'message': 'Missing data'}), 400
        # Determine table based on prefix
        if 'DTAI' in request_code:
            table = 'automation_improvement'
        elif 'DTAS' in request_code:
            table = 'automation_support'
        elif 'DTNA' in request_code:
            table = 'new_application'
        elif 'DTSI' in request_code:
            table = 'software_issue'
        elif 'DTIR' in request_code:
            table = 'software_internal_reports'
        else:
            return jsonify({'success': False, 'message': 'Invalid ticket code'}), 400
        # Verify user is the requester
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        conn = connect()
        cursor = conn.cursor()
        cursor.execute(f"SELECT requester, title, responsible FROM {table} WHERE internal_code = ?", (request_code,))
        row = cursor.fetchone()
        if not row:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Ticket not found'}), 404
        requester_name, ticket_title, responsible = row
        if requester_name != session.get('name'):
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'You can only reply to your own tickets'}), 403
        # Update: set requester_response and status back to Under Analysis (3)
        cursor.execute(f"""
            UPDATE {table}
            SET requester_response = ?, status = 3, updated_at = GETDATE()
            WHERE internal_code = ?
        """, (response_text, request_code))
        conn.commit()

        # Send notification to responsible person with additional information
        if responsible:
            try:
                cursor.execute("SELECT email, name FROM users WHERE username = ?", (responsible,))
                resp_row = cursor.fetchone()
                if resp_row:
                    resp_email, resp_name = resp_row
                    send_ticket_status_email(
                        to_email=resp_email,
                        to_name=resp_name,
                        ticket_code=request_code,
                        ticket_title=ticket_title,
                        new_status='under_analysis',
                        extra_msg=f"Requester has provided additional information:\n\n{response_text}",
                        notes=f"Requester added new information to the ticket"
                    )
            except Exception as email_err:
                logging.error(f"Error sending notification email: {str(email_err)}")

        cursor.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Response submitted successfully! The responsible person has been notified.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/edit_ticket/<ticket_code>', methods=['GET', 'POST'])
def edit_ticket(ticket_code):
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401

        conn = connect()
        cursor = conn.cursor()

        # Determinar tabela baseada no código
        if 'DTAI' in ticket_code:
            table = 'automation_improvement'
        elif 'DTAS' in ticket_code:
            table = 'automation_support'
        elif 'DTNA' in ticket_code:
            table = 'new_application'
        elif 'DTSI' in ticket_code:
            table = 'software_issue'
        elif 'DTIR' in ticket_code:
            table = 'software_internal_reports'
        else:
            return jsonify({'success': False, 'message': 'Invalid ticket code'}), 400

        if request.method == 'GET':
            # Buscar dados do ticket
            cursor.execute(f"""
                SELECT * FROM {table} WHERE internal_code = ?
            """, (ticket_code,))
            row = cursor.fetchone()
            if not row:
                cursor.close()
                conn.close()
                return jsonify({'success': False, 'message': 'Ticket not found'}), 404

            # Obter nomes das colunas
            columns = [column[0] for column in cursor.description]
            ticket_data = dict(zip(columns, row))

            # Verificar se o ticket está em waiting_requester (status = 5) ou se o usuário é o requester
            # Permitimos edição se status=5 e o usuário é o requester OU se o usuário é admin/DT
            status = ticket_data.get('status')
            requester = ticket_data.get('requester')
            user_category = session.get('category', '').lower()

            # Buscar nome do usuário logado
            cursor.execute("SELECT name FROM users WHERE username = ?", (username,))
            user_row = cursor.fetchone()
            user_name = user_row[0] if user_row else username

            can_edit = False
            if status == 5 and (user_name == requester or user_category in ['admin', 'dt']):
                can_edit = True
            elif user_category in ['admin', 'dt']:
                can_edit = True  # Admin/DT podem editar qualquer ticket

            if not can_edit:
                cursor.close()
                conn.close()
                return jsonify({'success': False, 'message': 'Ticket not editable. It must be in "Waiting from Requester" status and you must be the requester.'}), 403

            cursor.close()
            conn.close()
            return jsonify({'success': True, 'data': ticket_data})

        elif request.method == 'POST':
            # Atualizar o ticket com os dados enviados
            # Reconstruir query de UPDATE baseada na tabela
            # Vamos aceitar todos os campos exceto internal_code, id, created_at, updated_at
            update_fields = {}
            for key, value in request.form.items():
                if key not in ['internal_code', 'id', 'created_at', 'updated_at']:
                    update_fields[key] = value

            if not update_fields:
                return jsonify({'success': False, 'message': 'No fields to update'}), 400

            # Normalizar cc_emails: se for lista (select múltiplo), converter para string separada por vírgulas
            if 'cc_emails' in update_fields:
                cc_val = update_fields['cc_emails']
                if isinstance(cc_val, list):
                    update_fields['cc_emails'] = ', '.join(cc_val)

            # Primeiro, buscar dados atuais do ticket para notificação e para manter campos que não devem ser alterados (ex: cc_emails original se não enviado)
            cursor.execute(f"SELECT title, responsible, cc_emails FROM {table} WHERE internal_code = ?", (ticket_code,))
            current_ticket = cursor.fetchone()
            if not current_ticket:
                cursor.close()
                conn.close()
                return jsonify({'success': False, 'message': 'Ticket not found'}), 404

            current_title, current_responsible, current_cc_emails = current_ticket

            # Se cc_emails não foi enviado no formulário, manter o atual
            if 'cc_emails' not in update_fields:
                update_fields['cc_emails'] = current_cc_emails

            # Construir query dinamicamente
            set_clause = ', '.join([f"{field} = ?" for field in update_fields.keys()])
            query = f"UPDATE {table} SET {set_clause}, updated_at = GETDATE() WHERE internal_code = ?"

            values = list(update_fields.values()) + [ticket_code]

            cursor.execute(query, values)
            conn.commit()

            # Enviar notificação ao responsável/equipa de que o requerente adicionou informação
            if current_responsible:
                try:
                    cursor.execute("SELECT email, name FROM users WHERE username = ?", (current_responsible,))
                    user_row = cursor.fetchone()
                    if user_row:
                        resp_email, resp_name = user_row
                        # Enviar email ao responsável informando que o ticket foi atualizado com novas informações
                        send_ticket_status_email(
                            to_email=resp_email,
                            to_name=resp_name,
                            ticket_code=ticket_code,
                            ticket_title=current_title,
                            new_status='waiting_requester',  # Mantém o mesmo status
                            extra_msg='',
                            notes='The requester has provided additional information. Please review the ticket.',
                            responsible=resp_name,
                            cc_emails=update_fields.get('cc_emails', current_cc_emails)
                        )
                except Exception as e:
                    logging.error(f"Error sending update notification email: {str(e)}")

            cursor.close()
            conn.close()

            return jsonify({'success': True, 'message': 'Ticket updated successfully'})

    except Exception as e:
        return jsonify({'success': False, 'message': f"ERROR: {str(e)}"})


@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')


@app.route('/tasks/<int:task_id>/create-github-issue', methods=['POST'])
def create_task_github_issue(task_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401

        conn = connect()
        cursor = conn.cursor()

        # Buscar tarefa com informações do projeto
        cursor.execute("""
            SELECT t.title, t.description, t.project_id, t.github_issue_number,
                   p.github_repo, p.responsible
            FROM tasks t
            LEFT JOIN projects p ON t.project_id = p.id
            WHERE t.id = ?
        """, (task_id,))
        row = cursor.fetchone()
        if not row:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Task not found'}), 404

        title, description, project_id, existing_issue, github_repo, responsible_json = row

        # Se já tem Issue associada
        if existing_issue:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Issue already linked', 'issue_number': existing_issue}), 400

        # Se projeto não tem repositório configurado
        if not github_repo:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Project has no GitHub repository configured'}), 400

        # Buscar assignees dos responsáveis do projeto (suporta JSON, ';' e ',')
        assignees = get_user_github_usernames(responsible_json, cursor) if responsible_json else []

        # Criar Issue no GitHub
        issue_number = create_issue(
            repo=github_repo,
            title=title,
            description=description,
            assignees=assignees
        )

        if not issue_number:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Failed to create GitHub Issue'}), 500

        # Guardar número da Issue na tarefa
        cursor.execute("UPDATE tasks SET github_issue_number = ? WHERE id = ?", (issue_number, task_id))
        conn.commit()

        logging.info(f"GitHub Issue #{issue_number} criada manualmente para tarefa {task_id}")

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'issue_number': issue_number,
            'github_repo': github_repo,
            'assignees': assignees
        })

    except Exception as e:
        logging.error(f"Erro ao criar GitHub Issue para tarefa {task_id}: {str(e)}")
        if 'conn' in locals():
            conn.rollback()
            cursor.close()
            conn.close()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/profile/github-username', methods=['GET', 'PUT'])
def profile_github_username():
    username = session.get('username')
    if not username:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401

    conn = connect()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("SELECT github_username FROM users WHERE username = ?", (username,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'github_username': row[0] if row else None})

    # PUT
    data = request.get_json()
    github_username = data.get('github_username', '').strip() or None

    cursor.execute("UPDATE users SET github_username = ? WHERE username = ?", (github_username, username))
    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({'success': True, 'github_username': github_username})


@app.route('/api/profile', methods=['GET', 'PUT'])
def api_profile():
    username = session.get('username')
    if not username:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401

    conn = connect()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("SELECT username, name, email, github_username FROM users WHERE username = ?", (username,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        return jsonify({
            'success': True,
            'username': row[0],
            'name': row[1],
            'email': row[2],
            'github_username': row[3]
        })

    # PUT - update profile
    data = request.get_json()
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    github_username = data.get('github_username', '').strip() or None

    if not name or not email:
        return jsonify({'success': False, 'message': 'Name and email are required'}), 400

    cursor.execute("""
        UPDATE users
        SET name = ?, email = ?, github_username = ?
        WHERE username = ?
    """, (name, email, github_username, username))
    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({'success': True, 'message': 'Profile updated successfully'})


@app.route('/api/profile/password', methods=['PUT'])
def change_password():
    username = session.get('username')
    if not username:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401

    data = request.get_json()
    current_password = data.get('current_password')
    new_password = data.get('new_password')

    if not current_password or not new_password:
        return jsonify({'success': False, 'message': 'Current and new passwords are required'}), 400

    if len(new_password) < 6:
        return jsonify({'success': False, 'message': 'New password must be at least 6 characters'}), 400

    # Verify current password
    conn = connect()
    cursor = conn.cursor()
    cursor.execute("SELECT password FROM users WHERE username = ?", (username,))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'message': 'User not found'}), 404

    stored_hash = row[0]
    from flask_bcrypt import Bcrypt
    bcrypt = Bcrypt()
    if not bcrypt.check_password_hash(stored_hash, current_password):
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'message': 'Current password is incorrect'}), 401

    # Update password
    new_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
    cursor.execute("UPDATE users SET password = ? WHERE username = ?", (new_hash, username))
    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({'success': True, 'message': 'Password changed successfully'})


@app.route('/api/users/search', methods=['GET'])
def search_users():
    """API endpoint para buscar usuários por username, name ou email (autocomplete)"""
    query = request.args.get('q', '').strip()
    if not query or len(query) < 2:
        return jsonify({'success': True, 'users': []})

    try:
        conn = connect()
        cursor = conn.cursor()

        # Buscar usuários que correspondam à query parcial (username, name ou email)
        search_pattern = f"%{query}%"
        cursor.execute('''
            SELECT TOP 20 username, name, email
            FROM users
            WHERE username LIKE ? OR name LIKE ? OR email LIKE ?
            ORDER BY name
        ''', (search_pattern, search_pattern, search_pattern))

        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        users = []
        for row in rows:
            username, name, email = row
            display_name = name if name else username
            users.append({
                'username': username,
                'name': display_name,
                'email': email,
                'display': f"{display_name} <{email}>" if name else f"{username} <{email}>"
            })

        logging.info(f"search_users: query='{query}' returned {len(users)} users")
        return jsonify({'success': True, 'users': users})
    except Exception as e:
        logging.error(f"Error in search_users API: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500


@app.route('/tasks')
def tasks():
    try:
        username = session.get('username')
        user_category = session.get('category', 'admin')  # Get user category from session
        
        if not username:
            flash('Please log in to view tasks.', 'error')
            return redirect(url_for('index'))

        conn = connect()
        cursor = conn.cursor()

        software_tasks = []
        automation_tasks = []

        # Sempre mostrar todas as tarefas para todos os usuários
        # O filtro de categoria será aplicado no frontend como padrão, mas pode ser alterado
        show_software = True
        show_automation = True

        # Fetch Software tasks if needed
        if show_software:
            try:
                # Tentar com schema completo primeiro
                cursor.execute("""
                    SELECT t.id, t.week_number, t.title, t.description, t.responsible, t.equipment,
                           t.equipment_responsible, t.priority, t.status, t.start_date, t.end_date,
                           t.ticket_internal_code, t.ticket_table, t.time_spent, t.planned_start_date,
                           t.planned_end_date, t.estimated_hours, t.created_by, t.project_id, t.task_type,
                           t.is_principal_task, p.github_repo, p.name as project_name
                    FROM [DT_request].[dbo].[tasks] t
                    LEFT JOIN [DT_request].[dbo].[projects] p ON t.project_id = p.id
                    WHERE t.category = ?
                    ORDER BY t.responsible DESC
                """, ('Software',))
                software_tasks = [row_to_dict(cursor, row) for row in cursor.fetchall()]
            except Exception as e1:
                try:
                    # Tentar sem schema
                    cursor.execute("""
                        SELECT t.id, t.week_number, t.title, t.description, t.responsible, t.equipment,
                               t.equipment_responsible, t.priority, t.status, t.start_date, t.end_date,
                               t.ticket_internal_code, t.ticket_table, t.time_spent, t.planned_start_date,
                               t.planned_end_date, t.estimated_hours, t.created_by, t.project_id, t.task_type,
                               t.is_principal_task, p.github_repo, p.name as project_name
                        FROM tasks t
                        LEFT JOIN projects p ON t.project_id = p.id
                        WHERE t.category = ?
                        ORDER BY t.responsible
                    """, ('Software',))
                    software_tasks = [row_to_dict(cursor, row) for row in cursor.fetchall()]
                except Exception as e2:
                    # Se ainda falhar, vamos listar todas as tarefas sem filtro de categoria
                    cursor.execute("""
                        SELECT t.id, t.week_number, t.title, t.description, t.responsible, t.equipment,
                               t.equipment_responsible, t.priority, t.status, t.start_date, t.end_date,
                               t.ticket_internal_code, t.ticket_table, t.time_spent, t.planned_start_date,
                               t.planned_end_date, t.estimated_hours, t.created_by, t.project_id, t.task_type,
                               t.is_principal_task, p.github_repo, p.name as project_name
                        FROM tasks t
                        LEFT JOIN projects p ON t.project_id = p.id
                        WHERE t.category = 'Software' OR t.category IS NULL
                        ORDER BY t.responsible DESC
                    """)
                    software_tasks = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        # Fetch Automation tasks if needed
        if show_automation:
            try:
                cursor.execute("""
                    SELECT t.id, t.week_number, t.title, t.description, t.responsible, t.equipment,
                           t.equipment_responsible, t.priority, t.status, t.start_date, t.end_date,
                           t.ticket_internal_code, t.ticket_table, t.time_spent, t.planned_start_date,
                           t.planned_end_date, t.estimated_hours, t.created_by, t.project_id, t.task_type,
                           t.is_principal_task, p.github_repo, p.name as project_name
                    FROM [DT_request].[dbo].[tasks] t
                    LEFT JOIN [DT_request].[dbo].[projects] p ON t.project_id = p.id
                    WHERE t.category = ?
                    ORDER BY t.responsible DESC
                """, ('Automation',))
                automation_tasks = [row_to_dict(cursor, row) for row in cursor.fetchall()]
            except Exception as e1:
                try:
                    cursor.execute("""
                        SELECT t.id, t.week_number, t.title, t.description, t.responsible, t.equipment,
                               t.equipment_responsible, t.priority, t.status, t.start_date, t.end_date,
                               t.ticket_internal_code, t.ticket_table, t.time_spent, t.planned_start_date,
                               t.planned_end_date, t.estimated_hours, t.created_by, t.project_id, t.task_type,
                               t.is_principal_task, p.github_repo, p.name as project_name
                        FROM tasks t
                        LEFT JOIN projects p ON t.project_id = p.id
                        WHERE t.category = ?
                        ORDER BY t.responsible
                    """, ('Automation',))
                    automation_tasks = [row_to_dict(cursor, row) for row in cursor.fetchall()]
                except Exception as e2:
                    cursor.execute("""
                        SELECT t.id, t.week_number, t.title, t.description, t.responsible, t.equipment,
                               t.equipment_responsible, t.priority, t.status, t.start_date, t.end_date,
                               t.ticket_internal_code, t.ticket_table, t.time_spent, t.planned_start_date,
                               t.planned_end_date, t.estimated_hours, t.created_by, t.project_id, t.task_type,
                               t.is_principal_task, p.github_repo, p.name as project_name
                        FROM tasks t
                        LEFT JOIN projects p ON t.project_id = p.id
                        WHERE t.category = 'Automation' OR t.category IS NULL
                        ORDER BY t.responsible DESC
                    """)
                    automation_tasks = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        # Fetch Tickets for Software - apenas pending tasks do usuário logado
        cursor.execute("""
            SELECT internal_code, title, issue as description, status, expected_date, 
                   start_date, end_date, prod_line, app_name, reason, observations, time,
                   responsible, requester, 'software_issue' as table_name, NULL as observations_dt
            FROM software_issue
            WHERE responsible = ? AND status IN (0, 1, 4, 5, 6, 7) AND is_deleted = 0
            UNION
            SELECT internal_code, title, objective as description, status, expected_date,
                   start_date, end_date, department, NULL as app_name, reason, observations, 
                   time, responsible, requester, 'new_application' as table_name, NULL as observations_dt
            FROM new_application
            WHERE responsible = ? AND status IN (0, 1, 4, 5, 6, 7) AND is_deleted = 0
            UNION
            SELECT internal_code, title, description, status, NULL as expected_date,
                   NULL as start_date, NULL as end_date, NULL as prod_line, NULL as app_name,
                   NULL as reason, NULL as observations, NULL as time,
                   NULL as responsible, requester, 'software_internal_reports' as table_name, NULL as observations_dt
            FROM software_internal_reports
            WHERE responsible = ? AND status IN (0, 1, 4, 5, 6, 7) AND is_deleted = 0
        """, (username, username, username))
        software_tickets = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        # Fetch Tickets for Automation - apenas pending tasks do usuário logado
        cursor.execute("""
            SELECT internal_code, title, action_to_improve as description, status, expected_date,
                   start_date, end_date, prod_line, n_sap, reason, current_process,
                   observations_requester, observations_dt, time, responsible, requester,
                   'automation_support' as table_name
            FROM automation_support
            WHERE responsible = ? AND status IN (0, 1, 4, 5, 6, 7) AND is_deleted = 0
            UNION
            SELECT internal_code, title, action_to_improve as description, status, expected_date,
                   start_date, end_date, prod_line, n_sap, reason, current_process,
                   observations, NULL as observations_dt, time, responsible, requester,
                   'automation_improvement' as table_name
            FROM automation_improvement
            WHERE responsible = ? AND status IN (0, 1, 4, 5, 6, 7) AND is_deleted = 0
        """, (username, username))
        automation_tickets = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        cursor.close()

        # Fetch Projects for dropdown - only projects where the user is responsible
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, name, description, category, created_by, created_at, updated_at, responsible
            FROM [DT_request].[dbo].[projects]
            WHERE responsible LIKE ?
            ORDER BY name
        """, (f'%{username}%',))
        projects_raw = [row_to_dict(cursor, row) for row in cursor.fetchall()]
        logging.info(f"[DEBUG] Found {len(projects_raw)} raw projects matching username '{username}'")
        
        # Filter projects where the username is actually in the responsible JSON/list
        projects = []
        for proj in projects_raw:
            try:
                responsible_data = proj.get('responsible', '')
                if not responsible_data:
                    logging.info(f"[DEBUG] Project {proj.get('name')} (ID {proj.get('id')}): no responsible data")
                    continue
                
                # Try to parse as JSON
                responsible_list = []
                try:
                    responsible_list = json.loads(responsible_data)
                except json.JSONDecodeError:
                    # If not valid JSON, try parsing as comma-separated string
                    logging.info(f"[DEBUG] Project {proj.get('name')}: parsing as comma-separated (not JSON)")
                    responsible_list = [r.strip() for r in responsible_data.split(',') if r.strip()]
                
                # Ensure it's a list
                if not isinstance(responsible_list, list):
                    responsible_list = [responsible_list]
                
                # Check if the user's username is in the list
                if username in responsible_list:
                    logging.info(f"[DEBUG] Project {proj.get('name')} (ID {proj.get('id')}): user '{username}' found in responsible list")
                    projects.append(proj)
                else:
                    logging.info(f"[DEBUG] Project {proj.get('name')} (ID {proj.get('id')}): user '{username}' NOT in responsible list {responsible_list}")
            except Exception as e:
                logging.error(f"Error parsing responsible for project {proj.get('id')}: {str(e)}")
                continue
        
        logging.info(f"[DEBUG] Filtered to {len(projects)} projects for user '{username}'")
        # Fetch task counts per ticket
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ticket_internal_code, COUNT(*) as task_count
            FROM [DT_request].[dbo].[tasks]
            WHERE ticket_internal_code IS NOT NULL
            GROUP BY ticket_internal_code
        """)
        task_counts = {row[0]: row[1] for row in cursor.fetchall()}
        cursor.close()

        conn.close()

        # Combine all tickets
        tickets = software_tickets + automation_tickets

        # Format dates for JSON
        for task in software_tasks + automation_tasks:
            task['start_date'] = task['start_date'].strftime('%Y-%m-%d') if task['start_date'] else None
            task['end_date'] = task['end_date'].strftime('%Y-%m-%d') if task['end_date'] else None
            task['planned_start_date'] = task['planned_start_date'].strftime('%Y-%m-%d') if task['planned_start_date'] else None
            task['planned_end_date'] = task['planned_end_date'].strftime('%Y-%m-%d') if task['planned_end_date'] else None
            task['estimated_hours'] = float(task['estimated_hours']) if task['estimated_hours'] is not None else None
            task['time_spent'] = float(task['time_spent']) if task['time_spent'] is not None else None

        for ticket in tickets:
            ticket['start_date'] = ticket['start_date'].strftime('%Y-%m-%d') if ticket['start_date'] else None
            ticket['end_date'] = ticket['end_date'].strftime('%Y-%m-%d') if ticket['end_date'] else None
            ticket['expected_date'] = ticket['expected_date'].strftime('%Y-%m-%d') if ticket['expected_date'] else None
            ticket['category'] = 'Software' if ticket['table_name'] in ['software_issue', 'new_application', 'software_internal_reports'] else 'Automation'
            ticket['time'] = ticket['time'] if ticket['time'] is not None else 0
            ticket['task_count'] = task_counts.get(ticket['internal_code'], 0)

        return render_template('tasks.html',
                             software_tasks=software_tasks,
                             automation_tasks=automation_tasks,
                             tickets=tickets,
                             projects=projects,
                             user_category=user_category)
    except Exception as e:
        print(f"ERROR in tasks route: {str(e)}")
        flash(f"Error fetching tasks or tickets: {str(e)}", 'error')
        return redirect(url_for('index'))


@app.route('/profile')
def profile():
    username = session.get('username')
    if not username:
        return redirect(url_for('index'))

    # Verificar se o utilizador tem categoria "Software"
    user_category = session.get('category', '').lower()
    if user_category != 'software':
        flash('Access denied. Profile is only available for Software team members.', 'error')
        return redirect(url_for('index'))

    conn = connect()
    cursor = conn.cursor()
    cursor.execute("SELECT username, name, email, github_username FROM users WHERE username = ?", (username,))
    row = cursor.fetchone()
    if not row:
        cursor.close()
        conn.close()
        flash('User not found', 'error')
        return redirect(url_for('index'))
    user_data = {
        'username': row[0],
        'name': row[1],
        'email': row[2],
        'github_username': row[3]
    }
    cursor.close()
    conn.close()
    return render_template('profile.html', **user_data)


@app.route('/api/github/repos')
def list_github_repos():
    """List repositórios do utilizador (autenticado via token)"""
    token = os.getenv('GITHUB_TOKEN')
    logging.info(f"[DEBUG] GITHUB_TOKEN is set: {'YES' if token else 'NO'}")
    if token:
        logging.info(f"[DEBUG] Token prefix: {token[:10]}... length: {len(token)}")
    else:
        logging.error("[DEBUG] GITHUB_TOKEN is NOT set!")
        return jsonify({'success': False, 'message': 'GitHub token not configured on server'}), 500

    import requests
    url = "https://api.github.com/user/repos"
    headers = {"Authorization": f"token {token}"}
    params = {"per_page": 100}

    try:
        logging.info(f"[DEBUG] Requesting {url} with params {params}")
        resp = requests.get(url, headers=headers, params=params, timeout=10)
        logging.info(f"[DEBUG] GitHub API responded: {resp.status_code}")
        if resp.status_code == 200:
            repos = resp.json()
            simplified = [
                {
                    'full_name': r['full_name'],
                    'name': r['name'],
                    'owner': r['owner']['login'],
                    'private': r['private']
                }
                for r in repos
            ]
            logging.info(f"[DEBUG] Returning {len(simplified)} repos")
            return jsonify({'success': True, 'repos': simplified})
        else:
            logging.error(f"GitHub API error listing repos: {resp.status_code} - {resp.text}")
            return jsonify({'success': False, 'message': f'GitHub API error: {resp.status_code}', 'details': resp.text}), resp.status_code
    except requests.exceptions.Timeout:
        logging.error("Timeout connecting to GitHub API")
        return jsonify({'success': False, 'message': 'Timeout connecting to GitHub'}), 504
    except requests.exceptions.ConnectionError as e:
        logging.error(f"Connection error: {str(e)}")
        return jsonify({'success': False, 'message': 'Cannot connect to GitHub API'}), 503
    except Exception as e:
        logging.error(f"Error listing GitHub repos: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/export/completed_automation_tickets')
def export_completed_automation_tickets():
    """
    Export completed tickets to Excel (.xlsx) with proper formatting.
    Uses openpyxl for better control over column widths and text wrapping.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import re
        from datetime import datetime
        from io import BytesIO

        # Get filter parameters from query string
        date_ini = request.args.get('date_ini')
        date_fim = request.args.get('date_fim')
        category_filter = request.args.get('category')
        responsible_filter = request.args.get('responsible')

        # Função para normalizar cada row para as colunas do CSV
        def normalize_row(row):
            t = row.get('type', '')
            norm = {}
            norm['type'] = t
            norm['internal_code'] = row.get('internal_code', '')
            norm['title'] = row.get('title', '')
            norm['requester'] = row.get('requester', '')
            # Campo 'line_dept_app' varia por categoria
            if t in ['Automation Improvement', 'Automation Support']:
                norm['line_dept_app'] = row.get('prod_line', '')
            elif t == 'New Platform Software Development':
                norm['line_dept_app'] = row.get('department', '')
            elif t == 'Platform Software Improvement':
                norm['line_dept_app'] = row.get('app_name', '')
            elif t == 'Software Internal Report':
                norm['line_dept_app'] = ''
            else:
                norm['line_dept_app'] = ''
            # Campos comuns
            norm['reason'] = row.get('reason', '')
            norm['current_process'] = row.get('current_process', '')
            # action_to_improve varia
            if t == 'Platform Software Improvement':
                norm['action_to_improve'] = row.get('issue', '')
            else:
                norm['action_to_improve'] = row.get('action_to_improve', '')
            norm['start_date'] = row.get('start_date', '')
            norm['end_date'] = row.get('end_date', '')
            norm['time'] = row.get('time', '')
            norm['responsible'] = row.get('responsible', '')
            norm['observations'] = row.get('observations', '')
            norm['notes'] = row.get('notes', '')
            norm['hard_savings'] = row.get('hard_savings', '')
            norm['status'] = row.get('status', '')
            return norm

        # Função para limpar texto (remover quebras de linha e tabs)
        def clean_text(value, max_length=None):
            if value is None:
                return ''
            value = str(value)
            value = re.sub(r'[\r\n\t]+', ' ', value)
            value = re.sub(r' +', ' ', value)
            value = value.strip()
            if max_length and len(value) > max_length:
                value = value[:max_length]
            return value

        conn = connect()
        cursor = conn.cursor()

        results = []

        # 1. Buscar todos os completed requests usando o mesmo stored procedure da página
        try:
            cursor.execute('Exec GetAllCompletedRequests')
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            all_requests = [dict(zip(columns, row)) for row in rows]

            # Adicionar campo 'type' baseado no internal_code
            for r in all_requests:
                code = r.get('internal_code', '')
                if code.startswith('DTAI'):
                    r['type'] = 'Automation Improvement'
                elif code.startswith('DTAS'):
                    r['type'] = 'Automation Support'
                elif code.startswith('DTNA'):
                    r['type'] = 'New Platform Software Development'
                elif code.startswith('DTSI'):
                    r['type'] = 'Platform Software Improvement'
                else:
                    r['type'] = 'Other'
            results.extend(all_requests)
        except Exception as e:
            logging.error(f"Error fetching GetAllCompletedRequests: {str(e)}")

        # 2. Buscar DTIR separadamente (status = 2)
        try:
            cursor.execute("""
                SELECT
                    internal_code, title, requester, reporter, start_date, end_date,
                    time, responsible, observations, status
                FROM software_internal_reports
                WHERE status = 2 AND is_deleted = 0
            """)
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            dtir_requests = [dict(zip(columns, row)) for row in rows]
            for r in dtir_requests:
                r['type'] = 'Software Internal Report'
            results.extend(dtir_requests)
        except Exception as e:
            logging.error(f"Error fetching DTIR: {str(e)}")

        # Filtrar apenas tickets com status = 2 (Done/Completed)
        results = [r for r in results if r.get('status') == 2]

        cursor.close()
        conn.close()

        # Normalizar todas as rows para terem as colunas padrão
        normalized_results = [normalize_row(r) for r in results]

        # Apply filters if any
        if date_ini or date_fim or category_filter or responsible_filter:
            filtered_results = []
            ini_date = None
            fim_date = None
            if date_ini:
                try:
                    ini_date = datetime.strptime(date_ini, '%Y-%m-%d').date()
                except Exception as e:
                    logging.warning(f"Invalid date_ini format: {date_ini}: {e}")
            if date_fim:
                try:
                    fim_date = datetime.strptime(date_fim, '%Y-%m-%d').date()
                except Exception as e:
                    logging.warning(f"Invalid date_fim format: {date_fim}: {e}")

            type_map = {
                'DTAI': 'Automation Improvement',
                'DTAS': 'Automation Support',
                'DTNA': 'New Platform Software Development',
                'DTSI': 'Platform Software Improvement',
                'DTIR': 'Software Internal Report'
            }

            for r in normalized_results:
                # Category filter
                if category_filter:
                    if category_filter in type_map:
                        if r.get('type') != type_map[category_filter]:
                            continue
                    else:
                        # Unsupported category for export, skip all
                        continue
                # Date filter on start_date
                start_date_val = r.get('start_date')
                if start_date_val is None or start_date_val == '':
                    if ini_date or fim_date:
                        continue
                else:
                    # Convert to date if it's datetime or string
                    try:
                        if hasattr(start_date_val, 'date'):
                            start_date_val_date = start_date_val.date()
                        else:
                            # assume string in format YYYY-MM-DD
                            start_date_val_date = datetime.strptime(str(start_date_val), '%Y-%m-%d').date()
                    except Exception:
                        # if parsing fails, treat as no date
                        if ini_date or fim_date:
                            continue
                        start_date_val_date = None
                    if start_date_val_date:
                        if ini_date and start_date_val_date < ini_date:
                            continue
                        if fim_date and start_date_val_date > fim_date:
                            continue
                # Responsible filter
                if responsible_filter:
                    if r.get('responsible') != responsible_filter:
                        continue
                filtered_results.append(r)

            normalized_results = filtered_results

        if not normalized_results:
            return jsonify({'success': False, 'message': 'No completed tickets found'}), 404

        # Colunas na ordem desejada - correspondem exatamente ao que a página mostra
        csv_columns = [
            'type', 'internal_code', 'title', 'requester', 'line_dept_app',
            'reason', 'current_process', 'action_to_improve',
            'start_date', 'end_date', 'time', 'responsible', 'observations',
            'notes', 'hard_savings', 'status'
        ]

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Completed Tickets"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Escrever cabeçalho
        for col_idx, col_name in enumerate(csv_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Escrever dados
        for row_idx, r in enumerate(normalized_results, 2):
            for col_idx, col in enumerate(csv_columns, 1):
                value = r.get(col)
                if value is None:
                    value = ''
                # Converter datas
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')
                elif isinstance(value, (int, float)):
                    value = str(value)
                # Limpar texto
                value = clean_text(value, max_length=5000)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Ajustar largura das colunas
        for col_idx, col_name in enumerate(csv_columns, 1):
            column_letter = get_column_letter(col_idx)
            # Largura baseada no conteúdo
            max_length = len(col_name)
            for row_idx in range(2, len(results) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))
            # Ajustar largura (mínimo 10, máximo 50)
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

        # Congelar primeira linha (cabeçalho)
        ws.freeze_panes = 'A2'

        # Guardar ficheiro em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"completed_automation_tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        logging.error(f"Error exporting completed automation tickets: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/add_task', methods=['POST'])
def add_task():
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'No user logged in'}), 401

        data = request.get_json()
        category = data.get('category')
        week_number = data.get('week_number', '')
        title = data.get('title', 'New Task')
        description = data.get('description', '')
        equipment = data.get('equipment', '') or ''
        equipment_responsible = data.get('equipment_responsible', '') or ''
        time_spent = data.get('time_spent', 0)
        estimated_hours = data.get('estimated_hours', None)
        status = data.get('status', 'To Do')
        start_date = data.get('start_date', None)
        end_date = data.get('end_date', None)
        planned_start_date = data.get('planned_start_date', None)
        planned_end_date = data.get('planned_end_date', None)
        ticket_internal_code = data.get('ticket_internal_code', '')
        ticket_table = data.get('ticket_table', '')
        project_id = data.get('project_id', None)
        task_type = data.get('task_type', None)

        responsible = data.get('responsible', [username])
        responsible_str = format_responsible(responsible)
        priority = 'Medium'

        # Validate required fields
        if not category or category not in ['Software', 'Automation']:
            return jsonify({'success': False, 'message': 'Invalid or missing category'}), 400
        if status not in ['To Do', 'In Progress', 'Done', 'Standby']:
            return jsonify({'success': False, 'message': 'Invalid status'}), 400

        # Validate and parse dates
        start_date_obj = None
        end_date_obj = None
        planned_start_date_obj = None
        planned_end_date_obj = None
        try:
            if start_date:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            if end_date:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            if planned_start_date:
                planned_start_date_obj = datetime.strptime(planned_start_date, '%Y-%m-%d').date()
            if planned_end_date:
                planned_end_date_obj = datetime.strptime(planned_end_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format'}), 400

        # Validate lengths
        if len(description) > 2000:
            return jsonify({'success': False, 'message': 'Description exceeds 2000 characters'}), 400
        if len(title) > 255:
            return jsonify({'success': False, 'message': 'Title exceeds 255 characters'}), 400
        if len(equipment) > 100:
            return jsonify({'success': False, 'message': 'Equipment exceeds 100 characters'}), 400

        # Parse week_number
        week_num = None
        if week_number and str(week_number).strip():
            try:
                week_num = int(week_number)
                if week_num < 1 or week_num > 53:
                    week_num = None
            except (ValueError, TypeError):
                week_num = None
        else:
            if start_date_obj or planned_start_date_obj:
                week_num = (start_date_obj or planned_start_date_obj).isocalendar()[1]

        conn = connect()
        cursor = conn.cursor()

        # Determinar se é principal task
        is_principal = 0
        principal_task_id = None
        
        if ticket_internal_code:
            # Verificar se já existem tasks para este ticket
            cursor.execute("""
                SELECT id, is_principal_task FROM tasks
                WHERE ticket_internal_code = ?
                ORDER BY created_at ASC
            """, (ticket_internal_code,))
            existing_tasks = cursor.fetchall()

            if not existing_tasks:
                # É a primeira task para este ticket
                is_principal = 1
            else:
                # Encontrar a principal task existente
                principal_task = None
                for existing_task in existing_tasks:
                    if existing_task[1] == 1:  # is_principal_task = 1
                        principal_task = existing_task
                        break

                if principal_task:
                    # Já existe uma principal, nova task é subtarefa
                    principal_task_id = principal_task[0]
                    is_principal = 0
                else:
                    # Nenhuma principal existe -> a primeira task existente deve ser promovida a principal
                    # Primeiro, atualizar a primeira task para is_principal_task = 1
                    first_task_id = existing_tasks[0][0]
                    cursor.execute("""
                        UPDATE tasks SET is_principal_task = 1
                        WHERE id = ?
                    """, (first_task_id,))
                    conn.commit()
                    # A nova task será uma subtarefa apontando para a primeira (agora principal)
                    principal_task_id = first_task_id
                    is_principal = 0
        elif project_id:
            # Para projetos, aplicar mesma lógica
            cursor.execute("""
                SELECT id, is_principal_task FROM tasks
                WHERE project_id = ? AND principal_task_id IS NULL
                ORDER BY created_at ASC
            """, (project_id,))
            existing_tasks = cursor.fetchall()

            if not existing_tasks:
                # É a primeira task para este projeto
                is_principal = 1
            else:
                # Encontrar a principal task existente
                principal_task = None
                for existing_task in existing_tasks:
                    if existing_task[1] == 1:  # is_principal_task = 1
                        principal_task = existing_task
                        break

                if principal_task:
                    # Já existe uma principal, nova task é subtarefa
                    principal_task_id = principal_task[0]
                    is_principal = 0
                else:
                    # Nenhuma principal existe -> a primeira task existente deve ser promovida a principal
                    first_task_id = existing_tasks[0][0]
                    cursor.execute("""
                        UPDATE tasks SET is_principal_task = 1
                        WHERE id = ?
                    """, (first_task_id,))
                    conn.commit()
                    # A nova task será uma subtarefa apontando para a primeira (agora principal)
                    principal_task_id = first_task_id
                    is_principal = 0

        # Query corrigida com schema completo e apenas as colunas que existem
        query = """
            INSERT INTO [DT_request].[dbo].[tasks] (
                category, week_number, title, description, responsible, equipment,
                equipment_responsible, priority, status, start_date, end_date, 
                planned_start_date, planned_end_date, ticket_internal_code, ticket_table, 
                time_spent, estimated_hours, project_id, task_type, created_by, created_at, updated_at,
                is_principal_task, principal_task_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE(), ?, ?)
        """
        params = (
            category,
            week_num,
            title,
            description,
            responsible_str,  # Use formatted string with semicolon separator
            equipment,
            equipment_responsible,
            priority,
            status,
            start_date_obj,
            end_date_obj,
            planned_start_date_obj,
            planned_end_date_obj,
            ticket_internal_code if ticket_internal_code else None,
            ticket_table if ticket_table else None,
            time_spent,
            estimated_hours,
            project_id if project_id else None,
            task_type,
            username,
            is_principal,
            principal_task_id
        )

        cursor.execute(query, params)

        # Get the inserted task ID
        cursor.execute("SELECT @@IDENTITY AS id")
        task_id = cursor.fetchone().id

        issue_number = None

        # Se task está associada a projeto com github_repo, criar Issue obrigatoriamente
        if project_id:
            try:
                github_repo = get_project_github_repo(cursor, project_id)
                logging.info(f"[DEBUG] add_task - Task {task_id} - Project {project_id} - github_repo: {github_repo}")

                if github_repo:
                    cursor.execute("SELECT responsible FROM projects WHERE id = ?", (project_id,))
                    proj_resp = cursor.fetchone()
                    responsible_json = proj_resp[0] if proj_resp else None

                    assignees = []
                    if responsible_json:
                        assignees = get_user_github_usernames(responsible_json, cursor)

                    issue_number = create_issue(
                        repo=github_repo,
                        title=title,
                        description=description,
                        assignees=assignees
                    )

                    if issue_number:
                        cursor.execute("UPDATE tasks SET github_issue_number = ? WHERE id = ?", (issue_number, task_id))
                        logging.info(f"GitHub Issue #{issue_number} criada automaticamente para tarefa {task_id}")
                    else:
                        conn.rollback()
                        cursor.close()
                        conn.close()
                        return jsonify({
                            'success': False,
                            'message': 'Task not created because GitHub Issue creation failed for project repository'
                        }), 500
            except Exception as e:
                conn.rollback()
                cursor.close()
                conn.close()
                logging.error(f"Erro ao criar GitHub Issue automática para tarefa {task_id}: {str(e)}", exc_info=True)
                return jsonify({
                    'success': False,
                    'message': 'Task not created because GitHub Issue creation failed for project repository'
                }), 500

        conn.commit()

        # Fetch the complete task data to return to frontend
        cursor.execute("""
            SELECT id, category, week_number, title, description, responsible, equipment,
                   equipment_responsible, priority, status, start_date, end_date,
                   planned_start_date, planned_end_date, ticket_internal_code, ticket_table,
                   time_spent, estimated_hours, project_id, task_type, is_principal_task,
                   principal_task_id, created_by, created_at, updated_at
            FROM tasks
            WHERE id = ?
        """, (task_id,))
        row = cursor.fetchone()
        columns = [desc[0] for desc in cursor.description]
        task_dict = dict(zip(columns, row))

        # Convert date objects to strings
        date_fields = ['start_date', 'end_date', 'planned_start_date', 'planned_end_date', 'created_at', 'updated_at']
        for field in date_fields:
            if task_dict.get(field):
                task_dict[field] = task_dict[field].strftime('%Y-%m-%d') if hasattr(task_dict[field], 'strftime') else str(task_dict[field])

        cursor.close()
        conn.close()

        return jsonify({'success': True, 'task': task_dict})

    except pyodbc.Error as e:
        print(f"Database error in add_task: {str(e)}")
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500
    except Exception as e:
        print(f"General error in add_task: {str(e)}")
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500

@app.route('/update_task/<int:task_id>', methods=['POST'])
def update_task(task_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify(success=False, message="Not authenticated"), 401

        data = request.get_json()
        if not data:
            return jsonify(success=False, message="No data provided"), 400

        fields = [
            'week_number', 'title', 'description', 'responsible', 'priority', 'status',
            'start_date', 'end_date', 'planned_start_date', 'planned_end_date', 
            'time_spent', 'equipment', 'equipment_responsible', 'estimated_hours'
        ]
        updates = {field: data.get(field) for field in fields}

        for date_field in ['start_date', 'end_date', 'planned_start_date', 'planned_end_date']:
            if updates[date_field]:
                try:
                    updates[date_field] = datetime.strptime(updates[date_field], '%Y-%m-%d').date()
                except ValueError:
                    return jsonify(success=False, message=f"Invalid date format for {date_field}"), 400
            else:
                updates[date_field] = None

        # Format responsible to semicolon-separated string for multi-select support
        if updates['responsible'] is not None:
            updates['responsible'] = format_responsible(updates['responsible'])

        conn = connect()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT ticket_internal_code, ticket_table, status, is_principal_task, principal_task_id, project_id
            FROM tasks
            WHERE id = ?
        """, (task_id,))
        task_info = cursor.fetchone()
        
        if not task_info:
            return jsonify(success=False, message="Task not found"), 404

        current_ticket_code = task_info[0]
        current_ticket_table = task_info[1]
        old_status = task_info[2]
        is_principal_task = task_info[3]
        principal_task_id_val = task_info[4]
        project_id_val = task_info[5]
        new_status = updates['status']

        cursor.execute("""
            UPDATE tasks
            SET week_number = ?, title = ?, description = ?, responsible = ?, priority = ?, 
                status = ?, start_date = ?, end_date = ?, planned_start_date = ?, 
                planned_end_date = ?, time_spent = ?, equipment = ?, 
                equipment_responsible = ?, estimated_hours = ?, updated_at = GETDATE()
            WHERE id = ?
        """, (
            updates['week_number'], updates['title'], updates['description'], updates['responsible'],
            updates['priority'], updates['status'], updates['start_date'], updates['end_date'],
            updates['planned_start_date'], updates['planned_end_date'], updates['time_spent'],
            updates['equipment'], updates['equipment_responsible'], updates['estimated_hours'], task_id
        ))

        if cursor.rowcount == 0:
            cursor.close()
            conn.close()
            return jsonify(success=False, message="Task not found or no changes made"), 404

        # Se há ticket associado, a tarefa é principal e o status mudou, atualizar o status do ticket e enviar email
        if current_ticket_code and current_ticket_table and old_status != new_status and is_principal_task == 1:
            status_mapping = {
                'To Do': 0,
                'Standby': 0,
                'In Progress': 1,
                'Done': 2
            }
            ticket_status = status_mapping.get(new_status)
            if ticket_status is not None:
                # Determinar a tabela correta baseada no código do ticket
                if current_ticket_code.startswith('DTAI'):
                    table_name = 'automation_improvement'
                elif current_ticket_code.startswith('DTAS'):
                    table_name = 'automation_support'
                elif current_ticket_code.startswith('DTNA'):
                    table_name = 'new_application'
                elif current_ticket_code.startswith('DTSI'):
                    table_name = 'software_issue'
                elif current_ticket_code.startswith('DTIR'):
                    table_name = 'software_internal_reports'
                else:
                    table_name = None
                
                if table_name:
                    try:
                        cursor.execute(f"""
                            UPDATE {table_name}
                            SET status = ?
                            WHERE internal_code = ? AND approved = 1
                        """, (ticket_status, current_ticket_code))
                        # Buscar nome do responsável
                        cursor.execute("SELECT name FROM users WHERE username = ?", (updates['responsible'],))
                        row_resp = cursor.fetchone()
                        responsible_name = row_resp[0] if row_resp else None

                        # Buscar expected_date e notes
                        cursor.execute(f"SELECT expected_date, notes FROM {table_name} WHERE internal_code = ?", (current_ticket_code,))
                        row_date_notes = cursor.fetchone()
                        expected_date_fmt = ''
                        notes = ''
                        if row_date_notes:
                            if row_date_notes[0]:
                                try:
                                    expected_date_fmt = row_date_notes[0].strftime("%d/%m/%Y")
                                except Exception:
                                    expected_date_fmt = str(row_date_notes[0])
                            notes = row_date_notes[1] if row_date_notes[1] else ''

                        # Buscar email do requester
                        cursor.execute(f"""
                            SELECT u.email, u.name, t.title
                            FROM users u
                            JOIN {table_name} t ON (u.name = t.requester OR u.username = t.requester)
                            WHERE t.internal_code = ?
                        """, (current_ticket_code,))
                        row = cursor.fetchone()
                        if row:
                            to_email, to_name, ticket_title = row
                            # Mapear status da tarefa para status de email
                            if new_status == 'In Progress':
                                email_status = 'in_progress'
                            elif new_status == 'Done':
                                email_status = 'completed'
                            else:
                                email_status = 'approved'
                            send_ticket_status_email(
                                to_email, to_name, current_ticket_code, ticket_title, email_status,
                                responsible=responsible_name, expected_date=expected_date_fmt, notes=notes
                            )
                    except Exception as e:
                        print(f"Error updating ticket status or sending email: {str(e)}")

        conn.commit()

        # Se a tarefa é principal e o status mudou para 'Done', verificar se há Issue para fechar
        if is_principal_task == 1 and new_status == 'Done' and old_status != 'Done':
            cursor.execute("SELECT github_issue_number FROM tasks WHERE id = ?", (task_id,))
            task_row = cursor.fetchone()
            issue_number = task_row[0] if task_row else None

            if issue_number:
                # Buscar github_repo do projeto associado
                cursor.execute("""
                    SELECT p.github_repo
                    FROM projects p
                    WHERE p.id = (SELECT project_id FROM tasks WHERE id = ?)
                """, (task_id,))
                repo_row = cursor.fetchone()
                github_repo = repo_row[0] if repo_row else None

                if github_repo:
                    try:
                        close_issue(github_repo, issue_number)
                        logging.info(f"GitHub Issue #{issue_number} fechada (tarefa {task_id} concluída)")
                    except Exception as e:
                        logging.error(f"Erro ao fechar GitHub Issue #{issue_number}: {str(e)}")

        cursor.close()
        conn.close()

        # Se é principal task e mudou para Done, retornar as sub-tasks e ticket info
        response_data = {'success': True}
        if is_principal_task and new_status == 'Done' and old_status != 'Done':
            # Buscar todas as sub-tasks
            conn = connect()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, title, status FROM tasks 
                WHERE principal_task_id = ? OR (is_principal_task = 1 AND id = ?)
                ORDER BY created_at ASC
            """, (task_id, task_id))
            rows = cursor.fetchall()
            sub_tasks = [{'id': row[0], 'title': row[1], 'status': row[2]} for row in rows]
            cursor.close()
            conn.close()
            
            response_data['is_principal_task'] = True
            response_data['sub_tasks'] = sub_tasks
            response_data['ticket_code'] = current_ticket_code
            response_data['ticket_table'] = current_ticket_table

        return jsonify(response_data)
    except Exception as e:
        print(f"Error in update_task: {str(e)}")
        if 'conn' in locals():
            conn.rollback()
            cursor.close()
            conn.close()
        return jsonify(success=False, message=f"Unexpected error: {str(e)}"), 500
    
@app.route('/get_principal_task_details/<int:task_id>', methods=['GET'])
def get_principal_task_details(task_id):
    """Get principal task and all its sub-tasks"""
    try:
        conn = connect()
        cursor = conn.cursor()
        
        # Get principal task info
        cursor.execute("""
            SELECT id, title, status, ticket_internal_code, ticket_table, project_id
            FROM tasks 
            WHERE id = ? AND is_principal_task = 1
        """, (task_id,))
        principal = cursor.fetchone()
        
        if not principal:
            return jsonify({'success': False, 'error': 'Principal task not found'}), 404
        
        principal_data = {
            'id': principal[0],
            'title': principal[1],
            'status': principal[2],
            'ticket_code': principal[3],
            'ticket_table': principal[4],
            'project_id': principal[5]
        }
        
        # Get all sub-tasks
        cursor.execute("""
            SELECT id, title, status, responsible
            FROM tasks 
            WHERE principal_task_id = ? OR (is_principal_task = 1 AND id = ?)
            ORDER BY created_at ASC
        """, (task_id, task_id))
        
        rows = cursor.fetchall()
        sub_tasks = [{'id': row[0], 'title': row[1], 'status': row[2], 'responsible': row[3]} for row in rows]
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'principal_task': principal_data,
            'sub_tasks': sub_tasks
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/complete_principal_task/<int:task_id>', methods=['POST'])
def complete_principal_task(task_id):
    """Complete a principal task and all its sub-tasks, and close the ticket if applicable"""
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        data = request.get_json()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Parse dates
        start_date_obj = None
        end_date_obj = None
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        conn = connect()
        cursor = conn.cursor()
        
        # Get principal task info
        cursor.execute("""
            SELECT id, title, status, ticket_internal_code, ticket_table, project_id
            FROM tasks 
            WHERE id = ? AND is_principal_task = 1
        """, (task_id,))
        principal = cursor.fetchone()
        
        if not principal:
            conn.close()
            return jsonify({'success': False, 'message': 'Principal task not found'}), 404
        
        principal_id = principal[0]
        ticket_code = principal[3]
        ticket_table = principal[4]
        project_id_val = principal[5]
        
        # Update principal task to Done
        cursor.execute("""
            UPDATE tasks 
            SET status = 'Done', start_date = ?, end_date = ?, updated_at = GETDATE()
            WHERE id = ?
        """, (start_date_obj, end_date_obj, task_id))
        
        # Update all sub-tasks to Done
        cursor.execute("""
            UPDATE tasks 
            SET status = 'Done', start_date = ?, end_date = ?, updated_at = GETDATE()
            WHERE principal_task_id = ? OR (id = ? AND is_principal_task = 1)
        """, (start_date_obj, end_date_obj, task_id, task_id))
        
        # If there's a ticket, update its status to Done (2)
        if ticket_code and ticket_table:
            # Determine correct table name
            if ticket_code.startswith('DTAI'):
                table_name = 'automation_improvement'
            elif ticket_code.startswith('DTAS'):
                table_name = 'automation_support'
            elif ticket_code.startswith('DTNA'):
                table_name = 'new_application'
            elif ticket_code.startswith('DTSI'):
                table_name = 'software_issue'
            else:
                table_name = None
            
            if table_name:
                try:
                    cursor.execute(f"""
                        UPDATE {table_name}
                        SET status = 2
                        WHERE internal_code = ? AND approved = 1
                    """, (ticket_code,))
                except Exception as e:
                    logging.error(f"Error updating ticket status: {str(e)}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Principal task and all sub-tasks completed successfully'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/delete_task/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify(success=False, message="Not authenticated"), 401

        conn = connect()
        cursor = conn.cursor()

        # Fetch task to get category (no responsible check)
        cursor.execute("""
            SELECT category
            FROM [DT_request].[dbo].[tasks]
            WHERE id = ?
        """, (task_id,))
        
        task = cursor.fetchone()
        if not task:
            cursor.close()
            conn.close()
            return jsonify(success=False, message=f"Task with ID {task_id} not found"), 404

        category = task[0]

        # Delete the task
        cursor.execute("""
            DELETE FROM [DT_request].[dbo].[tasks]
            WHERE id = ?
        """, (task_id,))

        if cursor.rowcount == 0:
            cursor.close()
            conn.close()
            return jsonify(success=False, message=f"Task with ID {task_id} not found"), 404

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify(success=True, category=category)

    except pyodbc.Error as e:
        if 'conn' in locals():
            conn.rollback()
            cursor.close()
            conn.close()
        return jsonify(success=False, message=f"Database error: {str(e)}"), 500
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            cursor.close()
            conn.close()
        return jsonify(success=False, message=f"Unexpected error: {str(e)}"), 500
    
@app.route('/get_task/<int:task_id>')
def get_task(task_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify(success=False, message="Not authenticated"), 401

        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, category, title, description, responsible, priority, status, start_date, end_date,
                   week_number, ticket_internal_code, ticket_table, created_at, updated_at, equipment,
                   equipment_responsible, time_spent, planned_start_date, planned_end_date, project_id, 
                   comments, estimated_hours
            FROM [DT_request].[dbo].[tasks]
            WHERE id = ?
        """, (task_id,))
        row = cursor.fetchone()
        if not row:
            cursor.close()
            conn.close()
            return jsonify(success=False, message="Task not found"), 404

        columns = [desc[0] for desc in cursor.description]
        task = dict(zip(columns, row))
        # Format dates
        for f in ['start_date', 'end_date', 'planned_start_date', 'planned_end_date', 'created_at', 'updated_at']:
            if task.get(f): 
                task[f] = task[f].strftime('%Y-%m-%d') if hasattr(task[f], 'strftime') else task[f]
            else:
                task[f] = None
        task['estimated_hours'] = task['estimated_hours'] if task['estimated_hours'] is not None else None

        # Parse responsible to array for multi-select support
        task['responsible'] = parse_responsible(task.get('responsible'))

        cursor.close()
        conn.close()
        return jsonify(success=True, task=task)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/update_task_comment/<int:task_id>', methods=['POST'])
def update_task_comment(task_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify(success=False, message="Not authenticated"), 401
        data = request.get_json()
        comment = data.get('comments', '')
        if len(comment) > 2000:
            return jsonify(success=False, message="Comment too long (max 2000 chars)"), 400

        conn = connect()
        cursor = conn.cursor()
        
        # Primeiro, buscar informações da task, incluindo ticket associado
        cursor.execute("""
            SELECT ticket_internal_code, ticket_table, status, title
            FROM [DT_request].[dbo].[tasks]
            WHERE id = ?
        """, (task_id,))
        task_info = cursor.fetchone()
        
        if not task_info:
            cursor.close()
            conn.close()
            return jsonify(success=False, message="Task not found"), 404
        
        ticket_code, ticket_table, task_status, task_title = task_info
        
        # Atualizar o comment da task
        cursor.execute("""
            UPDATE [DT_request].[dbo].[tasks]
            SET comments = ?, updated_at = GETDATE()
            WHERE id = ?
        """, (comment, task_id))
        
        # Se a task está associada a um ticket e tem status que permite updates
        if ticket_code and ticket_table and task_status in ['To Do', 'Standby', 'In Progress'] and comment.strip():
            try:
                # Mapear nomes de tabelas
                table_mapping = {
                    'automation_improvement': 'automation_improvement',
                    'automation_support': 'automation_support', 
                    'new_application': 'new_application',
                    'software_issue': 'software_issue'
                }
                
                actual_table = table_mapping.get(ticket_table)
                if actual_table:
                    # Atualizar notes do ticket
                    cursor.execute(f"""
                        UPDATE [DT_request].[dbo].[{actual_table}]
                        SET notes = ?, updated_at = GETDATE()
                        WHERE internal_code = ?
                    """, (comment, ticket_code))
                    
                    # Buscar dados do requester, responsible e expected_date para enviar email
                    cursor.execute(f"""
                        SELECT u.email, u.name, t.title, t.responsible, t.expected_date, u2.name as responsible_name
                        FROM users u
                        JOIN {actual_table} t ON (u.name = t.requester OR u.username = t.requester)
                        LEFT JOIN users u2 ON t.responsible = u2.username
                        WHERE t.internal_code = ?
                    """, (ticket_code,))
                    row = cursor.fetchone()
                    
                    if row:
                        to_email, to_name, ticket_title, responsible_username, expected_date, responsible_name = row
                        
                        # Formatar a data se existir
                        formatted_date = None
                        if expected_date:
                            try:
                                formatted_date = expected_date.strftime('%Y-%m-%d') if hasattr(expected_date, 'strftime') else str(expected_date)
                            except:
                                formatted_date = str(expected_date)
                        
                        # Usar o nome do responsible se disponível, senão o username
                        responsible_display = responsible_name if responsible_name else responsible_username
                        
                        # Enviar email de atualização
                        send_ticket_status_email(
                            to_email, to_name, ticket_code, ticket_title, 'in_progress',
                            extra_msg=f"Task Update: {comment}", 
                            notes=comment,
                            responsible=responsible_display,
                            expected_date=formatted_date
                        )
                        print(f"[SYNC] Updated ticket {ticket_code} notes and sent email notification")
            except Exception as e:
                print(f"Error syncing comment to ticket: {str(e)}")
                # Não falhar a operação principal se a sincronização falhar
        
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/update_ticket_status', methods=['POST'])
def update_ticket_status():
    try:
        username = session.get('username')
        if not username:
            return jsonify(success=False, message="Not authenticated"), 401
            
        data = request.get_json()
        ticket_code = data.get('ticket_code')
        ticket_table = data.get('ticket_table')
        status = data.get('status')
        observations = data.get('observations', '')
        time_spent = data.get('time', 0)
        
        if not ticket_code or not ticket_table:
            return jsonify(success=False, message="Missing ticket code or table"), 400
            
        conn = connect()
        cursor = conn.cursor()
        
        # Map table names to actual table names
        table_mapping = {
            'automation_improvement': 'automation_improvement',
            'automation_support': 'automation_support',
            'new_application': 'new_application',
            'software_issue': 'software_issue',
            'software_internal_reports': 'software_internal_reports'
        }

        actual_table = table_mapping.get(ticket_table)
        if not actual_table:
            return jsonify(success=False, message="Invalid ticket table"), 400
            
        # Get additional data from request if available
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Update ticket status and related fields based on table type
        if actual_table == 'automation_support':
            # Para DTAS, observations vai para observations_dt
            if start_date and end_date:
                update_query = f"""
                    UPDATE [DT_request].[dbo].[{actual_table}]
                    SET status = ?, observations_dt = ?, time = ?, start_date = ?, end_date = ?, updated_at = GETDATE()
                    WHERE internal_code = ?
                """
                cursor.execute(update_query, (status, observations, time_spent, start_date, end_date, ticket_code))
            else:
                update_query = f"""
                    UPDATE [DT_request].[dbo].[{actual_table}]
                    SET status = ?, observations_dt = ?, time = ?, updated_at = GETDATE()
                    WHERE internal_code = ?
                """
                cursor.execute(update_query, (status, observations, time_spent, ticket_code))
        else:
            # Para DTAI, DTNA, DTSI, observations vai para observations
            if start_date and end_date:
                update_query = f"""
                    UPDATE [DT_request].[dbo].[{actual_table}]
                    SET status = ?, observations = ?, time = ?, start_date = ?, end_date = ?, updated_at = GETDATE()
                    WHERE internal_code = ?
                """
                cursor.execute(update_query, (status, observations, time_spent, start_date, end_date, ticket_code))
            else:
                update_query = f"""
                    UPDATE [DT_request].[dbo].[{actual_table}]
                    SET status = ?, observations = ?, time = ?, updated_at = GETDATE()
                    WHERE internal_code = ?
                """
                cursor.execute(update_query, (status, observations, time_spent, ticket_code))
        
        # Se o status é 2 (concluído), enviar email de conclusão
        if status == 2:
            try:
                cursor.execute(f"""
                    SELECT u.email, u.name, t.title
                    FROM users u
                    JOIN {actual_table} t ON (u.name = t.requester OR u.username = t.requester)
                    WHERE t.internal_code = ?
                """, (ticket_code,))
                row = cursor.fetchone()
                if row:
                    to_email, to_name, ticket_title = row
                    send_ticket_status_email(to_email, to_name, ticket_code, ticket_title, 'completed', observations)
            except Exception as e:
                print(f"Error sending completion email: {str(e)}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify(success=True)
        
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/get_ticket_data/<internal_code>')
def get_ticket_data(internal_code):
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401

        conn = connect()
        cursor = conn.cursor()
        
        ticket_data = None
        table_name = None
        
        if internal_code.startswith('DTSI'):
            table_name = 'software_issue'
            cursor.execute("""
                SELECT internal_code, title, issue as description, status, expected_date, 
                       start_date, end_date, prod_line, app_name, reason, observations, time,
                       responsible, requester, NULL as subcategory, NULL as [8D_Number],
                       NULL as complaint_number, NULL as hard_savings, NULL as other_description
                FROM [DT_request].[dbo].[software_issue]
                WHERE internal_code = ? AND (requester = ? OR responsible = ?) AND approved = 1
            """, (internal_code, username, username))
            
        elif internal_code.startswith('DTNA'):
            table_name = 'new_application'
            cursor.execute("""
                SELECT internal_code, title, objective as description, status, expected_date,
                       start_date, end_date, department, reason, current_process, 
                       observations, time, responsible, requester, NULL as subcategory,
                       NULL as [8D_Number], NULL as complaint_number, NULL as hard_savings,
                       NULL as other_description
                FROM [DT_request].[dbo].[new_application]
                WHERE internal_code = ? AND (requester = ? OR responsible = ?) AND approved = 1
            """, (internal_code, username, username))
            
        elif internal_code.startswith('DTAS'):
            table_name = 'automation_support'
            cursor.execute("""
                SELECT internal_code, title, action_to_improve as description, status, expected_date,
                       start_date, end_date, prod_line, n_sap, reason, current_process,
                       observations_requester, observations_dt, time, responsible, requester,
                       subcategory, [8D_Number], complaint_number, hard_savings, other_description
                FROM [DT_request].[dbo].[automation_support]
                WHERE internal_code = ? AND (requester = ? OR responsible = ?) AND approved = 1
            """, (internal_code, username, username))
            
        elif internal_code.startswith('DTAI'):
            table_name = 'automation_improvement'
            cursor.execute("""
                SELECT internal_code, title, action_to_improve as description, status, expected_date,
                       start_date, end_date, prod_line, n_sap, reason, current_process,
                       observations, time, responsible, requester, subcategory, [8D_Number],
                       complaint_number, hard_savings, other_description
                FROM [DT_request].[dbo].[automation_improvement]
                WHERE internal_code = ? AND (requester = ? OR responsible = ?) AND approved = 1
            """, (internal_code, username, username))

        elif internal_code.startswith('DTIR'):
            table_name = 'software_internal_reports'
            # Ensure priority column exists in software_internal_reports table
            try:
                cursor.execute("""
                    IF NOT EXISTS (SELECT * FROM sys.columns WHERE Name = N'priority' AND Object_ID = Object_ID(N'software_internal_reports'))
                    BEGIN
                        ALTER TABLE software_internal_reports ADD priority NVARCHAR(20) DEFAULT 'Medium' NULL
                    END
                """)
                conn.commit()
            except Exception as e:
                logging.warning(f"Could not ensure priority column exists: {str(e)}")
            cursor.execute("""
                SELECT internal_code, title, description, status, priority,
                       NULL as expected_date, NULL as start_date, NULL as end_date,
                       NULL as time, reporter as responsible, requester,
                       NULL as subcategory, NULL as [8D_Number],
                       NULL as complaint_number, NULL as hard_savings, NULL as other_description
                FROM software_internal_reports
                WHERE internal_code = ? AND is_deleted = 0
            """, (internal_code,))

        else:
            conn.close()
            return jsonify({'success': False, 'message': 'Invalid ticket code'}), 400

        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'message': 'Ticket not found or not accessible'}), 404

        columns = [desc[0] for desc in cursor.description]
        ticket_dict = dict(zip(columns, row))
        
        week_number = None
        if ticket_dict.get('start_date') or ticket_dict.get('expected_date'):
            date_to_use = ticket_dict.get('start_date') or ticket_dict.get('expected_date')
            if date_to_use:
                if isinstance(date_to_use, datetime):
                    date_to_use = date_to_use.date()
                week_number = date_to_use.isocalendar()[1]

        status_map = {0: 'To Do', 1: 'In Progress', 2: 'Done', None: 'To Do'}
        mapped_status = status_map.get(ticket_dict.get('status'), 'To Do')

        base_description = ticket_dict.get('description', '') or ''
        additional_info = []
        
        if table_name == 'software_issue':
            if ticket_dict.get('app_name'): additional_info.append(f"App: {ticket_dict['app_name']}")
            if ticket_dict.get('prod_line'): additional_info.append(f"Prod Line: {ticket_dict['prod_line']}")
        elif table_name == 'new_application':
            if ticket_dict.get('department'): additional_info.append(f"Department: {ticket_dict['department']}")
            if ticket_dict.get('current_process'): additional_info.append(f"Current Process: {ticket_dict['current_process']}")
        elif table_name in ['automation_support', 'automation_improvement']:
            if ticket_dict.get('prod_line'): additional_info.append(f"Prod Line: {ticket_dict['prod_line']}")
            if ticket_dict.get('n_sap'): additional_info.append(f"SAP: {ticket_dict['n_sap']}")
            if ticket_dict.get('current_process'): additional_info.append(f"Current Process: {ticket_dict['current_process']}")
        
        if ticket_dict.get('reason'): additional_info.append(f"Reason: {ticket_dict['reason']}")
        if ticket_dict.get('time'): additional_info.append(f"Time: {ticket_dict['time']}h")
        
        full_description = base_description
        if additional_info:
            full_description += '\n\n' + '\n'.join(additional_info)
        
        full_description = full_description[:2000] if full_description else ''

        response_data = {
            'internal_code': ticket_dict['internal_code'],
            'title': ticket_dict.get('title', 'New Task'),
            'description': full_description,
            'status': mapped_status,
            'start_date': ticket_dict['start_date'].strftime('%Y-%m-%d') if ticket_dict.get('start_date') else None,
            'end_date': ticket_dict['end_date'].strftime('%Y-%m-%d') if ticket_dict.get('end_date') else None,
            'expected_date': ticket_dict['expected_date'].strftime('%Y-%m-%d') if ticket_dict.get('expected_date') else None,
            'week_number': week_number,
            'table_name': table_name,
            'equipment': ticket_dict.get('n_sap', '') if table_name in ['automation_support', 'automation_improvement'] else '',
            'equipment_responsible': ticket_dict.get('requester', '') if table_name in ['automation_support', 'automation_improvement'] else '',
            'subcategory': ticket_dict.get('subcategory', ''),
            '8D_Number': ticket_dict.get('8D_Number', ''),
            'complaint_number': ticket_dict.get('complaint_number', ''),
            'hard_savings': ticket_dict.get('hard_savings', ''),
            'other_description': ticket_dict.get('other_description', '')
        }

        conn.close()
        return jsonify({'success': True, 'ticket': response_data})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    
@app.route('/update_ticket/<internal_code>', methods=['POST'])
def update_ticket(internal_code):
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401

        data = request.get_json()
        table_name = data.get('table_name')
        responsible = data.get('responsible')
        status = data.get('status')

        status_mapping = {
            'To Do': 0,
            'Standby': 0,
            'In Progress': 1,
            'Done': 2
        }
        ticket_status = status_mapping.get(status)
        if ticket_status is None:
            return jsonify({'success': False, 'message': 'Invalid status'}), 400

        valid_tables = ['software_issue', 'new_application', 'automation_support', 'automation_improvement']
        if table_name not in valid_tables:
            return jsonify({'success': False, 'message': 'Invalid table name'}), 400

        conn = connect()
        cursor = conn.cursor()

        # Buscar status atual antes de atualizar
        cursor.execute(f"SELECT status FROM [DT_request].[dbo].[{table_name}] WHERE internal_code = ?", (internal_code,))
        current_status_row = cursor.fetchone()
        current_status = current_status_row[0] if current_status_row else None

        # Atualizar ticket
        query = f"""
            UPDATE [DT_request].[dbo].[{table_name}]
            SET responsible = ?, status = ?, updated_at = GETDATE()
            WHERE internal_code = ? AND approved = 1
        """
        cursor.execute(query, (responsible, ticket_status, internal_code))

        if cursor.rowcount == 0:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'message': 'Ticket not found or not approved'}), 404

        # Só envia e-mail se o status mudou
        if current_status != ticket_status:
            # Buscar nome do responsável
            responsible_name = None
            if responsible:
                cursor.execute("SELECT name FROM users WHERE username = ?", (responsible,))
                row_resp = cursor.fetchone()
                if row_resp:
                    responsible_name = row_resp[0]

            # Buscar expected_date e notes
            cursor.execute(f"SELECT expected_date, notes FROM [DT_request].[dbo].[{table_name}] WHERE internal_code = ?", (internal_code,))
            row_date_notes = cursor.fetchone()
            expected_date_fmt = ''
            notes = ''
            if row_date_notes:
                if row_date_notes[0]:
                    try:
                        expected_date_fmt = row_date_notes[0].strftime("%d/%m/%Y")
                    except Exception:
                        expected_date_fmt = str(row_date_notes[0])
                notes = row_date_notes[1] if row_date_notes[1] else ''

            # Buscar email do requester
            cursor.execute(f"""
                SELECT u.email, u.name, t.title
                FROM users u
                JOIN [DT_request].[dbo].[{table_name}] t ON (u.name = t.requester OR u.username = t.requester)
                WHERE t.internal_code = ?
            """, (internal_code,))
            row = cursor.fetchone()
            if row:
                to_email, to_name, ticket_title = row
                email_status = 'in_progress' if status == 'In Progress' else 'approved'
                send_ticket_status_email(
                    to_email, to_name, internal_code, ticket_title, email_status,
                    responsible=responsible_name, expected_date=expected_date_fmt, notes=notes
                )

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({'success': True})

    except pyodbc.Error as e:
        if 'conn' in locals():
            conn.rollback()
            cursor.close()
            conn.close()
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            cursor.close()
            conn.close()
        return jsonify({'success': False, 'message': f'Unexpected error: {str(e)}'}), 500
    
@app.route('/get_responsible_users', methods=['GET'])
def get_responsible_users():
    
    try:
        username = session.get('username')
        if not username:
            return jsonify(success=False, message="Not authenticated"), 401

        conn = connect()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT username
            FROM [DT_request].[dbo].[users]
            WHERE role = 1
            ORDER BY username
        """)
        
        users = [{'username': row[0]} for row in cursor.fetchall()]

        cursor.close()
        conn.close()
        return jsonify(success=True, users=users)

    except pyodbc.Error as e:
        return jsonify(success=False, message=f"Database error: {str(e)}"), 500
    except Exception as e:
        return jsonify(success=False, message=f"Unexpected error: {str(e)}"), 500

@app.route('/get_current_user', methods=['GET'])
def get_current_user():
    try:
        username = session.get('username')
        if not username:
            return jsonify(success=False, message="Not authenticated"), 401

        return jsonify(success=True, username=username)

    except Exception as e:
        return jsonify(success=False, message=f"Error: {str(e)}"), 500
    
########################## PROJECTS #######################################################

@app.route('/projects')
def projects():
    try:
        username = session.get('username')
        if not username:
            flash('Please log in to view projects.', 'error')
            return redirect(url_for('index'))

        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, name, description, category, created_by, created_at, updated_at, status, responsible, total_hours, github_repo
            FROM [DT_request].[dbo].[projects]
            ORDER BY created_at DESC
        """)
        projects_list = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]

        # Calculate progress for each project
        for project in projects_list:
            project_id = project['id']
            # Get total estimated hours from tasks
            cursor.execute("""
                SELECT SUM(ISNULL(estimated_hours, 0)) as total_estimated,
                       SUM(CASE WHEN status IN ('Done', 'Completed') THEN ISNULL(estimated_hours, 0) ELSE 0 END) as completed_hours
                FROM [DT_request].[dbo].[tasks]
                WHERE project_id = ?
            """, (project_id,))
            hours_data = cursor.fetchone()
            total_estimated = hours_data[0] or 0
            completed_hours = hours_data[1] or 0

            # Calculate progress percentage
            if total_estimated > 0:
                progress_percentage = min(100, (completed_hours / total_estimated) * 100)
            else:
                progress_percentage = 0

            project['progress_percentage'] = round(progress_percentage, 1)
            project['total_estimated_hours'] = total_estimated
            project['completed_hours'] = completed_hours

            # Parse responsible field (JSON or comma-separated)
            if project['responsible']:
                try:
                    # Try to parse as JSON first
                    import json
                    responsible_list = json.loads(project['responsible'])
                    project['responsible_list'] = responsible_list if isinstance(responsible_list, list) else [responsible_list]
                except:
                    # Fall back to comma-separated
                    project['responsible_list'] = [r.strip() for r in project['responsible'].split(',') if r.strip()]
            else:
                project['responsible_list'] = []

        # Attach task timeline data to each project
        for project in projects_list:
            cursor.execute("""
                SELECT id, title, status, responsible,
                       start_date, end_date,
                       planned_start_date, planned_end_date,
                       description
                FROM [DT_request].[dbo].[tasks]
                WHERE project_id = ?
                ORDER BY COALESCE(planned_start_date, start_date, planned_end_date, end_date)
            """, (project['id'],))
            task_timeline = []
            for row in cursor.fetchall():
                t_id, t_title, t_status, t_resp, t_start, t_end, t_pstart, t_pend, t_desc = row
                # prefer actual dates; fall back to planned dates
                use_start = t_start  or t_pstart
                use_end   = t_end    or t_pend
                if not use_start and not use_end:
                    continue   # task has absolutely no date — skip
                use_start = use_start or use_end
                use_end   = use_end   or use_start
                task_timeline.append({
                    'id':          t_id,
                    'title':       t_title or '(no title)',
                    'description': t_desc  or '',
                    'status':      t_status or '',
                    'responsible': t_resp or '',
                    'start': use_start.strftime('%Y-%m-%d') if hasattr(use_start, 'strftime') else str(use_start),
                    'end':   use_end.strftime('%Y-%m-%d')   if hasattr(use_end,   'strftime') else str(use_end),
                })
            project['task_timeline'] = task_timeline

        cursor.close()
        conn.close()
        return render_template('projects.html', projects=projects_list)
    except Exception as e:
        flash(f"Error fetching projects: {str(e)}", 'error')
        return redirect(url_for('index'))


@app.route('/projects/add', methods=['POST'])
def add_project():
    try:
        data = request.get_json()
        name = data.get('name')
        category = data.get('category')
        description = data.get('description')
        responsible = data.get('responsible', [])
        total_hours = data.get('total_hours', 0)
        github_repo = data.get('github_repo')  # optional
        created_by = session.get('username', 'Unassigned')

        if not name or not category:
            return jsonify({'success': False, 'message': 'Name and category required'}), 400

        # Convert responsible list to JSON string
        import json
        responsible_json = json.dumps(responsible) if responsible else json.dumps([created_by])

        # Set default status based on category
        default_status = 'Waiting'

        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO [DT_request].[dbo].[projects] (name, category, description, created_by, status, responsible, total_hours, github_repo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, category, description, created_by, default_status, responsible_json, total_hours, github_repo))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/projects/<int:project_id>/edit', methods=['PUT'])
def edit_project(project_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401

        data = request.get_json()
        name = data.get('name')
        category = data.get('category')
        description = data.get('description')
        status = data.get('status')
        responsible = data.get('responsible', [])
        total_hours = data.get('total_hours', 0)
        github_repo = data.get('github_repo')  # optional

        if not name or not category:
            return jsonify({'success': False, 'message': 'Name and category required'}), 400

        # Convert responsible list to JSON string
        import json
        responsible_json = json.dumps(responsible) if responsible else None

        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE [DT_request].[dbo].[projects]
            SET name = ?, category = ?, description = ?, status = ?, responsible = ?, total_hours = ?, github_repo = ?, updated_at = GETDATE()
            WHERE id = ?
        """, (name, category, description, status, responsible_json, total_hours, github_repo, project_id))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/projects/<int:project_id>/delete', methods=['DELETE'])
def delete_project(project_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401

        conn = connect()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM [DT_request].[dbo].[tasks] WHERE project_id = ?", (project_id,))
        cursor.execute("DELETE FROM [DT_request].[dbo].[projects] WHERE id = ?", (project_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/projects/<int:project_id>/status', methods=['PUT'])
def update_project_status(project_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401

        data = request.get_json()
        status = data.get('status')

        if not status:
            return jsonify({'success': False, 'message': 'Status required'}), 400

        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE [DT_request].[dbo].[projects]
            SET status = ?, updated_at = GETDATE()
            WHERE id = ?
        """, (status, project_id))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/projects/<int:project_id>/tasks')
def project_tasks(project_id):
    try:
        username = session.get('username')
        if not username:
            flash('Please log in.', 'error')
            return redirect(url_for('index'))

        conn = connect()
        cursor = conn.cursor()
        # Buscar projeto
        cursor.execute("SELECT id, name, description, category, status, responsible, total_hours, github_repo FROM [DT_request].[dbo].[projects] WHERE id = ?", (project_id,))
        project = cursor.fetchone()
        if not project:
            flash('Project not found.', 'error')
            return redirect(url_for('projects'))

        project_data = dict(zip([col[0] for col in cursor.description], project))

        # Parse responsible field
        if project_data['responsible']:
            try:
                import json
                responsible_list = json.loads(project_data['responsible'])
                project_data['responsible_list'] = responsible_list if isinstance(responsible_list, list) else [responsible_list]
            except:
                project_data['responsible_list'] = [r.strip() for r in project_data['responsible'].split(',') if r.strip()]
        else:
            project_data['responsible_list'] = []

        # Buscar TODAS as tarefas associadas ao projeto
        cursor.execute("""
            SELECT t.id, t.category, t.title, t.description, t.responsible, t.priority, t.status,
                   t.start_date, t.end_date, t.week_number, t.equipment, t.equipment_responsible,
                   t.time_spent, t.planned_start_date, t.planned_end_date, t.estimated_hours,
                   p.github_repo
            FROM [DT_request].[dbo].[tasks] t
            LEFT JOIN [DT_request].[dbo].[projects] p ON t.project_id = p.id
            WHERE t.project_id = ?
            ORDER BY t.planned_start_date, t.created_at DESC
        """, (project_id,))
        tasks_list = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]

        # Calculate project progress
        total_estimated = sum(t['estimated_hours'] or 0 for t in tasks_list)
        completed_hours = sum((t['estimated_hours'] or 0) for t in tasks_list if t['status'] in ['Done', 'Completed'])
        progress_percentage = (completed_hours / total_estimated * 100) if total_estimated > 0 else 0

        project_data['progress_percentage'] = round(progress_percentage, 1)
        project_data['total_estimated_hours'] = total_estimated
        project_data['completed_hours'] = completed_hours

        # Prepare timeline data
        timeline_tasks = []
        for task in tasks_list:
            if task['planned_start_date'] or task['start_date']:
                start_date = task['planned_start_date'] or task['start_date']
                end_date = task['planned_end_date'] or task['end_date'] or start_date
                timeline_tasks.append({
                    'id': task['id'],
                    'title': task['title'],
                    'start': start_date.strftime('%Y-%m-%d') if hasattr(start_date, 'strftime') else str(start_date),
                    'end': end_date.strftime('%Y-%m-%d') if hasattr(end_date, 'strftime') else str(end_date),
                    'status': task['status'],
                    'responsible': task['responsible'],
                    'estimated_hours': task['estimated_hours'] or 0
                })

        cursor.close()
        conn.close()

        # Formatação das datas
        for t in tasks_list:
            for f in ['start_date', 'end_date', 'planned_start_date', 'planned_end_date']:
                t[f] = t[f].strftime('%Y-%m-%d') if t[f] else None
            t['estimated_hours'] = t['estimated_hours'] if t['estimated_hours'] is not None else None

        return render_template('project_tasks.html', project=project_data, tasks=tasks_list)
    except Exception as e:
        flash(f"Error loading project tasks: {str(e)}", 'error')
        return redirect(url_for('projects'))


@app.route('/projects/<int:project_id>/tasks/add', methods=['POST'])
def add_project_task(project_id):
    try:
        username = session.get('username')
        if not username:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401

        data = request.get_json()
        title = data.get('title', 'New Task')
        description = data.get('description', '')
        category = data.get('category', 'Software')
        status = data.get('status', 'To Do')
        planned_start_date = data.get('planned_start_date', None)
        planned_end_date = data.get('planned_end_date', None)
        start_date = data.get('start_date', None)
        end_date = data.get('end_date', None)
        equipment = data.get('equipment', '')
        equipment_responsible = data.get('equipment_responsible', '')
        time_spent = data.get('time_spent', 0)
        estimated_hours = data.get('estimated_hours', None)
        responsible = data.get('responsible', [username])
        responsible_str = format_responsible(responsible)

        # Validar e formatar as datas
        planned_start_date_obj = None
        planned_end_date_obj = None
        start_date_obj = None
        end_date_obj = None
        try:
            if planned_start_date:
                planned_start_date_obj = datetime.strptime(planned_start_date, '%Y-%m-%d').date()
            if planned_end_date:
                planned_end_date_obj = datetime.strptime(planned_end_date, '%Y-%m-%d').date()
            if start_date:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            if end_date:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format'}), 400

        # Calcula o week_number corretamente
        week_number = None
        if planned_start_date_obj or start_date_obj:
            week_number = (planned_start_date_obj or start_date_obj).isocalendar()[1]

        conn = connect()
        cursor = conn.cursor()
        
        # Determinar se é principal task para o projeto
        is_principal = 0
        principal_task_id = None
        
        # Verificar se já existem tasks para este projeto
        cursor.execute("""
            SELECT id, is_principal_task FROM tasks 
            WHERE project_id = ? AND principal_task_id IS NULL
            ORDER BY created_at ASC
        """, (project_id,))
        existing_tasks = cursor.fetchall()
        
        if not existing_tasks:
            # É a primeira task para este projeto
            is_principal = 1
        else:
            # Encontrar a principal task existente
            for existing_task in existing_tasks:
                if existing_task[1] == 1:  # is_principal_task = 1
                    principal_task_id = existing_task[0]
                    break
        
        cursor.execute("""
            INSERT INTO [DT_request].[dbo].[tasks] 
            (category, title, description, responsible, priority, status, start_date, end_date, 
             planned_start_date, planned_end_date, equipment, equipment_responsible, time_spent, 
             estimated_hours, project_id, week_number, created_at, updated_at, is_principal_task, principal_task_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, GETDATE(), GETDATE(), ?, ?)
        """, (
            category, title, description, responsible_str, 'Medium', status,
            start_date_obj, end_date_obj,
            planned_start_date_obj, planned_end_date_obj,
            equipment, equipment_responsible, time_spent, estimated_hours,
            project_id, week_number, is_principal, principal_task_id
        ))
        cursor.execute("SELECT @@IDENTITY AS id")
        task_id = cursor.fetchone().id

        # Se projeto tem github_repo, criação de Issue é obrigatória
        github_repo = get_project_github_repo(cursor, project_id)
        logging.info(f"[DEBUG] Task {task_id} - Project {project_id} - github_repo: {github_repo}")

        if github_repo:
            try:
                # Buscar assignees dos responsáveis do projeto
                cursor.execute("SELECT responsible FROM projects WHERE id = ?", (project_id,))
                proj_resp = cursor.fetchone()
                responsible_json = proj_resp[0] if proj_resp else None
                logging.info(f"[DEBUG] Project responsible JSON: {responsible_json}")

                assignees = []
                if responsible_json:
                    assignees = get_user_github_usernames(responsible_json, cursor)
                logging.info(f"[DEBUG] Assignees for GitHub Issue: {assignees}")

                # Criar Issue
                issue_number = create_issue(
                    repo=github_repo,
                    title=title,
                    description=description,
                    assignees=assignees
                )
                logging.info(f"[DEBUG] create_issue returned: {issue_number}")
                if issue_number:
                    # Guardar número da Issue na tarefa
                    cursor.execute("UPDATE tasks SET github_issue_number = ? WHERE id = ?", (issue_number, task_id))
                    logging.info(f"GitHub Issue #{issue_number} criada para tarefa {task_id}")
                else:
                    conn.rollback()
                    cursor.close()
                    conn.close()
                    return jsonify({
                        'success': False,
                        'message': 'Task not created because GitHub Issue creation failed for project repository'
                    }), 500
            except Exception as e:
                conn.rollback()
                cursor.close()
                conn.close()
                logging.error(f"Erro ao criar GitHub Issue para tarefa {task_id}: {str(e)}", exc_info=True)
                return jsonify({
                    'success': False,
                    'message': 'Task not created because GitHub Issue creation failed for project repository'
                }), 500

        conn.commit()

        cursor.close()
        conn.close()
        return jsonify({'success': True, 'task_id': task_id})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
    
###################### DT PLANNING #################################

@app.route('/planning')
def planning():
    try:
        conn = connect()
        cursor = conn.cursor()

        # Fetch Software (responsible for applications)
        cursor.execute("""
            SELECT id, application_name, responsible, notes
            FROM [DT_request].[dbo].[planning]
            WHERE category = 'software'
            ORDER BY application_name ASC
        """)
        
        columns = [column[0] for column in cursor.description]
        software = [dict(zip(columns, row)) for row in cursor.fetchall()]

        cursor.close()
        conn.close()

        return render_template('dt_planning.html', software=software)
    except Exception as e:
        flash(f"Error fetching planning data: {str(e)}", 'error')
        return redirect(url_for('index'))
    
@app.route('/update_automation', methods=['POST'])
def update_automation():
    try:
        data = request.get_json()
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE [DT_request].[dbo].[planning]
            SET responsible = ?
            WHERE id = ?
        """, (data['responsible'], data['id']))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/get_automation_planning')
def get_automation_planning():
    month = int(request.args.get('month'))
    year = int(request.args.get('year'))
    conn = connect()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, week_number, year, responsible, notes
        FROM [DT_request].[dbo].[planning]
        WHERE category = 'automation' AND year = ?
        ORDER BY week_number
    """, (year,))
    planning = []
    for row in cursor.fetchall():
        week_number = row[1]
        # Calcula o primeiro e último dia da semana
        jan1 = datetime(year, 1, 1)
        start_of_week = jan1 + timedelta(days=(week_number - 1) * 7)
        # Ajusta para segunda-feira
        day_of_week = start_of_week.weekday()
        start_of_week = start_of_week - timedelta(days=day_of_week)
        end_of_week = start_of_week + timedelta(days=6)
        # Se QUALQUER dia da semana cair no mês, inclui
        if (start_of_week.month == month or end_of_week.month == month or
            (start_of_week.month < month < end_of_week.month)):
            planning.append({
                'id': row[0],
                'week_number': week_number,
                'year': row[2],
                'responsible': row[3],
                'notes': row[4]
            })
    cursor.close()
    conn.close()
    return jsonify({'success': True, 'planning': planning})

@app.route('/get_software_users', methods=['GET'])
def get_software_users():
    try:
        conn = connect()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT username, email
            FROM [DT_request].[dbo].[users]
            WHERE category = 'Software'
            ORDER BY username
        """)
        
        users = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'users': users})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/get_automation_users', methods=['GET'])
def get_automation_users():
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT username, email
            FROM [DT_request].[dbo].[users]
            WHERE category = 'Automation'
            ORDER BY username
        """)
        users = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'users': users})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/add_software_application', methods=['POST'])
def add_software_application():
    try:
        data = request.get_json()
        application_name = data.get('application_name')
        responsible = data.get('responsible')
        notes = data.get('notes', '')
        now = datetime.now()

        if not application_name or not responsible:
            return jsonify({'success': False, 'message': 'Nome da aplicação e responsável são obrigatórios'}), 400

        conn = connect()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO [DT_request].[dbo].[planning] 
                (category, application_name, responsible, notes, year, week_number, created_at, updated_at)
            OUTPUT INSERTED.ID
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            'software',
            application_name,
            responsible,
            notes,
            now.year,
            None,
            now,
            now
        ))

        new_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/remove_software_application/<int:app_id>', methods=['DELETE'])
def remove_software_application(app_id):
    try:
        conn = connect()
        cursor = conn.cursor()
        
        cursor.execute("""
            DELETE FROM [DT_request].[dbo].[planning]
            WHERE id = ? AND category = 'software'
        """, (app_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/remove_automation_support/<int:support_id>', methods=['DELETE'])
def remove_automation_support(support_id):
    try:
        conn = connect()
        cursor = conn.cursor()
        
        cursor.execute("""
            DELETE FROM [DT_request].[dbo].[planning]
            WHERE id = ? AND category = 'automation'
        """, ( support_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/draw_automation_support', methods=['POST'])
def draw_automation_support():
    try:
        import random
        import datetime

        data = request.get_json()
        month = int(data.get('month', datetime.datetime.now().month))
        year = int(data.get('year', datetime.datetime.now().year))

        today = datetime.datetime.now()
        current_week = today.isocalendar()[1]

        # Calcular todas as semanas do mês 
        first_day = datetime.date(year, month, 1)
        if month == 12:
            last_day = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            last_day = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)

        # Lista de semanas do mês (número da semana ISO)
        weeks_in_month = []
        d = first_day
        while d <= last_day:
            week_num = d.isocalendar()[1]
            if week_num not in weeks_in_month:
                weeks_in_month.append(week_num)
            d += datetime.timedelta(days=1)

        conn = connect()
        cursor = conn.cursor()

        # Verificar se já existem atribuições para TODAS as semanas do mês
        cursor.execute("""
            SELECT week_number, responsible
            FROM [DT_request].[dbo].[planning]
            WHERE category = 'automation' AND year = ?
        """, (year,))
        all_assignments = cursor.fetchall()

        # Filtrar por mês
        assigned_weeks = set()
        for assignment in all_assignments:
            week_num = assignment[0]
            jan1 = datetime.date(year, 1, 1)
            week_start = jan1 + timedelta(weeks=week_num-1)
            if week_start.month == month:
                assigned_weeks.add(week_num)

        # Se todas as semanas do mês já tiverem atribuição, retorna erro
        if all(week in assigned_weeks for week in weeks_in_month):
            return jsonify({
                'success': False, 
                'message': 'Todas as semanas deste mês já possuem responsáveis designados.'
            })

        # Buscar todos os usuários da categoria 'Automation'
        cursor.execute("""
            SELECT username, email
            FROM [DT_request].[dbo].[users]
            WHERE category = 'Automation'
        """)
        all_users = [row[0] for row in cursor.fetchall()]

        if not all_users:
            return jsonify({
                'success': False, 
                'message': 'Nenhum usuário disponível para suporte à automação.'
            })

        sorteios = []
        used_users = set(u[1] for u in all_assignments if u[0] in weeks_in_month)
        available_users = [u for u in all_users if u not in used_users]
        week_user_map = {}

        # Para cada semana do mês, se não houver responsável, sorteia um diferente
        for week_num in weeks_in_month:
            if week_num in assigned_weeks:
                continue

            # Se todos já foram sorteados, reseta (mas nunca repete no mês)
            if not available_users:
                available_users = [u for u in all_users if u not in week_user_map.values()]

            if not available_users:
                available_users = all_users.copy()

            selected_user = random.choice(available_users)
            week_user_map[week_num] = selected_user
            available_users.remove(selected_user)

            # Adicionar ao banco
            cursor.execute("""
                INSERT INTO [DT_request].[dbo].[planning] 
                (category, week_number, year, application_name, responsible, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                'automation',
                week_num,
                year,
                None,
                selected_user,
                f"Responsável pelo suporte à automação da semana {week_num}/{year}"
            ))
            sorteios.append({'week_number': week_num, 'user': selected_user})

        conn.commit()
        cursor.close()
        conn.close()

        if sorteios:
            msg = "Sorteio realizado para as semanas: " + ", ".join(
                [f"{s['week_number']} ({s['user']})" for s in sorteios]
            )
        else:
            msg = "Não há semanas disponíveis para sorteio neste mês."

        return jsonify({
            'success': True,
            'message': msg,
            'sorteios': sorteios
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    
@app.route('/team_building')
def team_building():
    return render_template('team_building.html')

# ──────────────────────────────────────────────────────────
# Hangman (Jogo da Forca) – multiplayer online
# ──────────────────────────────────────────────────────────
# Rooms are stored in SQL so all IIS worker processes share state.
hangman_lock = threading.Lock()
HANGMAN_ALLOWED_USERS = {'pelopes', 'nnovais', 'josamorim'}

# ── DB helpers ────────────────────────────────────────────
def _hr_init():
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name='teste')
            CREATE TABLE [DT_request].[dbo].[teste] (
                room_id VARCHAR(8) NOT NULL,
                data NVARCHAR(MAX) NOT NULL,
                last_activity FLOAT NOT NULL,
                CONSTRAINT PK_hangman_rooms PRIMARY KEY (room_id)
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        logging.error(f"hangman _hr_init error: {e}")

def _hr_get(room_id):
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("SELECT data FROM [DT_request].[dbo].[teste] WHERE room_id = ?", (room_id,))
        row = cursor.fetchone()
        cursor.close(); conn.close()
        return json.loads(row[0]) if row else None
    except Exception as e:
        logging.error(f"_hr_get error: {e}")
        return None

def _hr_set(room_id, room):
    try:
        room['last_activity'] = time.time()
        data_json = json.dumps(room, default=str)
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            MERGE [DT_request].[dbo].[teste] AS t
            USING (SELECT ? AS rid, ? AS d, ? AS la) AS s
                ON t.room_id = s.rid
            WHEN MATCHED THEN UPDATE SET data = s.d, last_activity = s.la
            WHEN NOT MATCHED THEN INSERT (room_id, data, last_activity) VALUES (s.rid, s.d, s.la);
        """, (room_id, data_json, room['last_activity']))
        conn.commit()
        cursor.close(); conn.close()
    except Exception as e:
        logging.error(f"_hr_set error: {e}")

def _hr_delete(room_id):
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM [DT_request].[dbo].[teste] WHERE room_id = ?", (room_id,))
        conn.commit()
        cursor.close(); conn.close()
    except Exception as e:
        logging.error(f"_hr_delete error: {e}")

def _hr_items():
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("SELECT room_id, data FROM [DT_request].[dbo].[teste]")
        rows = cursor.fetchall()
        cursor.close(); conn.close()
        return [(rid, json.loads(d)) for rid, d in rows]
    except Exception as e:
        logging.error(f"_hr_items error: {e}")
        return []

def _hr_cleanup():
    try:
        expiry = time.time() - 3600
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM [DT_request].[dbo].[teste] WHERE last_activity < ?", (expiry,))
        conn.commit()
        cursor.close(); conn.close()
    except Exception as e:
        logging.error(f"_hr_cleanup error: {e}")

try:
    _hr_init()
except Exception:
    pass


def _hangman_check_access():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    if session['username'] not in HANGMAN_ALLOWED_USERS:
        return jsonify({'error': 'Acesso não autorizado'}), 403
    return None

def _normalize_word(word):
    nfkd = unicodedata.normalize('NFKD', word)
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).upper().strip()

def _advance_turn(room):
    players = room['players']
    if len(players) <= 1:
        return
    chooser = room['chooser']
    guessers = [p for p in players if p != chooser]
    if not guessers:
        return
    current = room.get('current_turn')
    if current in guessers:
        idx = guessers.index(current)
        room['current_turn'] = guessers[(idx + 1) % len(guessers)]
    else:
        room['current_turn'] = guessers[0]

def _calculate_scores(room):
    chooser = room['chooser']
    unique_letters = len(set(room['normalized_word']))
    errors = room['errors']
    if room['status'] == 'round_won':
        guesser_score = 100 + (unique_letters * 10) + (room['max_errors'] - errors) * 15
        chooser_score = errors * 20
    else:
        guesser_score = 0
        chooser_score = 150 + (unique_letters * 10)
    guessers = [p for p in room['players'] if p != chooser]
    for g in guessers:
        room['scores'][g] = room['scores'].get(g, 0) + guesser_score
    room['scores'][chooser] = room['scores'].get(chooser, 0) + chooser_score

@app.route('/hangman')
def hangman():
    if 'name' not in session:
        return redirect(url_for('login'))
    if session.get('username') not in HANGMAN_ALLOWED_USERS:
        return redirect(url_for('home'))
    return render_template('hangman.html')

@app.route('/hangman/api/rooms', methods=['GET'])
def hangman_list_rooms():
    denied = _hangman_check_access()
    if denied:
        return denied
    _hr_cleanup()
    rooms = []
    for rid, r in _hr_items():
        rooms.append({
            'id': rid,
            'name': r['name'],
            'host': r['host'],
            'players': r['players'],
            'status': r['status'],
            'player_count': len(r['players']),
            'max_players': 4,
        })
    return jsonify({'rooms': rooms})

@app.route('/hangman/api/rooms', methods=['POST'])
def hangman_create_room():
    denied = _hangman_check_access()
    if denied:
        return denied
    data = request.get_json(force=True)
    room_name = (data.get('name') or '').strip()
    if not room_name:
        return jsonify({'error': 'Nome da sala é obrigatório'}), 400
    username = session['username']
    display = session.get('name', username)
    room_id = uuid.uuid4().hex[:8]
    new_room = {
        'name': room_name,
        'host': username,
        'players': [username],
        'display_names': {username: display},
        'status': 'waiting_for_players',
        'created_at': time.time(),
        'scores': {},
        'round': 0,
        'word': None,
        'normalized_word': None,
        'guessed_letters': [],
        'errors': 0,
        'max_errors': 6,
        'chooser': None,
        'current_turn': None,
        'history': [],
        'last_activity': time.time(),
    }
    _hr_set(room_id, new_room)
    return jsonify({'success': True, 'room_id': room_id})

@app.route('/hangman/api/rooms/<room_id>/join', methods=['POST'])
def hangman_join_room(room_id):
    denied = _hangman_check_access()
    if denied:
        return denied
    username = session['username']
    display = session.get('name', username)
    with hangman_lock:
        room = _hr_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if username in room['players']:
            return jsonify({'success': True})
        if room['status'] != 'waiting_for_players':
            return jsonify({'error': 'Jogo já começou'}), 400
        if len(room['players']) >= 4:
            return jsonify({'error': 'Sala cheia (máximo 4 jogadores)'}), 400
        room['players'].append(username)
        room['display_names'][username] = display
        _hr_set(room_id, room)
    return jsonify({'success': True})

@app.route('/hangman/api/rooms/<room_id>/leave', methods=['POST'])
def hangman_leave_room(room_id):
    denied = _hangman_check_access()
    if denied:
        return denied
    username = session['username']
    with hangman_lock:
        room = _hr_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if username not in room['players']:
            return jsonify({'success': True})
        room['players'].remove(username)
        if not room['players']:
            _hr_delete(room_id)
            return jsonify({'success': True, 'room_deleted': True})
        if room['host'] == username:
            room['host'] = room['players'][0]
        if room['status'] in ('playing', 'choosing_word'):
            if room.get('chooser') == username:
                room['status'] = 'waiting_for_players'
                room['word'] = None
                room['normalized_word'] = None
                room['guessed_letters'] = []
                room['errors'] = 0
            elif room.get('current_turn') == username:
                _advance_turn(room)
        _hr_set(room_id, room)
    return jsonify({'success': True})

@app.route('/hangman/api/rooms/<room_id>/start', methods=['POST'])
def hangman_start_game(room_id):
    denied = _hangman_check_access()
    if denied:
        return denied
    username = session['username']
    with hangman_lock:
        room = _hr_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if room['host'] != username:
            return jsonify({'error': 'Apenas o host pode iniciar'}), 403
        if len(room['players']) < 2:
            return jsonify({'error': 'Mínimo 2 jogadores'}), 400
        room['round'] = 1
        room['chooser'] = room['players'][0]
        room['status'] = 'choosing_word'
        room['word'] = None
        room['normalized_word'] = None
        room['guessed_letters'] = []
        room['errors'] = 0
        room['current_turn'] = None
        room['last_activity'] = time.time()
        _hr_set(room_id, room)
    return jsonify({'success': True})

@app.route('/hangman/api/rooms/<room_id>/word', methods=['POST'])
def hangman_submit_word(room_id):
    denied = _hangman_check_access()
    if denied:
        return denied
    username = session['username']
    data = request.get_json(force=True)
    word = (data.get('word') or '').strip()
    if not word or len(word) < 2:
        return jsonify({'error': 'Palavra inválida (mínimo 2 letras)'}), 400
    normalized = _normalize_word(word)
    if not normalized.isalpha():
        return jsonify({'error': 'Apenas letras são permitidas'}), 400
    with hangman_lock:
        room = _hr_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if room['status'] != 'choosing_word':
            return jsonify({'error': 'Não é fase de escolher palavra'}), 400
        if room['chooser'] != username:
            return jsonify({'error': 'Não é a tua vez de escolher'}), 403
        room['word'] = word.upper()
        room['normalized_word'] = normalized
        room['guessed_letters'] = []
        room['errors'] = 0
        room['status'] = 'playing'
        guessers = [p for p in room['players'] if p != username]
        room['current_turn'] = guessers[0] if guessers else None
        room['last_activity'] = time.time()
        _hr_set(room_id, room)
    return jsonify({'success': True})

@app.route('/hangman/api/rooms/<room_id>/guess', methods=['POST'])
def hangman_guess(room_id):
    denied = _hangman_check_access()
    if denied:
        return denied
    username = session['username']
    data = request.get_json(force=True)
    letter = (data.get('letter') or '').strip().upper()
    if not letter or len(letter) != 1 or not letter.isalpha():
        return jsonify({'error': 'Letra inválida'}), 400
    with hangman_lock:
        room = _hr_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if room['status'] != 'playing':
            return jsonify({'error': 'Jogo não está em andamento'}), 400
        if room.get('current_turn') != username:
            return jsonify({'error': 'Não é a tua vez'}), 403
        if letter in room['guessed_letters']:
            return jsonify({'error': 'Letra já jogada'}), 400
        room['guessed_letters'].append(letter)
        hit = letter in room['normalized_word']
        if not hit:
            room['errors'] += 1
        all_revealed = all(l in room['guessed_letters'] for l in set(room['normalized_word']))
        if all_revealed:
            room['status'] = 'round_won'
            _calculate_scores(room)
            room['history'].append({
                'round': room['round'],
                'word': room['word'],
                'chooser': room['display_names'].get(room['chooser'], room['chooser']),
                'result': 'won',
                'errors': room['errors'],
            })
        elif room['errors'] >= room['max_errors']:
            room['status'] = 'round_lost'
            _calculate_scores(room)
            room['history'].append({
                'round': room['round'],
                'word': room['word'],
                'chooser': room['display_names'].get(room['chooser'], room['chooser']),
                'result': 'lost',
                'errors': room['errors'],
            })
        else:
            _advance_turn(room)
        room['last_activity'] = time.time()
        _hr_set(room_id, room)
    return jsonify({'success': True, 'hit': hit})

@app.route('/hangman/api/rooms/<room_id>/next_round', methods=['POST'])
def hangman_next_round(room_id):
    denied = _hangman_check_access()
    if denied:
        return denied
    username = session['username']
    with hangman_lock:
        room = _hr_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if room['host'] != username:
            return jsonify({'error': 'Apenas o host pode avançar'}), 403
        if room['status'] not in ('round_won', 'round_lost'):
            return jsonify({'error': 'Ronda ainda em curso'}), 400
        room['round'] += 1
        chooser_idx = room['players'].index(room['chooser'])
        room['chooser'] = room['players'][(chooser_idx + 1) % len(room['players'])]
        room['status'] = 'choosing_word'
        room['word'] = None
        room['normalized_word'] = None
        room['guessed_letters'] = []
        room['errors'] = 0
        room['current_turn'] = None
        room['last_activity'] = time.time()
        _hr_set(room_id, room)
    return jsonify({'success': True})

@app.route('/hangman/api/rooms/<room_id>/state', methods=['GET'])
def hangman_state(room_id):
    denied = _hangman_check_access()
    if denied:
        return denied
    username = session['username']
    with hangman_lock:
        room = _hr_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        is_chooser = (room.get('chooser') == username)
        normalized = room.get('normalized_word') or ''
        if room['status'] == 'playing' and not is_chooser:
            display_word = ''.join(l if l in room['guessed_letters'] else '_' for l in normalized)
        elif room['status'] in ('round_won', 'round_lost') or is_chooser:
            display_word = room.get('word') or ''
        else:
            display_word = ''
        state = {
            'name': room['name'],
            'host': room['host'],
            'players': room['players'],
            'display_names': room['display_names'],
            'status': room['status'],
            'round': room['round'],
            'scores': room['scores'],
            'chooser': room.get('chooser'),
            'current_turn': room.get('current_turn'),
            'display_word': display_word,
            'word_length': len(normalized),
            'guessed_letters': room['guessed_letters'],
            'errors': room['errors'],
            'max_errors': room['max_errors'],
            'history': room['history'],
            'is_chooser': is_chooser,
            'my_username': username,
        }
    return jsonify(state)

# ──────────────────────────────────────────────────────────
# Batalha Naval (Battleship) – multiplayer online (2-4 jogadores)
# Uses the same table 'teste' as hangman but different room structure
battleship_lock = threading.Lock()
BATTLESHIP_ALLOWED_USERS = {'pelopes', 'nnovais', 'josamorim'}

# Ship definitions
SHIPS_DEF = [
    {'name': 'Porta-Aviões', 'size': 5},
    {'name': 'Couraçado',    'size': 4},
    {'name': 'Cruzador',     'size': 3},
    {'name': 'Submarino',    'size': 3},
    {'name': 'Destroyer',    'size': 2},
]

# ── DB helpers ────────────────────────────────────────────
def _br_init():
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sys.tables WHERE name='teste')
            CREATE TABLE [DT_request].[dbo].[teste] (
                room_id VARCHAR(8) NOT NULL,
                data NVARCHAR(MAX) NOT NULL,
                last_activity FLOAT NOT NULL,
                CONSTRAINT PK_battleship_rooms PRIMARY KEY (room_id)
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        logging.error(f"battleship _br_init error: {e}")

def _br_get(room_id):
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("SELECT data FROM [DT_request].[dbo].[teste] WHERE room_id = ?", (room_id,))
        row = cursor.fetchone()
        cursor.close(); conn.close()
        return json.loads(row[0]) if row else None
    except Exception as e:
        logging.error(f"_br_get error: {e}")
        return None

def _br_set(room_id, room):
    try:
        room['last_activity'] = time.time()
        data_json = json.dumps(room, default=str)
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("""
            MERGE [DT_request].[dbo].[teste] AS t
            USING (SELECT ? AS rid, ? AS d, ? AS la) AS s
                ON t.room_id = s.rid
            WHEN MATCHED THEN UPDATE SET data = s.d, last_activity = s.la
            WHEN NOT MATCHED THEN INSERT (room_id, data, last_activity) VALUES (s.rid, s.d, s.la);
        """, (room_id, data_json, room['last_activity']))
        conn.commit()
        cursor.close(); conn.close()
    except Exception as e:
        logging.error(f"_br_set error: {e}")

def _br_delete(room_id):
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM [DT_request].[dbo].[teste] WHERE room_id = ?", (room_id,))
        conn.commit()
        cursor.close(); conn.close()
    except Exception as e:
        logging.error(f"_br_delete error: {e}")

def _br_items():
    try:
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("SELECT room_id, data FROM [DT_request].[dbo].[teste]")
        rows = cursor.fetchall()
        cursor.close(); conn.close()
        return [(rid, json.loads(d)) for rid, d in rows]
    except Exception as e:
        logging.error(f"_br_items error: {e}")
        return []

def _br_cleanup():
    try:
        expiry = time.time() - 3600
        conn = connect()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM [DT_request].[dbo].[teste] WHERE last_activity < ?", (expiry,))
        conn.commit()
        cursor.close(); conn.close()
    except Exception as e:
        logging.error(f"_br_cleanup error: {e}")

try:
    _br_init()
except Exception:
    pass

def _battleship_check_access():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    if session['username'] not in BATTLESHIP_ALLOWED_USERS:
        return jsonify({'error': 'Acesso não autorizado'}), 403
    return None

# ── Game logic helpers ────────────────────────────────────
def _generate_board():
    """Generate empty 10x10 board"""
    return [[None for _ in range(10)] for _ in range(10)]

def _can_place_ship(board, cells):
    """Check if ship can be placed on given cells"""
    for r, c in cells:
        if r < 0 or r >= 10 or c < 0 or c >= 10:
            return False
        if board[r][c] is not None:
            return False
    return True

def _get_ship_cells(size, start_r, start_c, orientation):
    """Get list of cells for a ship placement"""
    cells = []
    for i in range(size):
        r = start_r + (i if orientation == 'V' else 0)
        c = start_c + (i if orientation == 'H' else 0)
        cells.append([r, c])
    return cells

def _advance_turn(room):
    """Advance to next player turn (skip eliminated players)"""
    players = room['players']
    if len(players) <= 1:
        return
    current = room.get('current_turn')
    # Find next non-eliminated player
    for i in range(len(players)):
        next_idx = (players.index(current) + 1 + i) % len(players)
        next_player = players[next_idx]
        if next_player not in room.get('eliminated', []):
            room['current_turn'] = next_player
            return
    # All remaining players eliminated? (should not happen)
    room['current_turn'] = None

def _check_ship_sunk(board, ship_name):
    """Check if a ship has all cells hit"""
    for r in range(10):
        for c in range(10):
            if board[r][c] == ship_name:
                return False
    return True

def _calculate_score_for_hit():
    """Return score for a successful hit"""
    return 10

def _serialize_board(board):
    """Convert board to JSON-serializable format"""
    return [[board[r][c] for c in range(10)] for r in range(10)]

def _deserialize_board(data):
    """Convert JSON data back to board"""
    if not data or len(data) != 10:
        return _generate_board()
    return [[data[r][c] if data[r][c] is not None else None for c in range(10)] for r in range(10)]

def _normalize_name(name):
    """Normalize ship name for comparison (remove accents, uppercase)"""
    nfkd = unicodedata.normalize('NFKD', name)
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).upper().strip()

# ── Routes ───────────────────────────────────────────────
@app.route('/battleship')
def battleship():
    if 'name' not in session:
        return redirect(url_for('login'))
    if session.get('username') not in BATTLESHIP_ALLOWED_USERS:
        return redirect(url_for('home'))
    return render_template('battleship.html')

@app.route('/battleship/api/rooms', methods=['GET'])
def battleship_list_rooms():
    denied = _battleship_check_access()
    if denied:
        return denied
    _br_cleanup()
    rooms = []
    for rid, r in _br_items():
        # Only include valid battleship rooms
        if r.get('game_type') != 'battleship':
            continue
        rooms.append({
            'id': rid,
            'name': r['name'],
            'host': r['host'],
            'players': r['players'],
            'display_names': r['display_names'],
            'status': r['status'],
            'player_count': len(r['players']),
            'max_players': 4,
        })
    return jsonify({'rooms': rooms})

@app.route('/battleship/api/rooms', methods=['POST'])
def battleship_create_room():
    denied = _battleship_check_access()
    if denied:
        return denied
    data = request.get_json(force=True)
    room_name = (data.get('name') or '').strip()
    if not room_name:
        return jsonify({'error': 'Nome da sala é obrigatório'}), 400
    username = session['username']
    display = session.get('name', username)
    room_id = uuid.uuid4().hex[:8]
    new_room = {
        'game_type': 'battleship',
        'name': room_name,
        'host': username,
        'players': [username],
        'display_names': {username: display},
        'status': 'waiting',  # waiting, placing, playing, finished
        'created_at': time.time(),
        'scores': {username: 0},
        'boards': {},  # username -> {board: [[...]], ships: [...], ships_remaining: 5, shots_against: [...]}
        'ships_placed': [],  # list of usernames who have placed ships
        'current_turn': None,
        'eliminated': [],
        'shot_history': [],  # [{shooter, target, row, col, hit, sunk, timestamp}]
        'last_activity': time.time(),
    }
    _br_set(room_id, new_room)
    return jsonify({'success': True, 'room_id': room_id})

@app.route('/battleship/api/rooms/<room_id>/join', methods=['POST'])
def battleship_join_room(room_id):
    denied = _battleship_check_access()
    if denied:
        return denied
    username = session['username']
    display = session.get('name', username)
    with battleship_lock:
        room = _br_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if room['status'] not in ('waiting', 'placing'):
            return jsonify({'error': 'Sala não aceita mais jogadores'}), 400
        if len(room['players']) >= 4:
            return jsonify({'error': 'Sala cheia'}), 400
        if username in room['players']:
            return jsonify({'error': 'Já estás na sala'}), 400
        room['players'].append(username)
        room['display_names'][username] = display
        room['scores'][username] = 0
        # If joining during placing phase, initialize board for the new player
        if room['status'] == 'placing':
            room['boards'][username] = {
                'board': _generate_board(),
                'ships': [],
                'ships_remaining': len(SHIPS_DEF),
                'shots_against': [],
                'sunk_ships': [],
            }
            # Ensure ships_placed list does not include this player
            if username in room.get('ships_placed', []):
                room['ships_placed'].remove(username)
            # Update player_order to include new player at the end
            if 'player_order' in room:
                if username not in room['player_order']:
                    room['player_order'].append(username)
            else:
                room['player_order'] = room['players'][:]
        room['last_activity'] = time.time()
        _br_set(room_id, room)
    return jsonify({'success': True})

@app.route('/battleship/api/rooms/<room_id>/leave', methods=['POST'])
def battleship_leave_room(room_id):
    denied = _battleship_check_access()
    if denied:
        return denied
    username = session['username']
    with battleship_lock:
        room = _br_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if username not in room['players']:
            return jsonify({'success': True})  # Already gone
        room['players'].remove(username)
        room['display_names'].pop(username, None)
        room['scores'].pop(username, None)
        room['boards'].pop(username, None)
        if username in room.get('ships_placed', []):
            room['ships_placed'].remove(username)
        # Remove from player_order if present
        if 'player_order' in room and username in room['player_order']:
            try:
                room['player_order'].remove(username)
            except ValueError:
                pass
        # If host leaves, assign new host
        if room['host'] == username:
            if room['players']:
                room['host'] = room['players'][0]
            else:
                _br_delete(room_id)
                return jsonify({'success': True, 'deleted': True})
        # If game in playing phase and player leaves, mark as eliminated
        if room['status'] == 'playing':
            if username not in room.get('eliminated', []):
                room.setdefault('eliminated', []).append(username)
            # Check if all non-eliminated players eliminated, end game
            active_players = [p for p in room['players'] if p not in room.get('eliminated', [])]
            if len(active_players) <= 1:
                room['status'] = 'finished'
                if len(active_players) == 1:
                    room['winner'] = active_players[0]
                else:
                    room['winner'] = None
        room['last_activity'] = time.time()
        _br_set(room_id, room)
    return jsonify({'success': True, 'deleted': False})

@app.route('/battleship/api/rooms/<room_id>/start', methods=['POST'])
def battleship_start_game(room_id):
    denied = _battleship_check_access()
    if denied:
        return denied
    username = session['username']
    with battleship_lock:
        room = _br_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if room['host'] != username:
            return jsonify({'error': 'Apenas o host pode iniciar'}), 403
        if len(room['players']) < 2:
            return jsonify({'error': 'Precisas de pelo menos 2 jogadores'}), 400
        if len(room['players']) > 4:
            return jsonify({'error': 'Máximo de 4 jogadores'}), 400
        room['status'] = 'placing'
        room['ships_placed'] = []
        room['current_turn'] = None
        room['eliminated'] = []
        room['shot_history'] = []
        # Randomize turn order
        import random as rnd
        players = room['players'][:]
        rnd.shuffle(players)
        room['player_order'] = players  # For turn rotation
        # Initialize boards for each player
        room['boards'] = {}
        for p in room['players']:
            room['boards'][p] = {
                'board': _generate_board(),
                'ships': [],  # placed ships with cells
                'ships_remaining': len(SHIPS_DEF),
                'shots_against': [],  # [{row, col, hit}]
                'sunk_ships': [],
            }
        room['last_activity'] = time.time()
        _br_set(room_id, room)
    return jsonify({'success': True})

@app.route('/battleship/api/rooms/<room_id>/place', methods=['POST'])
def battleship_place_ships(room_id):
    denied = _battleship_check_access()
    if denied:
        return denied
    username = session['username']
    data = request.get_json(force=True)
    ships = data.get('ships', [])
    with battleship_lock:
        room = _br_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if room['status'] != 'placing':
            return jsonify({'error': 'Fase de colocação terminada'}), 400
        if username in room.get('ships_placed', []):
            return jsonify({'error': 'Já colocaste os teus navios'}), 400
        # Validate number of ships
        if len(ships) != len(SHIPS_DEF):
            return jsonify({'error': 'Número incorreto de navios'}), 400
        # Get player board
        player_data = room['boards'][username]
        board = _deserialize_board(player_data['board'])
        placed_ships = []
        for i, expected in enumerate(SHIPS_DEF):
            expected_name = expected['name']
            expected_size = expected['size']
            ship = ships[i]
            ship_name = ship['name']
            ship_size = ship['size']
            # Normalize names for comparison (ignore accents/case)
            if _normalize_name(ship_name) != _normalize_name(expected_name) or ship_size != expected_size:
                return jsonify({'error': f'Navio inválido: {expected_name}'}), 400
            cells = ship['cells']
            # Validate placement
            if not _can_place_ship(board, cells):
                return jsonify({'error': f'Posição inválida para {expected_name}'}), 400
            # Place on board using expected_name (canonical name)
            _place_ship_on_board(board, cells, expected_name)
            placed_ships.append({
                'name': expected_name,
                'size': expected_size,
                'cells': cells
            })
        # Save board and ships (with canonical names)
        player_data['board'] = _serialize_board(board)
        player_data['ships'] = placed_ships
        room['boards'][username] = player_data
        room.setdefault('ships_placed', []).append(username)
        room['last_activity'] = time.time()
        _br_set(room_id, room)
    return jsonify({'success': True})

def _place_ship_on_board(board, cells, ship_name):
    """Place ship on board (board is a 2D list)"""
    for r, c in cells:
        board[r][c] = ship_name

@app.route('/battleship/api/rooms/<room_id>/shoot', methods=['POST'])
def battleship_shoot(room_id):
    denied = _battleship_check_access()
    if denied:
        return denied
    username = session['username']
    data = request.get_json(force=True)
    target = data.get('target')
    row = data.get('row')
    col = data.get('col')
    if target is None or row is None or col is None:
        return jsonify({'error': 'Dados incompletos'}), 400
    if row < 0 or row >= 10 or col < 0 or col >= 10:
        return jsonify({'error': 'Coordenadas fora do tabuleiro'}), 400
    with battleship_lock:
        room = _br_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        if room['status'] != 'playing':
            return jsonify({'error': 'Jogo não está em andamento'}), 400
        if room['current_turn'] != username:
            return jsonify({'error': 'Não é a tua vez'}), 403
        if target not in room['players']:
            return jsonify({'error': 'Jogador alvo não encontrado'}), 400
        if target == username:
            return jsonify({'error': 'Não podes atirar em ti mesmo'}), 400
        if target in room.get('eliminated', []):
            return jsonify({'error': 'Jogador já eliminado'}), 400
        # Check if already shot at this cell
        target_board_data = room['boards'][target]
        already_shot = any(sh['row'] == row and sh['col'] == col for sh in target_board_data['shots_against'])
        if already_shot:
            return jsonify({'error': 'Já disparaste nesta célula'}), 400
        # Process shot
        target_board = _deserialize_board(target_board_data['board'])
        hit = target_board[row][col] is not None
        hit_ship_name = target_board[row][col] if hit else None
        sunk = False
        # Record shot
        target_board_data['shots_against'].append({'row': row, 'col': col, 'hit': hit})
        if hit:
            # Mark as hit on target's board (set to None to indicate destroyed cell)
            target_board[row][col] = None
            # Check if the ship was sunk
            if _check_ship_sunk(target_board, hit_ship_name):
                sunk = True
                # Find the ship object to get its cells
                ship_obj = None
                for s in target_board_data['ships']:
                    if s['name'] == hit_ship_name:
                        ship_obj = s
                        break
                if ship_obj:
                    # Check if already sunk (by name)
                    already_sunk = any(
                        s['name'] == hit_ship_name if isinstance(s, dict) else s == hit_ship_name
                        for s in target_board_data['sunk_ships']
                    )
                    if not already_sunk:
                        target_board_data['sunk_ships'].append({
                            'name': hit_ship_name,
                            'cells': ship_obj['cells']
                        })
                # Update remaining ships count based on sunk_ships
                target_board_data['ships_remaining'] = len(SHIPS_DEF) - len(target_board_data['sunk_ships'])
                if target_board_data['ships_remaining'] == 0:
                    room.setdefault('eliminated', []).append(target)
                    # Check if game should end
                    active_players = [p for p in room['players'] if p not in room.get('eliminated', [])]
                    if len(active_players) <= 1:
                        room['status'] = 'finished'
                        if len(active_players) == 1:
                            room['winner'] = active_players[0]
                        else:
                            room['winner'] = None
        # Record shot history
        room['shot_history'].append({
            'shooter': username,
            'target': target,
            'row': row,
            'col': col,
            'hit': hit,
            'sunk': sunk,
            'timestamp': time.time(),
        })
        # Update target's board
        target_board_data['board'] = _serialize_board(target_board)
        room['boards'][target] = target_board_data
        # Advance turn only if target not eliminated? Actually in battleship, turn passes after each shot regardless
        if not hit:
            # Miss: turn passes to next player
            _advance_turn(room)
        # If hit but not sunk, same player shoots again (house rule: continue on hit)
        # (If sunk, also continue? Usually you continue, but we follow same: continue shooting)
        room['last_activity'] = time.time()
        _br_set(room_id, room)
    return jsonify({'success': True, 'hit': hit, 'sunk': sunk})

@app.route('/battleship/api/rooms/<room_id>/state', methods=['GET'])
def battleship_state(room_id):
    denied = _battleship_check_access()
    if denied:
        return denied
    username = session['username']
    with battleship_lock:
        room = _br_get(room_id)
        if not room:
            return jsonify({'error': 'Sala não encontrada'}), 404
        # Auto-start: transition from placing to playing when all current players have placed ships
        if room['status'] == 'placing':
            current_players = room['players']
            ships_placed = room.get('ships_placed', [])
            # Check if every current player has placed their ships
            if len(current_players) > 0 and all(p in ships_placed for p in current_players):
                room['status'] = 'playing'
                # Use player_order (randomized at start) for turn rotation
                order = room.get('player_order', current_players)
                room['players'] = order  # reorder players list to match random order
                # Determine first turn from player_order
                for p in room['players']:
                    if p not in room.get('eliminated', []):
                        room['current_turn'] = p
                        break
                _br_set(room_id, room)  # persist state change
        # Build state for frontend
        boards = {}
        for p, bd in room['boards'].items():
            # Only send necessary info for each board
            is_own = (p == username)
            boards[p] = {
                'board': bd['board'] if is_own else None,  # only send own ship positions to owner
                'ships': bd['ships'] if is_own else None,  # only send ships to owner
                'ships_remaining': bd.get('ships_remaining', len(SHIPS_DEF)),
                'shots_against': bd['shots_against'],
                'sunk_ships': bd.get('sunk_ships', []),
            }
        state = {
            'name': room['name'],
            'host': room['host'],
            'players': room['players'],
            'display_names': room['display_names'],
            'status': room['status'],
            'scores': room['scores'],
            'boards': boards,
            'ships_placed': room.get('ships_placed', []),
            'current_turn': room.get('current_turn'),
            'eliminated': room.get('eliminated', []),
            'shot_history': room.get('shot_history', [])[-50:],  # last 50 shots
            'my_username': username,
            'winner': room.get('winner'),
        }
        # Add next turn info if playing
        if room['status'] == 'playing':
            state['player_order'] = room.get('player_order', [])
    return jsonify(state)

@app.route('/battleship/api/rooms/<room_id>/set_placement', methods=['POST'])
def battleship_set_placement(room_id):
    """Alias for /place, kept for compatibility"""
    return battleship_place_ships(room_id)

@app.route('/battleship/api/rooms/<room_id>/begin_game', methods=['POST'])
def battleship_begin_game(room_id):
    """Alternative endpoint to start game"""
    return battleship_start_game(room_id)

@app.route('/battleship/api/rooms/<room_id>/fire', methods=['POST'])
def battleship_fire(room_id):
    """Alias for /shoot"""
    return battleship_shoot(room_id)

@app.route('/battleship/api/test_migration', methods=['GET'])
def battleship_test_migration():
    """Test endpoint to ensure battleship endpoints are registered"""
    return jsonify({'status': 'ok', 'message': 'Battleship API is active'})

try:
    _br_init()
except Exception:
    pass

@app.route('/analytics')
def analytics():
    """RMS Analytics for Tickets"""
    if 'name' not in session:
        return redirect(url_for('login'))
    return render_template('analytics.html')

@app.route('/api/analytics/production-lines', methods=['GET'])
def get_production_lines():
    """API endpoint to fetch available production lines"""
    if 'name' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = connect()
        cursor = conn.cursor()
        
        # Get unique production lines from automation_improvement table
        query1 = """
        SELECT DISTINCT prod_line FROM [DT_request].[dbo].[automation_improvement]
        WHERE prod_line IS NOT NULL AND is_deleted = 0
        """
        
        cursor.execute(query1)
        lines_set = set([row[0] for row in cursor.fetchall()])
        
        # Get unique production lines from automation_support table
        query2 = """
        SELECT DISTINCT prod_line FROM [DT_request].[dbo].[automation_support]
        WHERE prod_line IS NOT NULL AND is_deleted = 0
        """
        
        cursor.execute(query2)
        lines_set.update([row[0] for row in cursor.fetchall()])
        
        # Sort and convert to list
        lines = sorted(list(lines_set))
        
        cursor.close()
        conn.close()
        
        logging.info(f"Production lines found: {lines}")
        return jsonify({'production_lines': lines})
    
    except Exception as e:
        logging.error(f"Error in get_production_lines: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/automation-responsibles', methods=['GET'])
def get_automation_responsibles():
    """API endpoint to fetch automation responsibles"""
    if 'name' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        conn = connect()
        cursor = conn.cursor()
        
        # Get unique responsibles from automation_improvement table
        query1 = """
        SELECT DISTINCT responsible FROM [DT_request].[dbo].[automation_improvement]
        WHERE responsible IS NOT NULL AND responsible != '' AND is_deleted = 0
        """
        
        cursor.execute(query1)
        responsibles_set = set([row[0] for row in cursor.fetchall() if row[0]])
        
        # Get unique responsibles from automation_support table
        query2 = """
        SELECT DISTINCT responsible FROM [DT_request].[dbo].[automation_support]
        WHERE responsible IS NOT NULL AND responsible != '' AND is_deleted = 0
        """
        
        cursor.execute(query2)
        responsibles_set.update([row[0] for row in cursor.fetchall() if row[0]])
        
        # Sort and convert to list
        responsibles = sorted(list(responsibles_set))
        
        cursor.close()
        conn.close()
        
        logging.info(f"Automation responsibles found: {responsibles}")
        return jsonify({'responsibles': responsibles})
    
    except Exception as e:
        logging.error(f"Error in get_automation_responsibles: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/data', methods=['GET'])
def get_analytics_data():
    """API endpoint to fetch analytics data with filters"""
    if 'name' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        logging.info("===== get_analytics_data START =====")
        
        # Get filter parameters from query string
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        ticket_type = request.args.get('ticket_types', '')  # Single ticket type now
        category = request.args.get('category', '')  # New category filter
        prod_line = request.args.get('prod_line', '')  # New production line filter
        responsible = request.args.get('responsible', '')  # Responsible filter
        status_filter = request.args.get('status', '')
        
        logging.info(f"Filters - Category: {category}, Type: {ticket_type}, Dates: {start_date} to {end_date}")
        
        # Determine which tables to query based on category or ticket_type
        tables_to_query = []
        
        if ticket_type:
            # If specific ticket type is selected
            ticket_type_map = {
                'DTAI': 'automation_improvement',
                'DTAS': 'automation_support',
                'DTNA': 'new_application',
                'DTSI': 'software_issue'
            }
            if ticket_type in ticket_type_map:
                tables_to_query = [ticket_type_map[ticket_type]]
        elif category:
            # If category is selected
            if category == 'automation':
                tables_to_query = ['automation_improvement', 'automation_support']
            elif category == 'software':
                tables_to_query = ['new_application', 'software_issue']
        else:
            # Default to all tables
            tables_to_query = ['automation_improvement', 'automation_support', 'new_application', 'software_issue']
        
        conn = connect()
        cursor = conn.cursor()
        
        # Build base filter query
        date_filter = ""
        if start_date and end_date:
            date_filter = f"WHERE created_at >= '{start_date}' AND created_at <= '{end_date}'"
        elif start_date:
            date_filter = f"WHERE created_at >= '{start_date}'"
        elif end_date:
            date_filter = f"WHERE created_at <= '{end_date}'"
        
        # 1. Distribution by Type (only for selected tables)
        type_map = {
            'automation_improvement': 'DTAI',
            'automation_support': 'DTAS',
            'new_application': 'DTNA',
            'software_issue': 'DTSI'
        }
        
        # Build distribution query with proper date filtering
        distribution_union_parts = []
        for table in tables_to_query:
            type_code = type_map[table]
            if date_filter:
                # Include created_at for filtering in subquery
                query_part = f"SELECT '{type_code}' as type, COUNT(*) as count FROM {table} {date_filter}"
            else:
                query_part = f"SELECT '{type_code}' as type, COUNT(*) as count FROM {table}"
            distribution_union_parts.append(query_part)
        
        distribution_query = " UNION ALL ".join(distribution_union_parts)
        
        try:
            cursor.execute(distribution_query)
            distribution = [{'type': row[0], 'count': row[1]} for row in cursor.fetchall()]
        except Exception as e:
            logging.error(f"Distribution query error: {str(e)}")
            distribution = []
        
        # 2. KPIs - Total Tickets Created (Last 30 days) - based on selected tables
        selected_tables_select = " UNION ALL ".join([f"SELECT status, created_at FROM {table}" for table in tables_to_query])
        
        kpi_query = f"""
        SELECT 
            COUNT(*) as total_created,
            SUM(CASE WHEN status IN (2, -1) THEN 1 ELSE 0 END) as completed,
            SUM(CASE WHEN status IN (0,1,3,4,5,6,7) THEN 1 ELSE 0 END) as pending,
            SUM(CASE WHEN status = 1 THEN 1 ELSE 0 END) as in_progress
        FROM (
            {selected_tables_select}
        ) AS all_tickets
        WHERE created_at >= DATEADD(day, -30, CAST(GETDATE() AS DATE))
        """
        
        cursor.execute(kpi_query)
        kpi_row = cursor.fetchone()
        kpis = {
            'total_created': kpi_row[0] or 0,
            'completed': kpi_row[1] or 0,
            'pending': kpi_row[2] or 0,
            'in_progress': kpi_row[3] or 0,
            'completion_rate': round((kpi_row[1] / kpi_row[0] * 100) if kpi_row[0] > 0 else 0, 1)
        }
        
        # 3. By Status - All time
        status_names = {
            -1: 'Rejected', 
            0: 'Waiting', 
            1: 'In Progress', 
            2: 'Completed', 
            3: 'Under Analysis',
            4: 'Waiting DT', 
            5: 'Waiting Requester', 
            6: 'Waiting Line', 
            7: 'Waiting Maintenance',
            None: 'Pending Approval'  # Newly created, not approved/rejected yet
        }
        
        selected_tables_select = " UNION ALL ".join([f"SELECT status FROM {table}" for table in tables_to_query])
        
        status_query = f"""
        SELECT status, COUNT(*) as count
        FROM (
            {selected_tables_select}
        ) AS all_tickets
        GROUP BY status
        ORDER BY status
        """
        
        cursor.execute(status_query)
        by_status = []
        for row in cursor.fetchall():
            status_id = row[0]
            status_name = status_names.get(status_id, f'Status {status_id}')
            by_status.append({'status': status_id, 'status_name': status_name, 'count': row[1]})
        
        # 4. Average completion time per type - only for selected tables
        avg_time_parts = []
        for table in tables_to_query:
            type_code = type_map[table]
            avg_time_parts.append(f"""SELECT '{type_code}' as type, AVG(DATEDIFF(day, created_at, updated_at)) as avg_days
            FROM {table}
            WHERE status = 2""")
        
        avg_time_query = " UNION ALL ".join(avg_time_parts)
        
        cursor.execute(avg_time_query)
        avg_time = [{'type': row[0], 'avg_days': round(row[1], 1) if row[1] else 0} for row in cursor.fetchall()]
        
        # 4.5 Tickets per Responsible (Total count)
        try:
            responsible_union_parts = []
            for table in tables_to_query:
                responsible_union_parts.append(f"SELECT responsible FROM {table} WHERE responsible IS NOT NULL AND responsible != ''")
            
            if responsible_union_parts:
                responsible_select = " UNION ALL ".join(responsible_union_parts)
                
                tickets_per_responsible_query = f"""
                SELECT responsible, COUNT(*) as ticket_count
                FROM ({responsible_select}) AS all_responsible
                GROUP BY responsible
                ORDER BY ticket_count DESC
                """
                
                cursor.execute(tickets_per_responsible_query)
                tickets_per_responsible = [
                    {'responsible': row[0] or 'Unknown', 'ticket_count': row[1] or 0}
                    for row in cursor.fetchall()
                ]
            else:
                tickets_per_responsible = []
        except Exception as e:
            logging.error(f"Tickets per responsible query error: {str(e)}")
            tickets_per_responsible = []
        
        # 5. Completed Tickets Analytics
        # Total hours worked on completed tickets
        try:
            # For completed tickets, aggregate hours and responsible info
            completed_union_parts = []
            for table in tables_to_query:
                # Use CAST to handle different data types, ISNULL for missing columns
                completed_union_parts.append(f"""
                SELECT 
                    CAST(ISNULL(responsible, 'Unassigned') as VARCHAR(255)) as responsible,
                    CAST(ISNULL([time], 0) as FLOAT) as hours
                FROM {table} 
                WHERE status = 2
                """)
            
            completed_select = " UNION ALL ".join(completed_union_parts)
            
            # Total hours worked
            total_hours_query = f"""
            SELECT SUM(hours) as total_hours
            FROM ({completed_select}) AS completed_data
            """
            
            cursor.execute(total_hours_query)
            hours_row = cursor.fetchone()
            total_hours_worked = float(hours_row[0]) if hours_row[0] else 0
            
            # Hours per responsible (completed tickets only)
            hours_per_responsible_query = f"""
            SELECT responsible, SUM(hours) as total_hours, COUNT(*) as completed_count
            FROM ({completed_select}) AS completed_data
            WHERE responsible IS NOT NULL AND responsible != ''
            GROUP BY responsible
            ORDER BY total_hours DESC
            """
            
            cursor.execute(hours_per_responsible_query)
            hours_per_responsible = [
                {'responsible': row[0] or 'Unknown', 'total_hours': float(row[1]) if row[1] else 0, 'completed_count': int(row[2]) if row[2] else 0}
                for row in cursor.fetchall()
            ]
        except Exception as e:
            logging.error(f"Completed tickets query error: {str(e)}")
            total_hours_worked = 0
            hours_per_responsible = []
        
        # 6. Projects: Time Spent per Project
        time_per_project_query = f"""
        SELECT p.id as project_id, p.name as project_name, SUM(ISNULL(t.time_spent, 0)) as total_hours
        FROM [DT_request].[dbo].[projects] p
        LEFT JOIN [DT_request].[dbo].[tasks] t ON p.id = t.project_id
        {f"WHERE t.created_at >= '{start_date}' AND t.created_at <= '{end_date}'" if start_date and end_date else (f"WHERE t.created_at >= '{start_date}'" if start_date else (f"WHERE t.created_at <= '{end_date}'" if end_date else ""))}
        GROUP BY p.id, p.name
        ORDER BY total_hours DESC
        """

        cursor.execute(time_per_project_query)
        time_per_project = [
            {'project_id': row[0], 'project_name': row[1] or 'Unknown', 'total_hours': row[2] or 0}
            for row in cursor.fetchall()
        ]
        
        # 7. Tasks per Project
        tasks_per_project_query = f"""
        SELECT p.name as project_name, COUNT(t.id) as task_count
        FROM [DT_request].[dbo].[projects] p
        LEFT JOIN [DT_request].[dbo].[tasks] t ON p.id = t.project_id
        {f"WHERE t.created_at >= '{start_date}' AND t.created_at <= '{end_date}'" if start_date and end_date else (f"WHERE t.created_at >= '{start_date}'" if start_date else (f"WHERE t.created_at <= '{end_date}'" if end_date else ""))}
        GROUP BY p.id, p.name
        ORDER BY task_count DESC
        """
        
        cursor.execute(tasks_per_project_query)
        tasks_per_project = [
            {'project_name': row[0] or 'Unknown', 'task_count': row[1] or 0}
            for row in cursor.fetchall()
        ]
        
        # 8. Average Time per Task by Project
        avg_time_per_task_query = f"""
        SELECT p.name as project_name, AVG(ISNULL(t.time_spent, 0)) as avg_hours
        FROM [DT_request].[dbo].[projects] p
        LEFT JOIN [DT_request].[dbo].[tasks] t ON p.id = t.project_id
        {f"WHERE t.created_at >= '{start_date}' AND t.created_at <= '{end_date}'" if start_date and end_date else (f"WHERE t.created_at >= '{start_date}'" if start_date else (f"WHERE t.created_at <= '{end_date}'" if end_date else ""))}
        GROUP BY p.id, p.name
        ORDER BY avg_hours DESC
        """
        
        cursor.execute(avg_time_per_task_query)
        avg_time_per_task = [
            {'project_name': row[0] or 'Unknown', 'avg_hours': row[1] or 0}
            for row in cursor.fetchall()
        ]
        
        # 9. Tasks by Status (all projects)
        task_status_names = {
            0: 'Not Started', 
            1: 'In Progress', 
            2: 'Completed', 
            3: 'On Hold', 
            4: 'Cancelled'
        }
        
        tasks_status_query = f"""
        SELECT ISNULL(status, 0) as status, COUNT(*) as count
        FROM [DT_request].[dbo].[tasks]
        {f"WHERE created_at >= '{start_date}' AND created_at <= '{end_date}'" if start_date and end_date else (f"WHERE created_at >= '{start_date}'" if start_date else (f"WHERE created_at <= '{end_date}'" if end_date else ""))}
        GROUP BY status
        ORDER BY status
        """
        
        cursor.execute(tasks_status_query)
        tasks_status = []
        for row in cursor.fetchall():
            status_id = row[0]
            status_name = task_status_names.get(status_id, f'Status {status_id}')
            tasks_status.append({'status': status_id, 'status_name': status_name, 'count': row[1]})
        
        # 10. Tasks by Type
        task_type_names = {
            'planned': 'Planned',
            'unplanned': 'Unplanned',
            'project': 'Project',
            'ticket': 'From Ticket'
        }
        
        tasks_type_query = f"""
        SELECT 
            CASE 
                WHEN project_id IS NOT NULL THEN 'project'
                WHEN ticket_table IS NOT NULL THEN 'ticket'
                WHEN ISNULL(task_type, '') = 'unplanned' THEN 'unplanned'
                ELSE 'planned'
            END as task_type,
            COUNT(*) as count
        FROM [DT_request].[dbo].[tasks]
        {f"WHERE created_at >= '{start_date}' AND created_at <= '{end_date}'" if start_date and end_date else (f"WHERE created_at >= '{start_date}'" if start_date else (f"WHERE created_at <= '{end_date}'" if end_date else ""))}
        GROUP BY 
            CASE 
                WHEN project_id IS NOT NULL THEN 'project'
                WHEN ticket_table IS NOT NULL THEN 'ticket'
                WHEN ISNULL(task_type, '') = 'unplanned' THEN 'unplanned'
                ELSE 'planned'
            END
        ORDER BY count DESC
        """
        
        cursor.execute(tasks_type_query)
        tasks_type = []
        for row in cursor.fetchall():
            type_key = row[0]
            type_name = task_type_names.get(type_key, type_key if type_key else 'Planned')
            tasks_type.append({'task_type': type_key, 'task_type_name': type_name, 'count': row[1]})
        
        # 11. Production Line Analytics (Automation Support + Automation Improvement)
        # Build date filter for production line queries
        line_date_filter = ""
        line_where_parts = []
        
        if start_date and end_date:
            line_where_parts.append(f"created_at >= '{start_date}' AND created_at <= '{end_date}'")
        elif start_date:
            line_where_parts.append(f"created_at >= '{start_date}'")
        elif end_date:
            line_where_parts.append(f"created_at <= '{end_date}'")
        
        if prod_line:
            line_where_parts.append(f"prod_line = '{prod_line}'")
        
        if responsible:
            line_where_parts.append(f"responsible = '{responsible}'")
        
        if line_where_parts:
            line_date_filter = "WHERE " + " AND ".join(line_where_parts)
        
        # Determine which line tables to include based on filters
        line_tables = []
        if ticket_type:
            # If specific ticket type is selected
            if ticket_type == 'DTAI':
                line_tables = ['[DT_request].[dbo].[automation_improvement]']
            elif ticket_type == 'DTAS':
                line_tables = ['[DT_request].[dbo].[automation_support]']
            elif ticket_type in ['DTNA', 'DTSI']:
                # Software categories don't have prod_line data, show empty
                line_tables = []
        elif category == 'automation':
            line_tables = ['[DT_request].[dbo].[automation_support]', '[DT_request].[dbo].[automation_improvement]']
        elif category == 'software':
            # Software category doesn't have prod_line, so show no data
            line_tables = []
        else:
            # No filter - include automation tables (software tables don't have prod_line)
            line_tables = ['[DT_request].[dbo].[automation_support]', '[DT_request].[dbo].[automation_improvement]']
        
        # Daily Savings by Production Line
        daily_savings = []
        daily_tickets = []
        daily_hours = []
        
        if line_tables:
            # Build individual queries for each table
            union_parts = []
            for table in line_tables:
                where_parts = []
                
                if start_date and end_date:
                    where_parts.append(f"CAST(updated_at AS DATE) >= '{start_date}' AND CAST(updated_at AS DATE) <= '{end_date}'")
                elif start_date:
                    where_parts.append(f"CAST(updated_at AS DATE) >= '{start_date}'")
                elif end_date:
                    where_parts.append(f"CAST(updated_at AS DATE) <= '{end_date}'")
                
                if prod_line:
                    where_parts.append(f"prod_line = '{prod_line}'")
                
                if responsible:
                    where_parts.append(f"responsible = '{responsible}'")
                
                where_clause = ""
                if where_parts:
                    where_clause = "WHERE " + " AND ".join(where_parts)
                
                # Use hard_savings if available, otherwise calculate from [time] * 18.75
                query_part = f"""SELECT 
                    CAST(updated_at AS DATE) as date, 
                    ISNULL(prod_line, 'Unknown') as prod_line,
                    CAST(ISNULL(hard_savings, ISNULL([time], 0) * 18.75) AS FLOAT) as savings
                FROM {table} {where_clause}"""
                union_parts.append(query_part)
            
            # Daily Savings Query - now using actual savings values
            daily_savings_query = f"""
            SELECT 
                date,
                prod_line,
                SUM(savings) as total_savings
            FROM (
                {' UNION ALL '.join(union_parts)}
            ) as all_data
            GROUP BY date, prod_line
            ORDER BY date ASC
            """
            
            try:
                cursor.execute(daily_savings_query)
                rows = cursor.fetchall()
                logging.info(f"Daily savings query returned {len(rows)} rows")
                
                for row in rows:
                    date = str(row[0]) if row[0] else 'Unknown'
                    prod_line = row[1]
                    total_savings = float(row[2]) if row[2] else 0
                    
                    daily_savings.append({
                        'date': date,
                        'prod_line': prod_line,
                        'savings': round(total_savings, 2)
                    })
                    
                logging.info(f"Daily savings data: {daily_savings}")
            except Exception as e:
                logging.error(f"Error in daily_savings_query: {str(e)}")
                logging.error(f"Query was: {daily_savings_query}")
                daily_savings = []
            
            # Daily Tickets Query - using old logic
            daily_tickets_union_parts = []
            for table in line_tables:
                daily_tickets_union_parts.append(f"SELECT created_at, prod_line FROM {table}")

            
            daily_tickets_query = f"""
            SELECT CAST(created_at AS DATE) as date, ISNULL(prod_line, 'Unknown') as prod_line, COUNT(*) as ticket_count
            FROM (
                {' UNION ALL '.join(daily_tickets_union_parts)}
            ) as all_tickets
            {("WHERE created_at >= '" + start_date + "' AND created_at <= '" + end_date + "'") if start_date and end_date else (("WHERE created_at >= '" + start_date + "'") if start_date else (("WHERE created_at <= '" + end_date + "'") if end_date else ""))}
            GROUP BY CAST(created_at AS DATE), prod_line
            ORDER BY date ASC
            """
            
            logging.info(f"Daily Tickets Query: {daily_tickets_query}")
            
            try:
                cursor.execute(daily_tickets_query)
                rows = cursor.fetchall()
                logging.info(f"Daily Tickets Query returned {len(rows)} rows")
                for row in rows:
                    daily_tickets.append({
                        'date': str(row[0]),
                        'prod_line': row[1],
                        'ticket_count': row[2]
                    })
                    logging.debug(f"Row: date={row[0]}, prod_line={row[1]}, count={row[2]}")
            except Exception as e:
                logging.error(f"Error in daily_tickets_query: {str(e)}")
                daily_tickets = []
            
            # Daily Hours Query - using old logic with updated_at
            daily_hours_union_parts = []
            for table in line_tables:
                daily_hours_union_parts.append(f"SELECT updated_at, prod_line, [time] FROM {table}")

            daily_hours_query = f"""
            SELECT CAST(updated_at AS DATE) as date, ISNULL(prod_line, 'Unknown') as prod_line, SUM(ISNULL([time], 0)) as total_hours
            FROM (
                {' UNION ALL '.join(daily_hours_union_parts)}
            ) as all_data
            {("WHERE updated_at >= '" + start_date + "' AND updated_at <= '" + end_date + "'") if start_date and end_date else (("WHERE updated_at >= '" + start_date + "'") if start_date else (("WHERE updated_at <= '" + end_date + "'") if end_date else ""))}
            GROUP BY CAST(updated_at AS DATE), prod_line
            ORDER BY date ASC
            """
            
            logging.info(f"Daily Hours Query: {daily_hours_query}")
            
            try:
                cursor.execute(daily_hours_query)
                rows = cursor.fetchall()
                logging.info(f"Daily Hours Query returned {len(rows)} rows")
                for row in rows:
                    daily_hours.append({
                        'date': str(row[0]),
                        'prod_line': row[1],
                        'total_hours': float(row[2]) if row[2] else 0
                    })
                    logging.debug(f"Row: date={row[0]}, prod_line={row[1]}, hours={row[2]}")
            except Exception as e:
                logging.error(f"Error in daily_hours_query: {str(e)}")
                daily_hours = []
        
        
        # Tickets by production line
        if line_tables:
            union_queries = [f"SELECT prod_line FROM {table} {line_date_filter}" for table in line_tables]
            line_tickets_query = f"""
            SELECT ISNULL(prod_line, 'Unknown') as prod_line, COUNT(*) as ticket_count
            FROM (
                {' UNION ALL '.join(union_queries)}
            ) as all_tickets
            GROUP BY prod_line
            ORDER BY ticket_count DESC
            """
            
            cursor.execute(line_tickets_query)
            tickets_by_line = [
                {'prod_line': row[0], 'ticket_count': row[1]}
                for row in cursor.fetchall()
            ]
        else:
            tickets_by_line = []
        
        # Time spent by production line
        if line_tables:
            union_queries = [f"SELECT prod_line, [time] FROM {table} {line_date_filter}" for table in line_tables]
            line_time_query = f"""
            SELECT ISNULL(prod_line, 'Unknown') as prod_line, SUM(ISNULL([time], 0)) as total_hours
            FROM (
                {' UNION ALL '.join(union_queries)}
            ) as all_tickets
            GROUP BY prod_line
            ORDER BY total_hours DESC
            """
            
            cursor.execute(line_time_query)
            time_by_line = [
                {'prod_line': row[0], 'total_hours': row[1] or 0}
                for row in cursor.fetchall()
            ]
        else:
            time_by_line = []
        
        # Savings by production line (8 hours = 150€)
        if line_tables:
            union_queries = [f"SELECT prod_line, [time], hard_savings FROM {table} {line_date_filter}" for table in line_tables]
            line_savings_query = f"""
            SELECT ISNULL(prod_line, 'Unknown') as prod_line, 
                   SUM(ISNULL([time], 0)) as total_hours,
                   COUNT(*) as ticket_count,
                   SUM(ISNULL(hard_savings, 0)) as existing_savings
            FROM (
                {' UNION ALL '.join(union_queries)}
            ) as all_tickets
            GROUP BY prod_line
            ORDER BY total_hours DESC
            """
            
            cursor.execute(line_savings_query)
            savings_by_line = []
            avg_savings_by_line = []
            
            for row in cursor.fetchall():
                prod_line = row[0]
                total_hours = row[1] or 0
                ticket_count = row[2] or 1
                existing_savings = row[3] or 0
                
                # Calculate savings: existing_savings if > 0, else 8h = 150€
                calculated_savings = existing_savings if existing_savings > 0 else (total_hours / 8) * 150
                avg_savings = calculated_savings / ticket_count if ticket_count > 0 else 0
                
                savings_by_line.append({
                    'prod_line': prod_line,
                    'total_savings': calculated_savings
                })
                
                avg_savings_by_line.append({
                    'prod_line': prod_line,
                    'avg_savings': avg_savings
                })
        else:
            savings_by_line = []
            avg_savings_by_line = []
        
        cursor.close()
        conn.close()
        
        # Get status evolution data (for automation)
        status_evolution = []
        if category == 'automation' or not category:
            try:
                conn = connect()
                cursor = conn.cursor()
                
                evolution_query = """
                SELECT CAST(created_at AS DATE) as date, 
                       CASE WHEN status = 2 THEN 'completed'
                            WHEN status = 1 THEN 'in_progress'
                            ELSE 'pending' 
                       END as status,
                       COUNT(*) as count
                FROM (
                    SELECT created_at, status, responsible FROM automation_improvement 
                    WHERE is_deleted = 0
                    UNION ALL
                    SELECT created_at, status, responsible FROM automation_support
                    WHERE is_deleted = 0
                ) as all_tickets
                """
                
                filters = []
                if responsible:
                    filters.append(f"responsible = '{responsible}'")
                if start_date and end_date:
                    filters.append(f"created_at >= '{start_date}' AND created_at <= '{end_date}'")
                elif start_date:
                    filters.append(f"created_at >= '{start_date}'")
                elif end_date:
                    filters.append(f"created_at <= '{end_date}'")
                
                if filters:
                    evolution_query += " WHERE " + " AND ".join(filters)
                
                evolution_query += " GROUP BY CAST(created_at AS DATE), CASE WHEN status = 2 THEN 'completed' WHEN status = 1 THEN 'in_progress' ELSE 'pending' END"
                evolution_query += " ORDER BY date ASC"
                
                cursor.execute(evolution_query)
                for row in cursor.fetchall():
                    status_evolution.append({
                        'date': str(row[0]),
                        'status': row[1],
                        'count': row[2]
                    })
                
                cursor.close()
                conn.close()
            except Exception as e:
                logging.error(f"Error in status_evolution_query: {str(e)}")
                status_evolution = []
        
        # Get software analytics data
        software_analytics = {}
        if category == 'software' or not category:
            try:
                conn = connect()
                cursor = conn.cursor()
                
                # Tickets by status - Software and Automation category users
                tickets_by_status = []
                
                # Build date filter for tickets
                ticket_date_filter = ""
                if start_date and end_date:
                    ticket_date_filter = f"AND created_at >= '{start_date}' AND created_at <= '{end_date}'"
                elif start_date:
                    ticket_date_filter = f"AND created_at >= '{start_date}'"
                elif end_date:
                    ticket_date_filter = f"AND created_at <= '{end_date}'"
                
                query = f"""
                SELECT u.[name] as responsible, 
                       CASE WHEN status = 2 THEN 'Completed'
                            WHEN status = 1 THEN 'In Progress'
                            WHEN status = 0 THEN 'Waiting'
                            WHEN status = -1 THEN 'Rejected'
                            ELSE 'Unknown'
                       END as status_name,
                       COUNT(*) as count
                FROM (
                    SELECT responsible, status, created_at FROM software_issue WHERE is_deleted = 0 {ticket_date_filter}
                    UNION ALL
                    SELECT responsible, status, created_at FROM new_application WHERE is_deleted = 0 {ticket_date_filter}
                ) as all_tickets
                INNER JOIN users u ON all_tickets.responsible = u.username
                WHERE u.category = 'Software'
                GROUP BY u.[name], status
                ORDER BY u.[name], status
                """
                
                logging.info(f"Software tickets by status query: {query}")
                cursor.execute(query)
                rows = cursor.fetchall()
                logging.info(f"Software tickets by status result: {rows}")
                
                for row in rows:
                    tickets_by_status.append({
                        'responsible': row[0],
                        'status': row[1],
                        'count': row[2]
                    })
                
                # Projects with task breakdown - Software category users with their tasks (including in Automation projects)
                # Task status mapping: 'To do'=Not Started, 'In Progress'=In Progress, 'Done'=Completed, 'On Hold'=On Hold, 'Cancelled'=Cancelled
                projects_with_tasks = []
                
                # Build date filter for tasks
                task_date_filter = ""
                if start_date and end_date:
                    task_date_filter = f"AND t.created_at >= '{start_date}' AND t.created_at <= '{end_date}'"
                elif start_date:
                    task_date_filter = f"AND t.created_at >= '{start_date}'"
                elif end_date:
                    task_date_filter = f"AND t.created_at <= '{end_date}'"
                
                query = f"""
                SELECT u.[name] as created_by,
                       SUM(CASE WHEN LOWER(TRIM(t.status)) = 'done' THEN 1 ELSE 0 END) as completed_tasks,
                       SUM(CASE WHEN LOWER(TRIM(t.status)) = 'in progress' THEN 1 ELSE 0 END) as in_progress_tasks,
                       SUM(CASE WHEN LOWER(TRIM(t.status)) = 'to do' THEN 1 ELSE 0 END) as not_started_tasks,
                       SUM(CASE WHEN LOWER(TRIM(t.status)) = 'on hold' THEN 1 ELSE 0 END) as on_hold_tasks,
                       SUM(CASE WHEN LOWER(TRIM(t.status)) = 'cancelled' THEN 1 ELSE 0 END) as cancelled_tasks,
                       SUM(CASE WHEN t.planned_end_date < CAST(GETDATE() AS DATE) AND LOWER(TRIM(t.status)) IN ('to do', 'in progress', 'on hold') THEN 1 ELSE 0 END) as overdue_tasks,
                       COUNT(t.id) as total_tasks
                FROM users u
                LEFT JOIN tasks t ON t.responsible = u.username AND t.project_id IS NOT NULL {task_date_filter}
                LEFT JOIN projects p ON t.project_id = p.id
                WHERE u.category = 'Software'
                  AND (p.category IN ('Software', 'Automation') OR t.id IS NULL)
                GROUP BY u.[name]
                ORDER BY u.[name]
                """
                
                logging.info(f"Projects with tasks query executed")
                cursor.execute(query)
                rows = cursor.fetchall()
                logging.info(f"Projects with tasks result: {rows}")
                
                for row in rows:
                    projects_with_tasks.append({
                        'created_by': row[0],
                        'completed_tasks': row[1] or 0,
                        'in_progress_tasks': row[2] or 0,
                        'not_started_tasks': row[3] or 0,
                        'on_hold_tasks': row[4] or 0,
                        'cancelled_tasks': row[5] or 0,
                        'overdue_tasks': row[6] or 0,
                        'total_tasks': row[7] or 0
                    })
                
                # Planned vs Unplanned tasks by user - Software and Automation category users
                # Unplanned: task_type = 'unplanned'
                # Planned: task_type IS NULL OR ticket_table IS NULL OR project_id IS NULL
                planned_tasks_by_user = []
                
                # Build date filter for tasks
                task_date_filter = ""
                if start_date and end_date:
                    task_date_filter = f"AND t.created_at >= '{start_date}' AND t.created_at <= '{end_date}'"
                elif start_date:
                    task_date_filter = f"AND t.created_at >= '{start_date}'"
                elif end_date:
                    task_date_filter = f"AND t.created_at <= '{end_date}'"
                
                query = f"""
                SELECT u.[name] as responsible,
                       CASE WHEN t.task_type = 'unplanned' THEN 'Unplanned'
                            ELSE 'Planned'
                       END as task_classification,
                       COUNT(*) as count
                FROM tasks t
                INNER JOIN users u ON t.responsible = u.username
                WHERE t.project_id IS NULL
                  AND u.category = 'Software'
                  AND t.week_number IS NOT NULL AND t.week_number != ''
                  AND t.planned_end_date IS NOT NULL
                  {task_date_filter}
                GROUP BY u.[name], CASE WHEN t.task_type = 'unplanned' THEN 'Unplanned' ELSE 'Planned' END
                ORDER BY responsible
                """
                
                logging.info(f"Planned tasks query executed")
                cursor.execute(query)
                rows = cursor.fetchall()
                logging.info(f"Planned tasks result: {rows}")
                
                for row in rows:
                    planned_tasks_by_user.append({
                        'responsible': row[0],
                        'classification': row[1],
                        'count': row[2]
                    })
                
                software_analytics = {
                    'tickets_by_status': tickets_by_status,
                    'projects_with_tasks': projects_with_tasks,
                    'planned_tasks_by_user': planned_tasks_by_user
                }
                
                logging.info(f"Software analytics complete: {len(tickets_by_status)} ticket statuses, {len(projects_with_tasks)} projects, {len(planned_tasks_by_user)} planned tasks")
                
                cursor.close()
                conn.close()
            except Exception as e:
                logging.error(f"Error in software_analytics_query: {str(e)}")
                import traceback
                logging.error(traceback.format_exc())
                software_analytics = {}
        
        logging.info(f"===== Returning analytics data with software_analytics: {bool(software_analytics)} =====")
        
        return jsonify({
            'success': True,
            'line_analytics': {
                'daily_savings': daily_savings,
                'daily_tickets': daily_tickets,
                'daily_hours': daily_hours,
                'status_evolution': status_evolution,
                'tickets_by_line': tickets_by_line,
                'time_by_line': time_by_line,
                'savings_by_line': savings_by_line,
                'avg_savings_by_line': avg_savings_by_line
            },
            'projects': {
                'time_per_project': time_per_project,
                'tasks_per_project': tasks_per_project,
                'avg_time_per_task': avg_time_per_task,
                'tasks_status': tasks_status,
                'tasks_type': tasks_type
            },
            'software_analytics': software_analytics
        })
    
    except Exception as e:
        logging.error(f"Error in get_analytics_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/software-tickets/<username>', methods=['GET'])
def get_software_tickets_for_user(username):
    """Get individual software tickets for a user that are not completed"""
    if 'name' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get date filters from query parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        conn = connect()
        cursor = conn.cursor()
        
        # Build date filter
        date_filter = ""
        if start_date and end_date:
            date_filter = f"AND created_at >= '{start_date}' AND created_at <= '{end_date}'"
        elif start_date:
            date_filter = f"AND created_at >= '{start_date}'"
        elif end_date:
            date_filter = f"AND created_at <= '{end_date}'"
        
        # Get tickets from software_issue and new_application where responsible matches and status is not Completed
        query = f"""
        SELECT 'DTSI' as type, 
               id as ticket_id,
               internal_code,
               title,
               CASE WHEN status = 2 THEN 'Completed'
                    WHEN status = 1 THEN 'In Progress'
                    WHEN status = 0 THEN 'Waiting'
                    WHEN status = -1 THEN 'Rejected'
                    WHEN status = 3 THEN 'Under Analysis'
                    WHEN status = 4 THEN 'Waiting DT'
                    WHEN status = 5 THEN 'Waiting Requester'
                    WHEN status = 6 THEN 'Waiting Line'
                    WHEN status = 7 THEN 'Waiting Maintenance'
                    ELSE 'Unknown'
               END as status,
               priority,
               created_at,
               updated_at
        FROM software_issue
        WHERE responsible = (SELECT username FROM users WHERE [name] = ?)
          AND is_deleted = 0
          AND status != 2
          {date_filter}
        UNION ALL
        SELECT 'DTNA' as type,
               id as ticket_id,
               internal_code,
               title,
               CASE WHEN status = 2 THEN 'Completed'
                    WHEN status = 1 THEN 'In Progress'
                    WHEN status = 0 THEN 'Waiting'
                    WHEN status = -1 THEN 'Rejected'
                    WHEN status = 3 THEN 'Under Analysis'
                    WHEN status = 4 THEN 'Waiting DT'
                    WHEN status = 5 THEN 'Waiting Requester'
                    WHEN status = 6 THEN 'Waiting Line'
                    WHEN status = 7 THEN 'Waiting Maintenance'
                    ELSE 'Unknown'
               END as status,
               priority,
               created_at,
               updated_at
        FROM new_application
        WHERE responsible = (SELECT username FROM users WHERE [name] = ?)
          AND is_deleted = 0
          AND status != 2
          {date_filter}
        ORDER BY created_at DESC
        """
        
        cursor.execute(query, (username, username))
        rows = cursor.fetchall()
        
        tickets = []
        for row in rows:
            tickets.append({
                'type': row[0],
                'ticket_id': row[1],
                'internal_code': row[2],
                'title': row[3],
                'status': row[4],
                'priority': row[5],
                'created_at': row[6],
                'updated_at': row[7]
            })
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'user': username,
            'tickets': tickets,
            'count': len(tickets)
        })
    
    except Exception as e:
        logging.error(f"Error in get_software_tickets_for_user: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/user-tickets/<username>', methods=['GET'])
def get_user_tickets(username):
    """Get individual tasks for a user that are not completed, from projects they created"""
    if 'name' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get date filters from query parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        conn = connect()
        cursor = conn.cursor()
        
        # Build date filter
        date_filter = ""
        if start_date and end_date:
            date_filter = f"AND t.created_at >= '{start_date}' AND t.created_at <= '{end_date}'"
        elif start_date:
            date_filter = f"AND t.created_at >= '{start_date}'"
        elif end_date:
            date_filter = f"AND t.created_at <= '{end_date}'"
        
        # Get tasks from projects where the user is the creator and task status is not 'done'
        query = f"""
        SELECT t.id as task_id,
               t.title,
               t.description,
               t.status,
               t.priority,
               t.created_at,
               t.updated_at,
               p.name as project_name,
               p.id as project_id
        FROM tasks t
        INNER JOIN projects p ON t.project_id = p.id
        INNER JOIN users u ON p.created_by = u.username
        WHERE u.[name] = ?
          AND LOWER(TRIM(t.status)) NOT IN ('done', 'standby')
          AND t.project_id IS NOT NULL
          {date_filter}
        ORDER BY t.created_at DESC
        """
        
        cursor.execute(query, (username,))
        rows = cursor.fetchall()
        
        tasks = []
        for row in rows:
            tasks.append({
                'task_id': row[0],
                'title': row[1],
                'description': row[2],
                'status': row[3],
                'priority': row[4],
                'created_at': row[5],
                'updated_at': row[6],
                'project_name': row[7],
                'project_id': row[8]
            })
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'user': username,
            'tickets': tasks,  # Keep 'tickets' key for frontend compatibility
            'count': len(tasks)
        })
    
    except Exception as e:
        logging.error(f"Error in get_user_tickets: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/planned-tasks', methods=['GET'])
def get_planned_tasks():
    """Get planned and unplanned tasks for a Software category user that are not completed"""
    if 'name' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        username = request.args.get('username', '')
        if not username:
            return jsonify({'error': 'Username is required'}), 400
        # Get date filters from query parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        conn = connect()
        cursor = conn.cursor()
        
        # Build date filter for tasks
        date_filter = ""
        if start_date and end_date:
            date_filter = f"AND t.created_at >= '{start_date}' AND t.created_at <= '{end_date}'"
        elif start_date:
            date_filter = f"AND t.created_at >= '{start_date}'"
        elif end_date:
            date_filter = f"AND t.created_at <= '{end_date}'"
        
        # Get planned and unplanned tasks for the user (exclude project tasks and tasks without week/due date)
        query = f"""
        SELECT t.id as task_id,
               t.title,
               t.description,
               t.status,
               t.priority,
               t.created_at,
               t.planned_end_date,
               CASE WHEN t.task_type = 'unplanned' THEN 'Unplanned'
                    ELSE 'Planned'
               END as task_classification,
               t.week_number,
               ISNULL(CAST(t.estimated_hours AS FLOAT), 0) as estimated_hours
        FROM tasks t
        INNER JOIN users u ON t.responsible = u.username
        WHERE u.[name] = ?
          AND u.category = 'Software'
          AND t.project_id IS NULL
          AND t.week_number IS NOT NULL AND t.week_number != ''
          AND t.planned_end_date IS NOT NULL
          {date_filter}
        ORDER BY
            CASE LOWER(TRIM(t.status))
                WHEN 'in progress' THEN 1
                WHEN 'to do' THEN 2
                WHEN 'done' THEN 3
                ELSE 4
            END,
            t.created_at DESC
        """

        cursor.execute(query, (username,))
        rows = cursor.fetchall()

        tasks = []
        for row in rows:
            tasks.append({
                'task_id': row[0],
                'title': row[1],
                'description': row[2],
                'status': row[3],
                'priority': row[4],
                'created_at': row[5].strftime('%Y-%m-%d') if row[5] else None,
                'planned_end_date': row[6].strftime('%Y-%m-%d') if row[6] else None,
                'task_classification': row[7],
                'week_number': row[8],
                'estimated_hours': row[9] or 0
            })
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'user': username,
            'tasks': tasks,
            'count': len(tasks)
        })
    
    except Exception as e:
        logging.error(f"Error in get_planned_tasks: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/user-pending-tasks', methods=['GET'])
def get_user_pending_tasks():
    """Get pending tasks (Not Started and Overdue) for a user from projects"""
    if 'name' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        username = request.args.get('username', '')
        if not username:
            return jsonify({'error': 'Username is required'}), 400
        # Get date filters from query parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        conn = connect()
        cursor = conn.cursor()
        
        # Build date filter
        date_filter = ""
        if start_date and end_date:
            date_filter = f"AND t.created_at >= '{start_date}' AND t.created_at <= '{end_date}'"
        elif start_date:
            date_filter = f"AND t.created_at >= '{start_date}'"
        elif end_date:
            date_filter = f"AND t.created_at <= '{end_date}'"
        
        # Get all tasks for this user's projects
        query = f"""
        SELECT t.id as task_id,
               t.title,
               t.description,
               t.status,
               t.priority,
               t.created_at,
               t.planned_end_date,
               p.name as project_name,
               p.id as project_id,
               CASE
                   WHEN LOWER(TRIM(t.status)) = 'done' THEN 'Completed'
                   WHEN LOWER(TRIM(t.status)) = 'to do' AND t.planned_end_date < CAST(GETDATE() AS DATE) THEN 'Overdue'
                   WHEN LOWER(TRIM(t.status)) = 'to do' THEN 'Not Started'
                   WHEN LOWER(TRIM(t.status)) = 'in progress' AND t.planned_end_date < CAST(GETDATE() AS DATE) THEN 'Overdue'
                   WHEN LOWER(TRIM(t.status)) = 'in progress' THEN 'In Progress'
                   ELSE 'Pending'
               END as task_category,
               t.week_number,
               ISNULL(CAST(t.estimated_hours AS FLOAT), 0) as estimated_hours
        FROM tasks t
        INNER JOIN projects p ON t.project_id = p.id
        WHERE p.created_by = (SELECT username FROM users WHERE [name] = ?)
          AND t.project_id IS NOT NULL
          {date_filter}
        ORDER BY
            CASE LOWER(TRIM(t.status))
                WHEN 'in progress' THEN 1
                WHEN 'to do' THEN 2
                WHEN 'done' THEN 3
                ELSE 4
            END,
            t.created_at DESC
        """

        cursor.execute(query, (username,))
        rows = cursor.fetchall()

        tasks = []
        not_started = 0
        overdue = 0
        completed = 0
        in_progress = 0

        for row in rows:
            category = row[9]
            if category == 'Completed':
                completed += 1
            elif category == 'Overdue':
                overdue += 1
            elif category == 'In Progress':
                in_progress += 1
            else:
                not_started += 1

            tasks.append({
                'task_id': row[0],
                'title': row[1],
                'description': row[2],
                'status': row[3],
                'priority': row[4],
                'created_at': row[5].strftime('%Y-%m-%d') if row[5] else None,
                'planned_end_date': row[6].strftime('%Y-%m-%d') if row[6] else None,
                'project_name': row[7],
                'project_id': row[8],
                'category': category,
                'week_number': row[10],
                'estimated_hours': row[11] or 0
            })

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'user': username,
            'tasks': tasks,
            'summary': {
                'not_started': not_started,
                'overdue': overdue,
                'completed': completed,
                'in_progress': in_progress,
                'total': len(tasks)
            }
        })
    
    except Exception as e:
        logging.error(f"Error in get_user_pending_tasks: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/project-tasks/<int:project_id>', methods=['GET'])
def get_project_tasks(project_id):
    """Get all tasks for a given project"""
    if 'name' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        conn = connect()
        cursor = conn.cursor()

        query = """
        SELECT t.id, t.title, t.description, t.status, t.priority,
               t.responsible, ISNULL(CAST(t.estimated_hours AS FLOAT), 0) as estimated_hours, t.week_number,
               t.planned_start_date, t.planned_end_date,
               p.name as project_name
        FROM tasks t
        INNER JOIN projects p ON t.project_id = p.id
        WHERE t.project_id = ?
        ORDER BY
            CASE LOWER(TRIM(t.status))
                WHEN 'in progress' THEN 1
                WHEN 'to do' THEN 2
                WHEN 'done' THEN 3
                ELSE 4
            END,
            t.planned_end_date
        """

        cursor.execute(query, (project_id,))
        rows = cursor.fetchall()

        tasks = []
        for row in rows:
            task_data = {
                'task_id': row[0],
                'title': row[1],
                'description': row[2],
                'status': row[3],
                'priority': row[4],
                'responsible': row[5],
                'estimated_hours': row[6] or 0,
                'week_number': row[7],
                'planned_start_date': row[8].strftime('%Y-%m-%d') if row[8] else None,
                'planned_end_date': row[9].strftime('%Y-%m-%d') if row[9] else None,
                'project_name': row[10]
            }
            # Parse responsible to array for multi-select support
            task_data['responsible'] = parse_responsible(task_data['responsible'])
            tasks.append(task_data)

        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'project_id': project_id,
            'tasks': tasks,
            'total': len(tasks)
        })

    except Exception as e:
        logging.error(f"Error in get_project_tasks: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)